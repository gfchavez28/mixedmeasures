"""Excel export endpoints — split from export.py for maintainability."""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import and_, func, or_
import io
import json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_db
from ..models.user import User
from ..models.project import Project
from ..models.conversation import Conversation
from ..models.document import Document
from ..models.observation import Observation
from ..models.segment import Segment
from ..models.code import Code
from ..models.code_application import CodeApplication
from ..models.audit import AuditEntry
from ..models.speaker import Speaker
from ..models.note import Note
from ..models.memo import Memo
from ..models.dataset import Dataset, DatasetColumn, DatasetRow, DatasetValue, ColumnType
from ..models.metric import MetricDefinition
from ..models.excerpt import Excerpt, segment_has_any_quote_filter
from ..models.analysis_domain import AnalysisDomain, AnalysisDomainMember
from ..models.equivalence_group import EquivalenceGroup
from ..services.grouping import MISSING_GROUP_LABEL
from ..services.recode import compute_value
from ..services.missing_values import (
    describe_missing_rules,
    is_missing,
    parse_missing_rules,
)
from ..services.metrics import resolve_input_source_labels
from ..services.coding_layers import (
    CONSENSUS_ORIGIN,
    code_usage_count_expr,
    non_consensus_filter,
    project_scoped_segments,
    visible_target_filter,
)
from ..auth import get_current_user
from .helpers import _get_project_or_404, sanitize_content_disposition
from ..services.timestamp import format_timecode
from .excerpts import _base_excerpt_query, _excerpt_to_response
from .export_helpers import (
    _build_category_tree_and_chains,
    build_code_source_matrix,
    segment_source_pair,
    build_code_cooccurrence_matrix,
    EXPORT_VALUE_PRECISION,
    excel_set_safe,
    local_wall_time,
)

router = APIRouter(tags=["export"])


@router.get("/excel")
def export_study_excel(
    project_id: int,
    include_coded_data: bool = True,
    include_matrix: bool = True,
    include_cooccurrence: bool = True,
    include_codebook: bool = True,
    include_memos: bool = True,
    include_notes: bool = True,
    include_quotes: bool = True,
    include_summaries: bool = True,
    include_audit: bool = True,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export project data as Excel with up to 9 sheets: Coded Data, Code-Source Matrix,
    Code Co-occurrence, Codebook, Memos, Notes, Quotes, Summaries, Audit Trail.

    #620: the Coded Data and Notes sheets span all THREE segment parents /
    FOUR note parents; before this they inner-joined Conversation and silently
    omitted documents (since documents shipped) and observation clips. Quotes is
    a new sheet — excerpts previously reached Excel only as a Yes/No flag.

    #629: the matrix sheet was the last one #620 left conversation-only. It is
    now "Code-Source Matrix" — every segment parent on the axis, each column
    naming its source TYPE, and the sheet no longer disappears on a project
    without conversations.
    """
    project = _get_project_or_404(db, project_id, user.id)

    wb = Workbook()
    worksheets = []

    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    diagonal_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Get all codes for this project (used by multiple sheets)
    codes = db.query(Code).filter(
        Code.project_id == project_id,
        Code.is_active == True
    ).order_by(Code.numeric_id).all()

    # Get all conversations (used by multiple sheets)
    conversations = db.query(Conversation).filter(
        Conversation.project_id == project_id
    ).order_by(Conversation.created_at).all()

    # The other two source types, for the three-parent sheets (#620).
    documents = db.query(Document).filter(
        Document.project_id == project_id
    ).order_by(Document.created_at).all()
    observations = db.query(Observation).filter(
        Observation.project_id == project_id
    ).order_by(Observation.created_at).all()

    # Create lookup dicts for entity name resolution
    code_id_to_name = {c.id: c.name for c in codes}
    conversation_id_to_name = {c.id: c.name for c in conversations}
    document_id_to_name = {d.id: d.name for d in documents}
    observation_id_to_name = {o.id: o.name for o in observations}

    # A dataset-value note names its dataset + column, not the value's id (#620).
    dataset_value_source = {
        vid: f"{ds_name} · {col_name}"
        for vid, col_name, ds_name in db.query(
            DatasetValue.id, DatasetColumn.column_name, Dataset.name
        )
        .join(DatasetColumn, DatasetValue.column_id == DatasetColumn.id)
        .join(Dataset, DatasetColumn.dataset_id == Dataset.id)
        .filter(Dataset.project_id == project_id)
        .all()
    }

    # Pre-query all whole-segment excerpts for this project (whole-segment excerpt lookup)
    quoted_seg_ids = set(
        eid for (eid,) in db.query(Excerpt.segment_id).filter(
            Excerpt.project_id == project_id,
            segment_has_any_quote_filter(),
        ).all()
    )

    # ==================== Sheet 1: Coded Data ====================
    if include_coded_data:
        ws_coded = wb.active
        ws_coded.title = "Coded Data"
        worksheets.append(ws_coded)

        # #620: "Conversation" became the honest Source Type + Source pair —
        # this sheet spans all THREE segment parents now, exactly as the
        # coded-segments CSV has since #616, so the source column must say WHAT
        # it names. Conversation rows keep every value they had.
        coded_headers = [
            "Source Type", "Source", "Segment ID", "Sequence", "Speaker", "Is Facilitator",
            "Start Time", "End Time", "Text", "Quoted"
        ]
        coded_headers.extend([f"{c.numeric_id} - {c.name}" for c in codes])
        first_code_col = len(coded_headers) - len(codes) + 1

        for col, header in enumerate(coded_headers, 1):
            cell = ws_coded.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Bulk-load every visible segment in the project, whatever its parent
        # (#620). `project_scoped_segments` is the ONE three-parent scope — never
        # a hand-rolled third `or_`. Before this, `Segment.conversation_id.in_()`
        # silently excluded document segments (since documents shipped) and
        # observation clips, so a sheet titled "Coded Data" was a partial answer
        # presenting itself as a complete one.
        all_segments = project_scoped_segments(
            db.query(Segment).options(
                selectinload(Segment.code_applications),
                joinedload(Segment.speaker),
                joinedload(Segment.conversation),
                joinedload(Segment.document),
                joinedload(Segment.observation),
            ),
            project_id,
        ).filter(
            Segment.merged_into_id == None,
            Segment.split_into_id == None,
        ).order_by(
            func.coalesce(Conversation.name, Document.name, Observation.name),
            Segment.sequence_order,
        ).all()

        row = 2
        for segment in all_segments:
            # J2-B: human layer only (#490 latent site — an "X" can't flip
            # today since consensus ⊆ agreed codes, but keep the layer rule).
            applied_code_ids = set(
                ca.code_id for ca in segment.code_applications
                if ca.origin != CONSENSUS_ORIGIN
            )

            source_kind, source_name = segment_source_pair(segment)

            # Speaker/facilitator are conversation-only concepts; they degrade to
            # blank on the other two parents rather than inventing a value.
            speaker_name = ""
            is_facilitator = ""
            if segment.speaker:
                speaker_name = segment.speaker.name
                is_facilitator = "Yes" if segment.speaker.is_facilitator else "No"

            start_time = f"{segment.start_time:.2f}" if segment.start_time is not None else ""
            end_time = f"{segment.end_time:.2f}" if segment.end_time is not None else ""

            ws_coded.cell(row=row, column=1, value=source_kind)
            excel_set_safe(ws_coded.cell(row=row, column=2), source_name)
            ws_coded.cell(row=row, column=3, value=segment.id)
            ws_coded.cell(row=row, column=4, value=segment.sequence_order)
            excel_set_safe(ws_coded.cell(row=row, column=5), speaker_name)
            ws_coded.cell(row=row, column=6, value=is_facilitator)
            ws_coded.cell(row=row, column=7, value=start_time)
            ws_coded.cell(row=row, column=8, value=end_time)
            excel_set_safe(ws_coded.cell(row=row, column=9), segment.text)
            ws_coded.cell(row=row, column=10, value="Yes" if segment.id in quoted_seg_ids else "")

            for col_offset, code in enumerate(codes):
                value = "X" if code.id in applied_code_ids else ""
                ws_coded.cell(row=row, column=first_code_col + col_offset, value=value)

            row += 1
    else:
        # Remove the default sheet if not including coded data
        wb.remove(wb.active)

    # ==================== Sheet 2: Code-Source Matrix ====================
    # #629: was "Code-Conversation Matrix", gated on `conversations` and keyed by
    # bare conversation id. Three defects in one sheet: documents and observation
    # clips were absent from the axis; a document-only or observation-only
    # project got NO SHEET AT ALL (the `and conversations` gate — the #626/#627
    # `isEmpty` shape, silent, with nothing saying why); and the bare-id key
    # would have summed conversation 5 into document 5 the moment it widened.
    matrix_sources = (
        [("conversation", c.id, c.name) for c in conversations]
        + [("document", d.id, d.name) for d in documents]
        + [("observation", o.id, o.name) for o in observations]
    )
    if include_matrix and codes and matrix_sources:
        ws_matrix = wb.create_sheet("Code-Source Matrix")
        worksheets.append(ws_matrix)

        matrix_data = build_code_source_matrix(db, project_id)

        # Header row: "Code" + one column per source + Total. The source's TYPE
        # rides its header because two sources of different types may share a
        # name ("Session 1" as both a transcript and a recording) and the column
        # would otherwise be ambiguous — the matrix analogue of the Source Type +
        # Source column pair #620 introduced on the row-shaped sheets.
        # Source names are user-supplied AND lead the string, so excel_set_safe
        # is load-bearing here, not decorative.
        matrix_headers = (
            ["Code"]
            + [f"{name} ({kind})" for kind, _sid, name in matrix_sources]
            + ["Total"]
        )
        for col, header in enumerate(matrix_headers, 1):
            cell = ws_matrix.cell(row=1, column=col)
            excel_set_safe(cell, header)
            cell.fill = header_fill
            cell.font = header_font

        # Keep the Code column and the header row on screen: with three source
        # types the axis is materially wider than it was, and scrolling right
        # used to strand the reader among unlabelled numbers.
        ws_matrix.freeze_panes = "B2"

        for row_num, code in enumerate(codes, 2):
            # Leading "{numeric_id} - " makes formula-prefix risk negligible
            # (numeric_id is auto-increment positive int), but defang anyway
            # for consistency.
            excel_set_safe(ws_matrix.cell(row=row_num, column=1), f"{code.numeric_id} - {code.name}")
            row_total = 0
            for col_num, (kind, source_id, _name) in enumerate(matrix_sources, 2):
                count = matrix_data.get(((kind, source_id), code.id), 0)
                ws_matrix.cell(row=row_num, column=col_num, value=count if count > 0 else "")
                row_total += count
            ws_matrix.cell(row=row_num, column=len(matrix_sources) + 2, value=row_total)

    # ==================== Sheet 3: Code Co-occurrence ====================
    if include_cooccurrence and codes:
        ws_cooccur = wb.create_sheet("Code Co-occurrence")
        worksheets.append(ws_cooccur)

        cooccur_data = build_code_cooccurrence_matrix(db, project_id)

        # Header row: scope label in the corner + code names. The corner cell
        # states the facilitator scope so the sheet's numbers carry their claim
        # (#493) without shifting the matrix layout.
        cooccur_headers = ["Participant segments only"] + [
            f"{c.numeric_id} - {c.name}" for c in codes
        ]
        for col, header in enumerate(cooccur_headers, 1):
            cell = ws_cooccur.cell(row=1, column=col)
            excel_set_safe(cell, header)
            cell.fill = header_fill
            cell.font = header_font

        for row_num, code_row in enumerate(codes, 2):
            excel_set_safe(ws_cooccur.cell(row=row_num, column=1), f"{code_row.numeric_id} - {code_row.name}")
            for col_num, code_col in enumerate(codes, 2):
                count = cooccur_data.get((code_row.id, code_col.id), 0)
                cell = ws_cooccur.cell(row=row_num, column=col_num, value=count if count > 0 else "")
                # Highlight diagonal
                if code_row.id == code_col.id:
                    cell.fill = diagonal_fill

    # ==================== Sheet 4: Codebook ====================
    if include_codebook:
        ws_codebook = wb.create_sheet("Codebook")
        worksheets.append(ws_codebook)

        codebook_headers = ["ID", "Name", "Description", "Universal", "Active", "Usage Count", "Created", "Category", "Category Path", "Category Depth"]
        for col, header in enumerate(codebook_headers, 1):
            cell = ws_codebook.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Batch usage counts (avoid N+1). Track J · J2: distinct targets, not raw
        # rows, excluding the consensus layer (single-sourced with the codes list).
        code_ids = [c.id for c in codes]
        usage_counts = {}
        if code_ids:
            usage_rows = db.query(
                CodeApplication.code_id, code_usage_count_expr()
            ).outerjoin(
                Segment, CodeApplication.segment_id == Segment.id
            ).filter(
                CodeApplication.code_id.in_(code_ids),
                non_consensus_filter(),
                visible_target_filter(),  # #500
            ).group_by(CodeApplication.code_id).all()
            usage_counts = dict(usage_rows)

        # Category chain lookup
        cb_chain_map, _, _ = _build_category_tree_and_chains(db, project_id)

        for row_num, code in enumerate(codes, 2):
            ws_codebook.cell(row=row_num, column=1, value=code.numeric_id)
            excel_set_safe(ws_codebook.cell(row=row_num, column=2), code.name)
            excel_set_safe(ws_codebook.cell(row=row_num, column=3), code.description or "")
            ws_codebook.cell(row=row_num, column=4, value="Yes" if code.is_universal else "No")
            ws_codebook.cell(row=row_num, column=5, value="Yes" if code.is_active else "No")
            ws_codebook.cell(row=row_num, column=6, value=usage_counts.get(code.id, 0))
            ws_codebook.cell(row=row_num, column=7, value=local_wall_time(code.created_at))
            # Category columns
            cat_chain = cb_chain_map.get(code.category_id, []) if code.category_id else []
            excel_set_safe(ws_codebook.cell(row=row_num, column=8), cat_chain[-1] if cat_chain else "")
            excel_set_safe(ws_codebook.cell(row=row_num, column=9), " › ".join(cat_chain) if cat_chain else "")
            ws_codebook.cell(row=row_num, column=10, value=len(cat_chain) - 1 if cat_chain else "")

    # ==================== Sheet 5: Memos ====================
    if include_memos:
        ws_memos = wb.create_sheet("Memos")
        worksheets.append(ws_memos)

        memo_headers = ["ID", "Content", "Link Type", "Link Name", "Created", "Updated"]
        for col, header in enumerate(memo_headers, 1):
            cell = ws_memos.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        memos = db.query(Memo).filter(
            Memo.project_id == project_id,
            Memo.is_archived == False
        ).order_by(Memo.created_at).all()

        for row_num, memo in enumerate(memos, 2):
            # Resolve entity name
            # #620: this ladder had arms for FOUR of the ten memo-able entity
            # types (see MEMO_ENTITY_REMAP), so a memo on a document,
            # observation, dataset or saved analysis exported with Link Type
            # filled in and Link Name BLANK — the row named a link it couldn't
            # resolve. Unknown types now fall back to "{Type} {id}", which is
            # ugly but never silently empty.
            link_name = ""
            if memo.entity_type == "project":
                link_name = project.name
            elif memo.entity_type == "conversation":
                link_name = conversation_id_to_name.get(memo.entity_id, f"Conversation {memo.entity_id}")
            elif memo.entity_type == "document":
                link_name = document_id_to_name.get(memo.entity_id, f"Document {memo.entity_id}")
            elif memo.entity_type == "observation":
                link_name = observation_id_to_name.get(memo.entity_id, f"Observation {memo.entity_id}")
            elif memo.entity_type == "code":
                link_name = code_id_to_name.get(memo.entity_id, f"Code {memo.entity_id}")
            elif memo.entity_type == "code_category":
                link_name = f"Category {memo.entity_id}"
            else:
                link_name = f"{memo.entity_type.replace('_', ' ').capitalize()} {memo.entity_id}"

            ws_memos.cell(row=row_num, column=1, value=f"M-{memo.numeric_id}")
            excel_set_safe(ws_memos.cell(row=row_num, column=2), memo.content)
            ws_memos.cell(row=row_num, column=3, value=memo.entity_type.capitalize())
            excel_set_safe(ws_memos.cell(row=row_num, column=4), link_name)
            ws_memos.cell(row=row_num, column=5, value=local_wall_time(memo.created_at))
            ws_memos.cell(row=row_num, column=6, value=local_wall_time(memo.updated_at))

    # ==================== Sheet 6: Notes ====================
    if include_notes:
        ws_notes = wb.create_sheet("Notes")
        worksheets.append(ws_notes)

        # #620: the "Conversation" column became Source Type + Source. A Note
        # hangs off FOUR parents (conversation · document · observation ·
        # dataset value), and the old `.join(Conversation)` was an INNER join, so
        # every document note (live since documents shipped), every clip note and
        # every dataset-value note was silently absent from a sheet titled
        # "Notes".
        note_headers = [
            "ID", "Content", "Source Type", "Source", "Segment #", "Segment Text",
            "Created", "Updated",
        ]
        for col, header in enumerate(note_headers, 1):
            cell = ws_notes.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # ONE query with outer joins rather than four per-parent queries: a note
        # has exactly one parent, so the ORs are disjoint and nothing multiplies.
        # The dataset-value arm needs three hops (value → column → dataset) —
        # DatasetValue carries no project_id of its own.
        notes = (
            db.query(Note)
            .options(joinedload(Note.segment))
            .outerjoin(Conversation, Note.conversation_id == Conversation.id)
            .outerjoin(Document, Note.document_id == Document.id)
            .outerjoin(Observation, Note.observation_id == Observation.id)
            .outerjoin(DatasetValue, Note.dataset_value_id == DatasetValue.id)
            .outerjoin(DatasetColumn, DatasetValue.column_id == DatasetColumn.id)
            .outerjoin(Dataset, DatasetColumn.dataset_id == Dataset.id)
            .filter(
                Note.is_archived == False,
                or_(
                    Conversation.project_id == project_id,
                    Document.project_id == project_id,
                    Observation.project_id == project_id,
                    Dataset.project_id == project_id,
                ),
            )
            .order_by(Note.created_at)
            .all()
        )

        for row_num, note in enumerate(notes, 2):
            if note.conversation_id is not None:
                source_kind = "conversation"
                source_name = conversation_id_to_name.get(note.conversation_id, "")
            elif note.document_id is not None:
                source_kind, source_name = "document", document_id_to_name.get(note.document_id, "")
            elif note.observation_id is not None:
                source_kind = "observation"
                source_name = observation_id_to_name.get(note.observation_id, "")
            elif note.dataset_value_id is not None:
                # A response-level note: the dataset+column is the useful
                # locator, not the value's own id.
                source_kind = "dataset value"
                source_name = dataset_value_source.get(note.dataset_value_id, "")
            else:
                source_kind, source_name = "", ""

            segment_num = ""
            segment_text = ""
            if note.segment:
                segment_num = note.segment.sequence_order
                # Truncate segment text to 200 chars
                segment_text = note.segment.text[:200] + "..." if len(note.segment.text) > 200 else note.segment.text

            ws_notes.cell(row=row_num, column=1, value=f"N-{note.sequence_number}")
            excel_set_safe(ws_notes.cell(row=row_num, column=2), note.content)
            ws_notes.cell(row=row_num, column=3, value=source_kind)
            excel_set_safe(ws_notes.cell(row=row_num, column=4), source_name)
            ws_notes.cell(row=row_num, column=5, value=segment_num)
            excel_set_safe(ws_notes.cell(row=row_num, column=6), segment_text)
            # #513: localize like the Codebook/Memos sheets — raw strftime emits UTC
            ws_notes.cell(row=row_num, column=7, value=local_wall_time(note.created_at))
            ws_notes.cell(row=row_num, column=8, value=local_wall_time(note.updated_at))

    # ==================== Sheet 7: Quotes ====================
    #
    # #620: excerpts reached this workbook only as the Coded Data sheet's
    # Yes/No "Quoted" flag — the quote's own TEXT, its range and its note were
    # nowhere, so the qualitative payload a researcher most wants to work with
    # in a spreadsheet was the one thing the study export omitted. The
    # `/export/excerpts` CSV already emits exactly these columns; this sheet
    # reuses ITS builders (`_base_excerpt_query` / `_excerpt_to_response`)
    # rather than re-deriving shape and text, because "which shape is this and
    # what text does it carry" is single-sourced by rule (slab 5a) and a second
    # copy is a drift waiting to happen.
    if include_quotes:
        ws_quotes = wb.create_sheet("Quotes")
        worksheets.append(ws_quotes)

        quote_headers = [
            "Excerpt ID", "Source Type", "Source", "Speaker", "Timestamp",
            "Start Time", "End Time", "Duration", "Quote Text", "Type", "Note", "Created",
        ]
        for col, header in enumerate(quote_headers, 1):
            cell = ws_quotes.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Excerpts on soft-deleted segments are UI-unreachable and must not
        # export (the outerjoin keeps comment excerpts, which have no segment).
        excerpts = _base_excerpt_query(db, project_id).outerjoin(
            Segment, Excerpt.segment_id == Segment.id
        ).filter(
            or_(
                Excerpt.segment_id.is_(None),
                and_(Segment.merged_into_id == None, Segment.split_into_id == None),
            )
        ).order_by(Excerpt.created_at).all()

        for row_num, exc in enumerate(excerpts, 2):
            resp = _excerpt_to_response(exc)
            # #736: one dispatch, shared with the response builder and the
            # excerpts CSV. This block used to hand-roll its own copy.
            source_kind, source_name = resp.source_kind, resp.source_name
            range_start = range_end = duration = ""
            if exc.segment_id is not None:
                seg = exc.segment
                if exc.start_offset is not None:
                    excerpt_type = "sub-segment"
                elif exc.start_time is not None:
                    excerpt_type = "time-range"
                    range_start = format_timecode(exc.start_time)
                    range_end = format_timecode(exc.end_time)
                    duration = format_timecode(exc.end_time - exc.start_time)
                else:
                    excerpt_type = "whole-segment"
                if seg is not None and seg.observation_id is not None:
                    # A whole-clip quote's identity IS the clip's range — its
                    # label is often blank, so without this the row would be
                    # emptiest exactly where it matters most (mirrors the CSV).
                    if exc.start_time is None and seg.start_time is not None and seg.end_time is not None:
                        range_start = format_timecode(seg.start_time)
                        range_end = format_timecode(seg.end_time)
                        duration = format_timecode(seg.end_time - seg.start_time)
            else:
                excerpt_type = "text"

            ws_quotes.cell(row=row_num, column=1, value=exc.id)
            ws_quotes.cell(row=row_num, column=2, value=source_kind)
            excel_set_safe(ws_quotes.cell(row=row_num, column=3), source_name)
            excel_set_safe(ws_quotes.cell(row=row_num, column=4), resp.speaker_name or "")
            # #733: `or ""` blanked a segment starting at exactly 0.0 — the
            # first turn of a timestamped transcript, or a clip at the top of a
            # recording. The CSV sibling carried the identical defect.
            ws_quotes.cell(
                row=row_num, column=5,
                value=resp.segment_timestamp if resp.segment_timestamp is not None else "",
            )
            ws_quotes.cell(row=row_num, column=6, value=range_start)
            ws_quotes.cell(row=row_num, column=7, value=range_end)
            ws_quotes.cell(row=row_num, column=8, value=duration)
            excel_set_safe(ws_quotes.cell(row=row_num, column=9), resp.excerpt_text)
            ws_quotes.cell(row=row_num, column=10, value=excerpt_type)
            excel_set_safe(ws_quotes.cell(row=row_num, column=11), resp.note.content if resp.note else "")
            ws_quotes.cell(row=row_num, column=12, value=local_wall_time(exc.created_at))

    # ==================== Sheet 8: Summaries ====================
    if include_summaries:
        ws_summaries = wb.create_sheet("Summaries")
        worksheets.append(ws_summaries)

        # #620: Document carries its OWN `summary` column and was absent here.
        # Observations are deliberately NOT included: they have `description`,
        # which is a different field with a different meaning — folding it into a
        # sheet called "Summaries" would mislabel it.
        summary_headers = ["Source Type", "Source", "Subject ID", "Date", "Status", "Summary"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summaries.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row_num = 2
        for conversation in conversations:
            ws_summaries.cell(row=row_num, column=1, value="conversation")
            excel_set_safe(ws_summaries.cell(row=row_num, column=2), conversation.name)
            excel_set_safe(ws_summaries.cell(row=row_num, column=3), conversation.subject_id or "")
            ws_summaries.cell(row=row_num, column=4, value=conversation.conversation_date.strftime("%Y-%m-%d") if conversation.conversation_date else "")
            ws_summaries.cell(row=row_num, column=5, value=conversation.status.value if conversation.status else "")
            excel_set_safe(ws_summaries.cell(row=row_num, column=6), conversation.summary or "")
            row_num += 1

        for document in documents:
            # Subject ID / Date / Status are conversation-only fields; they stay
            # blank rather than being faked from something adjacent.
            ws_summaries.cell(row=row_num, column=1, value="document")
            excel_set_safe(ws_summaries.cell(row=row_num, column=2), document.name)
            excel_set_safe(ws_summaries.cell(row=row_num, column=6), document.summary or "")
            row_num += 1

    # ==================== Sheet 8: Audit Trail ====================
    if include_audit:
        ws_audit = wb.create_sheet("Audit Trail")
        worksheets.append(ws_audit)

        audit_headers = ["Timestamp", "Action", "Entity Type", "Entity ID", "Details"]
        for col, header in enumerate(audit_headers, 1):
            cell = ws_audit.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        audit_entries = db.query(AuditEntry).filter(
            AuditEntry.project_id == project_id
        ).order_by(AuditEntry.timestamp.desc()).limit(1000).all()

        for row_num, entry in enumerate(audit_entries, 2):
            # #513: localized; keeps second precision for the audit trail
            ws_audit.cell(row=row_num, column=1, value=local_wall_time(entry.timestamp, "%Y-%m-%d %H:%M:%S"))
            excel_set_safe(ws_audit.cell(row=row_num, column=2), entry.action)
            excel_set_safe(ws_audit.cell(row=row_num, column=3), entry.entity_type)
            ws_audit.cell(row=row_num, column=4, value=entry.entity_id)
            excel_set_safe(ws_audit.cell(row=row_num, column=5), entry.details or "")

    # Adjust column widths for all worksheets
    for ws in worksheets:
        for column in ws.columns:
            if not hasattr(column[0], 'column_letter'):
                continue
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sanitize_content_disposition(project.name)}_export_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/datasets-excel")
def export_datasets_excel(
    project_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all datasets as Excel with one sheet per dataset plus a Data Dictionary."""
    project = _get_project_or_404(db, project_id, user.id)

    # Get all datasets for this project
    datasets = db.query(Dataset).filter(
        Dataset.project_id == project_id
    ).order_by(Dataset.name).all()

    if not datasets:
        raise HTTPException(status_code=404, detail="No datasets found")

    wb = Workbook()

    # Styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    recode_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    recode_font = Font(bold=True)

    worksheets = []
    # Track data for Data Dictionary sheet
    dict_rows = []

    # Excel disallows / \ : * ? [ ] in sheet titles (and a 31-char cap).
    _SHEET_INVALID = ("/", "\\", ":", "*", "?", "[", "]")

    def _sanitize_sheet_title(name: str) -> str:
        cleaned = name
        for ch in _SHEET_INVALID:
            cleaned = cleaned.replace(ch, "_")
        cleaned = cleaned[:31] or "Sheet"
        return cleaned

    for ds_idx, dataset in enumerate(datasets):
        # Create sheet (use active for first, create for rest)
        if ds_idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = _sanitize_sheet_title(dataset.name)
        worksheets.append(ws)

        # Load columns (skip SKIP type), ordered by sequence_order
        columns = [col for col in dataset.columns
                   if col.column_type != ColumnType.SKIP]

        # #592: per-column missing rules for on-the-fly recode computation
        # (parsed once, not per cell). None = the recognized-N/A defaults.
        missing_rules_by_col = {
            col.id: parse_missing_rules(col.missing_values) for col in columns
        }

        # Build column layout: for each column, determine export columns
        # Each entry: (header_text, is_recode, column, recode_definition_or_None)
        col_defs = []
        # Leading columns
        col_defs.append(("Record", False, None, None))
        col_defs.append(("Participant", False, None, None))

        for column in columns:
            # Raw value column
            col_defs.append((column.column_code or f"C{column.sequence_order + 1:03d}",
                             False, column, None))

            # Primary recode column (value_numeric)
            primary_def = None
            non_primary_defs = []
            for rd in column.recode_definitions:
                if rd.is_primary:
                    primary_def = rd
                else:
                    non_primary_defs.append(rd)

            if primary_def:
                code = column.column_code or f"C{column.sequence_order + 1:03d}"
                col_defs.append((f"{code} [numeric]", True, column, primary_def))

            # Non-primary recode columns
            for rd in non_primary_defs:
                code = column.column_code or f"C{column.sequence_order + 1:03d}"
                col_defs.append((f"{code} [{rd.name}]", True, column, rd))

            # Collect data dictionary rows
            if not primary_def and not non_primary_defs:
                # No recodes — one row with empty recode columns
                scale_labels_str = ""
                if column.scale_labels:
                    try:
                        labels = json.loads(column.scale_labels)
                        scale_labels_str = ", ".join(str(l) for l in labels)
                    except (json.JSONDecodeError, TypeError):
                        scale_labels_str = column.scale_labels
                dict_rows.append({
                    "dataset": dataset.name,
                    "code": column.column_code or "",
                    "text": column.column_text,
                    "type": column.column_type.value,
                    "scale_labels": scale_labels_str,
                    "missing": describe_missing_rules(missing_rules_by_col.get(column.id)),
                    "source": column.source or "imported",
                    "recode_name": "",
                    "recode_type": "",
                    "mapping": "",
                })
            else:
                all_defs = ([primary_def] if primary_def else []) + non_primary_defs
                for rd in all_defs:
                    scale_labels_str = ""
                    if column.scale_labels:
                        try:
                            labels = json.loads(column.scale_labels)
                            scale_labels_str = ", ".join(str(l) for l in labels)
                        except (json.JSONDecodeError, TypeError):
                            scale_labels_str = column.scale_labels
                    mapping_str = ""
                    if rd.mapping:
                        try:
                            m = json.loads(rd.mapping)
                            mapping_str = "; ".join(f"{k} -> {v}" for k, v in m.items())
                        except (json.JSONDecodeError, TypeError):
                            mapping_str = rd.mapping
                    dict_rows.append({
                        "dataset": dataset.name,
                        "code": column.column_code or "",
                        "text": column.column_text,
                        "type": column.column_type.value,
                        "scale_labels": scale_labels_str,
                        "missing": describe_missing_rules(missing_rules_by_col.get(column.id)),
                        "source": column.source or "imported",
                        "recode_name": rd.name + (" (primary)" if rd.is_primary else ""),
                        "recode_type": rd.recode_type.value if hasattr(rd.recode_type, 'value') else str(rd.recode_type),
                        "mapping": mapping_str,
                    })

        # Write header row. Column codes flow into headers; defang.
        for col_idx, (header_text, is_recode, _, _) in enumerate(col_defs, 1):
            cell = ws.cell(row=1, column=col_idx)
            excel_set_safe(cell, header_text)
            if is_recode:
                cell.fill = recode_fill
                cell.font = recode_font
            else:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Load responses with answers and participant
        responses = db.query(DatasetRow).filter(
            DatasetRow.dataset_id == dataset.id
        ).order_by(DatasetRow.id).all()

        # Build answer lookup: {(row_id, column_id): answer}
        answer_lookup = {}
        for response in responses:
            for answer in response.values:
                answer_lookup[(response.id, answer.column_id)] = answer

        # Write data rows
        for row_idx, response in enumerate(responses, 2):
            # Record
            excel_set_safe(ws.cell(row=row_idx, column=1), response.row_identifier or "")
            # Participant
            participant_name = ""
            if response.participant:
                participant_name = response.participant.display_name or response.participant.identifier or ""
            excel_set_safe(ws.cell(row=row_idx, column=2), participant_name)

            # Data columns
            col_idx = 3
            for header_text, is_recode, column, recode_def in col_defs[2:]:
                answer = answer_lookup.get((response.id, column.id))
                if answer is None:
                    ws.cell(row=row_idx, column=col_idx, value="")
                elif recode_def is None:
                    # Raw value column.
                    #
                    # 🔴 **A missing cell is EMPTY here, matching the R export
                    # (#822, 2026-08-25) — and this REVERSES the #592 §I.10 /
                    # #611e decision that used to sit in this comment.** That
                    # decision made this sheet a raw-data escape hatch where a
                    # declared "99" or a recognized "N/A" stayed visible; the
                    # cost, measured on a real survey, is that the two exports
                    # of one project disagreed about what the sample IS. R
                    # blanked and Excel wrote the sentinel text (1,099,939
                    # cells), so a recipient comparing them got a different N,
                    # and anyone averaging a column in Excel silently included
                    # "No answer" rows — a text sentinel in a numeric column is
                    # a trap in the tool people actually open it in.
                    #
                    # ⚠️ **IDENTIFIER columns are the exception, and it is
                    # carried over from R deliberately (#533): an ID is a join
                    # key, not an analysis value.** Blanking an "N/A"-looking
                    # identifier would make the Excel export unjoinable exactly
                    # where the R export stays joinable — the same divergence
                    # this fix exists to remove, pointing the other way.
                    #
                    # The DISTINCTION between "Do not know", "No answer" and
                    # "Inapplicable" is preserved by the Data Dictionary's
                    # Missing Values column, not by the cell.
                    raw_text = answer.value_text or ""
                    if (
                        column.column_type != ColumnType.IDENTIFIER
                        and is_missing(raw_text, missing_rules_by_col.get(column.id))
                    ):
                        raw_text = ""
                    excel_set_safe(ws.cell(row=row_idx, column=col_idx), raw_text)
                elif recode_def.is_primary:
                    # Primary recode: use stored value_numeric
                    val = answer.value_numeric
                    ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
                else:
                    # Non-primary recode: compute on the fly (#592: the
                    # column's missing declaration governs the null-set,
                    # same as the stored primary values).
                    if answer.value_text:
                        computed = compute_value(
                            answer.value_text, recode_def,
                            missing_rules=missing_rules_by_col.get(column.id),
                        )
                        ws.cell(row=row_idx, column=col_idx, value=computed if computed is not None else "")
                    else:
                        ws.cell(row=row_idx, column=col_idx, value="")
                col_idx += 1

    # ==================== Data Dictionary sheet ====================
    ws_dict = wb.create_sheet("Data Dictionary")
    worksheets.append(ws_dict)

    # #822: "Missing Values" sits beside the scale it qualifies. Without it the
    # export whose stated purpose is "with Data Dictionary" shipped a file in
    # which the single most important interpretive fact about a column — which
    # of its codes are non-answers — was absent, and (since the cells are now
    # blanked to match R) unrecoverable from the data sheet.
    dict_headers = ["Dataset", "Code", "Column label", "Type", "Scale Labels",
                     "Missing Values", "Source", "Recode Name", "Recode Type",
                     "Mapping"]
    for col_idx, header in enumerate(dict_headers, 1):
        cell = ws_dict.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(dict_rows, 2):
        excel_set_safe(ws_dict.cell(row=row_idx, column=1), row_data["dataset"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=2), row_data["code"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=3), row_data["text"])
        ws_dict.cell(row=row_idx, column=4, value=row_data["type"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=5), row_data["scale_labels"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=6), row_data["missing"])
        ws_dict.cell(row=row_idx, column=7, value=row_data["source"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=8), row_data["recode_name"])
        ws_dict.cell(row=row_idx, column=9, value=row_data["recode_type"])
        excel_set_safe(ws_dict.cell(row=row_idx, column=10), row_data["mapping"])

    # ==================== Computed Metrics Sheets ====================

    METRIC_TYPE_DISPLAY = {
        "frequency_distribution": "Frequency Distribution",
        "proportion": "Proportion (% Meeting Threshold)",
        "mean": "Mean",
        "domain_aggregate": "Domain Aggregate",
    }

    stale_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    def safe_sheet_name(name: str) -> str:
        """Ensure sheet name is unique (max 31 chars) and doesn't collide with existing sheets."""
        name = name[:31]
        if name in wb.sheetnames:
            name = f"{name[:18]} (Computed)"[:31]
        return name

    # Load all metrics with results for this project
    all_metrics = (
        db.query(MetricDefinition)
        .options(joinedload(MetricDefinition.results))
        .filter(MetricDefinition.project_id == project_id)
        .order_by(MetricDefinition.sequence_order)
        .all()
    )

    metrics_with_results = [m for m in all_metrics if len(m.results) > 0]

    if metrics_with_results:
        # Batch resolve labels
        label_map = resolve_input_source_labels(db, metrics_with_results)

        # Build domain membership map: (member_type, member_id) → domain_name
        domain_members = (
            db.query(AnalysisDomainMember, AnalysisDomain.name)
            .join(AnalysisDomain)
            .filter(AnalysisDomain.project_id == project_id)
            .all()
        )
        domain_name_map = {
            (dm.member_type, dm.member_id): dname
            for dm, dname in domain_members
        }

        # Build column → equivalence_group_id map for indirection lookup
        metric_col_ids = [
            m.input_source_id for m in metrics_with_results
            if m.input_source_type == "dataset_column"
        ]
        col_equiv_map = {}
        if metric_col_ids:
            col_eq_rows = (
                db.query(DatasetColumn.id, DatasetColumn.equivalence_group_id)
                .filter(DatasetColumn.id.in_(metric_col_ids))
                .all()
            )
            col_equiv_map = {cid: eq_id for cid, eq_id in col_eq_rows if eq_id is not None}

        # Build domain name map for AnalysisDomain IDs (for dataset_domain metrics)
        domain_id_name_map = {}
        domain_ids = [m.input_source_id for m in metrics_with_results if m.input_source_type == "dataset_domain"]
        if domain_ids:
            domain_rows = (
                db.query(AnalysisDomain.id, AnalysisDomain.name)
                .filter(AnalysisDomain.id.in_(domain_ids))
                .all()
            )
            domain_id_name_map = {did: dname for did, dname in domain_rows}

        def resolve_domain_name(metric: MetricDefinition) -> str:
            """Resolve domain name for a metric."""
            if metric.input_source_type == "dataset_domain":
                return domain_id_name_map.get(metric.input_source_id, "")
            # dataset_column: check direct column membership
            dn = domain_name_map.get(("column", metric.input_source_id))
            if dn:
                return dn
            # Check via equivalence_group_id indirection
            eq_id = col_equiv_map.get(metric.input_source_id)
            if eq_id:
                dn = domain_name_map.get(("equivalence_group", eq_id))
                if dn:
                    return dn
            return ""

        # Partition metrics
        ungrouped = [m for m in metrics_with_results if m.grouping_column_id is None and m.grouping_column_id_2 is None and m.grouping_mode != "dataset"]
        grouped = [m for m in metrics_with_results if m.grouping_column_id is not None or m.grouping_column_id_2 is not None or m.grouping_mode == "dataset"]

        # ── Sheet: Metrics Summary (ungrouped only) ──
        if ungrouped:
            ws_summary = wb.create_sheet(safe_sheet_name("Metrics Summary"))
            worksheets.append(ws_summary)

            summary_headers = ["#", "Name", "Type", "Input Source", "Domain",
                               "Value", "Valid N", "Total N", "Stale", "Computed At"]
            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, metric in enumerate(ungrouped, 2):
                result = metric.results[0]
                try:
                    rd = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                except (json.JSONDecodeError, TypeError):
                    rd = {}

                source_label = label_map.get(
                    (metric.input_source_type, metric.input_source_id), ""
                )
                domain_name = resolve_domain_name(metric)

                # Value depends on metric type
                value = None
                if metric.metric_type == "proportion":
                    value = rd.get("percentage")
                elif metric.metric_type == "mean":
                    value = rd.get("mean")
                elif metric.metric_type == "domain_aggregate":
                    value = rd.get("aggregate_value")
                # frequency_distribution: leave blank

                ws_summary.cell(row=row_idx, column=1, value=row_idx - 1)
                # metric.name and domain_name flow from user-typed input; the
                # Tier-3 auto-create path also synthesizes f"{domain.name} Score"
                # so the risk is hot here.
                excel_set_safe(ws_summary.cell(row=row_idx, column=2), metric.name)
                ws_summary.cell(row=row_idx, column=3, value=METRIC_TYPE_DISPLAY.get(metric.metric_type, metric.metric_type))
                excel_set_safe(ws_summary.cell(row=row_idx, column=4), source_label)
                excel_set_safe(ws_summary.cell(row=row_idx, column=5), domain_name)
                if value is not None:
                    ws_summary.cell(row=row_idx, column=6, value=round(value, EXPORT_VALUE_PRECISION))
                ws_summary.cell(row=row_idx, column=7, value=result.valid_n)
                ws_summary.cell(row=row_idx, column=8, value=result.total_n)
                ws_summary.cell(row=row_idx, column=9, value="Yes" if metric.stale else "No")
                ws_summary.cell(
                    row=row_idx, column=10,
                    value=local_wall_time(result.computed_at) if result.computed_at else ""  # #513
                )

                if metric.stale:
                    for c in range(1, len(summary_headers) + 1):
                        ws_summary.cell(row=row_idx, column=c).fill = stale_fill

        # ── Sheet: Metrics Detail (freq dist sections + domain agg breakdowns) ──
        ungrouped_freq = [m for m in ungrouped if m.metric_type == "frequency_distribution"]
        ungrouped_domain_agg = [m for m in ungrouped if m.metric_type == "domain_aggregate"]

        if ungrouped_freq or ungrouped_domain_agg:
            ws_detail = wb.create_sheet(safe_sheet_name("Metrics Detail"))
            worksheets.append(ws_detail)
            detail_row = 1

            # Section A: Frequency Distributions
            if ungrouped_freq:
                # Section header
                ws_detail.cell(row=detail_row, column=1, value="FREQUENCY DISTRIBUTIONS")
                ws_detail.cell(row=detail_row, column=1).font = Font(bold=True, size=12)
                ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=7)
                detail_row += 1

                # queue #42: the interval must say what KIND it is, and a
                # spreadsheet has no tooltip to put that in — so the disclosure
                # is a row of its own. Without it a reader would reasonably take
                # seven categories' intervals as jointly covering at 95%, which
                # is the one thing per-category binomial intervals do not do.
                ws_detail.cell(
                    row=detail_row, column=1,
                    value="95% Wilson score interval, computed separately for each response "
                          "option (binomial, not simultaneous across options).",
                )
                ws_detail.cell(row=detail_row, column=1).font = Font(italic=True, size=9)
                ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=7)
                detail_row += 1

                # Column headers
                freq_headers = [
                    "Metric", "Response Option", "Count", "Percentage",
                    "95% CI Lower", "95% CI Upper", "Valid N",
                ]
                for col_idx, header in enumerate(freq_headers, 1):
                    cell = ws_detail.cell(row=detail_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                detail_row += 1

                for metric in ungrouped_freq:
                    result = metric.results[0]
                    try:
                        rd = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                    except (json.JSONDecodeError, TypeError):
                        rd = {}

                    source_label = label_map.get(
                        (metric.input_source_type, metric.input_source_id), metric.name
                    )
                    # #766: this read `rd.get("distribution", {})` — a key NO
                    # producer has ever written. `compute_frequency_distribution`
                    # returns `counts` / `percentages` / `scale_order`, so the
                    # loop below ran zero times and every Excel export emitted
                    # this section as bare headers. `frequency_distribution` is
                    # the DEFAULT metric type, so that is the most common metric
                    # in the app, absent from the export, silently, with no test
                    # covering the section.
                    #
                    # Iterate `scale_order`, not the dict: that is the ordering
                    # seam (#406, numeric-aware) and it is what puts a declared
                    # level nobody chose in its declared place rather than at the
                    # end — the zero-fill (#577/#591) means those keys exist here.
                    counts = rd.get("counts", {}) or {}
                    percentages = rd.get("percentages", {}) or {}
                    order = rd.get("scale_order") or list(counts.keys())
                    # queue #42. Absent on rows computed before it — those export
                    # blank CI cells rather than a fabricated interval.
                    ci_lower = rd.get("ci_lower_by_label", {}) or {}
                    ci_upper = rd.get("ci_upper_by_label", {}) or {}
                    valid_n = result.valid_n

                    first_row = True
                    for option in order:
                        count = counts.get(option, 0)
                        excel_set_safe(
                            ws_detail.cell(row=detail_row, column=1),
                            source_label if first_row else "",
                        )
                        # `option` is a scale-label key, user-supplied.
                        excel_set_safe(ws_detail.cell(row=detail_row, column=2), option)
                        ws_detail.cell(row=detail_row, column=3, value=count)
                        # Read the percentage the server computed; do NOT
                        # recompute it from count/valid_n. That was a second
                        # copy of one fact (#733) — and the two would diverge the
                        # moment the denominator rule changes, which #592 already
                        # did once for declared-missing values.
                        pct = percentages.get(option)
                        if pct is not None:
                            ws_detail.cell(row=detail_row, column=4,
                                           value=round(pct, EXPORT_VALUE_PRECISION))
                        lo, hi = ci_lower.get(option), ci_upper.get(option)
                        if lo is not None:
                            ws_detail.cell(row=detail_row, column=5,
                                           value=round(lo, EXPORT_VALUE_PRECISION))
                        if hi is not None:
                            ws_detail.cell(row=detail_row, column=6,
                                           value=round(hi, EXPORT_VALUE_PRECISION))
                        if first_row:
                            ws_detail.cell(row=detail_row, column=7, value=valid_n)
                        first_row = False
                        detail_row += 1

                    if metric.stale:
                        for r in range(detail_row - len(order), detail_row):
                            for c in range(1, 8):
                                ws_detail.cell(row=r, column=c).fill = stale_fill

                    detail_row += 1  # blank row separator

            # Section B: Domain Aggregate Breakdowns
            if ungrouped_domain_agg:
                detail_row += 1
                ws_detail.cell(row=detail_row, column=1, value="DOMAIN AGGREGATE BREAKDOWNS")
                ws_detail.cell(row=detail_row, column=1).font = Font(bold=True, size=12)
                ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=5)
                detail_row += 1

                agg_headers = ["Domain", "Item", "Item Mean", "Domain Score", "Valid N"]
                for col_idx, header in enumerate(agg_headers, 1):
                    cell = ws_detail.cell(row=detail_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                detail_row += 1

                for metric in ungrouped_domain_agg:
                    result = metric.results[0]
                    try:
                        rd = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                    except (json.JSONDecodeError, TypeError):
                        rd = {}

                    domain_name = resolve_domain_name(metric)
                    agg_value = rd.get("aggregate_value")
                    col_means = rd.get("column_means", {})

                    first_row = True
                    for item_name, item_mean in col_means.items():
                        excel_set_safe(
                            ws_detail.cell(row=detail_row, column=1),
                            domain_name if first_row else "",
                        )
                        # `item_name` is a column code/name, user-supplied.
                        excel_set_safe(ws_detail.cell(row=detail_row, column=2), item_name)
                        ws_detail.cell(row=detail_row, column=3,
                                       value=round(item_mean, EXPORT_VALUE_PRECISION) if item_mean is not None else "")
                        if first_row:
                            ws_detail.cell(row=detail_row, column=4,
                                           value=round(agg_value, EXPORT_VALUE_PRECISION) if agg_value is not None else "")
                            ws_detail.cell(row=detail_row, column=5, value=result.valid_n)
                        first_row = False
                        detail_row += 1

                    if metric.stale:
                        for r in range(detail_row - len(col_means), detail_row):
                            for c in range(1, 6):
                                ws_detail.cell(row=r, column=c).fill = stale_fill

                    detail_row += 1  # blank row separator

        # ── Sheet(s): Grouped metrics ──
        # Build grouping column name lookup
        grp_col_ids = set()
        for m in grouped:
            if m.grouping_column_id:
                grp_col_ids.add(m.grouping_column_id)
            if m.grouping_column_id_2:
                grp_col_ids.add(m.grouping_column_id_2)
        grp_col_name_map = {}
        if grp_col_ids:
            grp_col_rows = (
                db.query(DatasetColumn.id, DatasetColumn.column_code, DatasetColumn.column_name)
                .filter(DatasetColumn.id.in_(grp_col_ids))
                .all()
            )
            grp_col_name_map = {cid: (ccode or cname or f"Col {cid}") for cid, ccode, cname in grp_col_rows}

        if grouped:
            ws_grouped = wb.create_sheet(safe_sheet_name("Grouped Metrics"))
            worksheets.append(ws_grouped)

            group_headers = ["#", "Name", "Type", "Input Source", "Domain",
                             "Group By", "Group Value", "Value", "Valid N", "Total N", "Stale"]
            for col_idx, header in enumerate(group_headers, 1):
                cell = ws_grouped.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            group_row = 2
            for metric in grouped:
                source_label = label_map.get(
                    (metric.input_source_type, metric.input_source_id), metric.name
                )
                domain_name = resolve_domain_name(metric)

                # Build group-by label
                grp_parts = []
                if metric.grouping_mode == "dataset":
                    grp_parts.append("Dataset")
                if metric.grouping_column_id:
                    grp_parts.append(grp_col_name_map.get(metric.grouping_column_id, f"Col {metric.grouping_column_id}"))
                if metric.grouping_column_id_2:
                    grp_parts.append(grp_col_name_map.get(metric.grouping_column_id_2, f"Col {metric.grouping_column_id_2}"))
                grp_label = " × ".join(grp_parts) if grp_parts else ""

                for result in metric.results:
                    try:
                        rd = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                    except (json.JSONDecodeError, TypeError):
                        rd = {}

                    value = None
                    if metric.metric_type == "proportion":
                        value = rd.get("percentage")
                    elif metric.metric_type == "mean":
                        value = rd.get("mean")
                    elif metric.metric_type == "domain_aggregate":
                        value = rd.get("aggregate_value")

                    ws_grouped.cell(row=group_row, column=1, value=group_row - 1)
                    excel_set_safe(ws_grouped.cell(row=group_row, column=2), metric.name)
                    ws_grouped.cell(row=group_row, column=3, value=METRIC_TYPE_DISPLAY.get(metric.metric_type, metric.metric_type))
                    excel_set_safe(ws_grouped.cell(row=group_row, column=4), source_label)
                    excel_set_safe(ws_grouped.cell(row=group_row, column=5), domain_name)
                    excel_set_safe(ws_grouped.cell(row=group_row, column=6), grp_label)
                    # group_value is the partition key, often a user-typed
                    # column-value label (e.g., a treatment-arm name). A NULL
                    # group_value is the None listwise-deletion bucket (rows
                    # whose grouping value was missing) — label it explicitly
                    # rather than emitting a blank-looking group (#506).
                    excel_set_safe(
                        ws_grouped.cell(row=group_row, column=7),
                        MISSING_GROUP_LABEL if result.group_value is None else result.group_value,
                    )
                    if value is not None:
                        ws_grouped.cell(row=group_row, column=8, value=round(value, EXPORT_VALUE_PRECISION))
                    ws_grouped.cell(row=group_row, column=9, value=result.valid_n)
                    ws_grouped.cell(row=group_row, column=10, value=result.total_n)
                    ws_grouped.cell(row=group_row, column=11, value="Yes" if metric.stale else "No")

                    if metric.stale:
                        for c in range(1, len(group_headers) + 1):
                            ws_grouped.cell(row=group_row, column=c).fill = stale_fill

                    group_row += 1

    # Adjust column widths for all worksheets
    for ws in worksheets:
        for column in ws.columns:
            # Skip merged cells (MergedCell has no column_letter)
            if not hasattr(column[0], 'column_letter'):
                continue
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError, ValueError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sanitize_content_disposition(project.name)}_datasets_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
