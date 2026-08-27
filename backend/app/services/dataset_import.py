"""
Dataset CSV import service for Mixed Measures.

Follows the same preview -> import philosophy as csv_import.py but handles
dataset/questionnaire data rather than conversation transcripts.
"""

import csv
import io
import json
import logging
import math
import re
from itertools import islice
from sqlalchemy import insert as sa_insert
from sqlalchemy.orm import Session

from ..models.dataset import (
    ColumnType,
    Dataset,
    DatasetColumn,
    DatasetRow,
    DatasetValue,
)
from ..models.recode import RecodeDefinition, RecodeType, OutputType
from .missing_values import (  # noqa: F401 — _NA_PREFIXES/_is_na re-export (#592 slab 1)
    _NA_PREFIXES,
    _is_na,
    is_missing,
    matched_missing_label,
)

# #691: this module called logger.warning() at two sites with no `logger` in scope,
# so a malformed .sav — the exact case the warnings exist to report — raised
# NameError mid-import instead of degrading gracefully. Both branches are
# user-reachable (scale_values/scale_labels length mismatch; a cells_are_codes
# column with mismatched labels/values). Guarded by tests/test_logger_defined_sweep.py,
# which fails the suite for ANY module that uses `logger.` without defining it —
# the instance was a singleton backend-wide, and the class is what stays closed.
logger = logging.getLogger(__name__)

# ── Rounding precision constants ─────────────────────────────────────────────
PREVIEW_STATS_PRECISION = 1  # round(x, 1) for import preview statistics


# ═══════════════════════════════════════════════════════════════════════════════
# Known Scale Library
# ═══════════════════════════════════════════════════════════════════════════════

"""
KNOWN_SCALES — Expanded scale library for survey auto-detection.
Sources: Vagias (2006), Brown (2010).

Each entry:
  - name: unique key, used for display and alphabetical tiebreaker
  - labels: ordered list, low-to-high (1 = first label, N = last label)
  - canonical: True for the most standard version of each construct type

Matching rules (in priority order):
  1. Case-insensitive subset: all unique substantive values must be in the scale
  2. Minimum 2 unique substantive values required
  3. Minimum 50% coverage: data values must cover >= 50% of scale labels
  4. Tightest fit: fewest labels wins
  5. Best coverage: highest % of labels present in data wins
  6. Canonical preference: canonical=True wins over canonical=False
  7. Alphabetical tiebreaker on name
"""

KNOWN_SCALES: list[dict] = [
    # ── Agreement ────────────────────────────────────────────────────────
    {
        "name": "agreement-2pt",
        "labels": ["Disagree", "Agree"],
        "canonical": False,
    },
    {
        "name": "agreement-3pt",
        "labels": ["Disagree", "Undecided", "Agree"],
        "canonical": False,
    },
    {
        "name": "agreement-4pt",
        "labels": ["Strongly Disagree", "Disagree", "Agree", "Strongly Agree"],
        "canonical": True,
    },
    {
        "name": "agreement-5pt",
        "labels": [
            "Strongly Disagree", "Disagree",
            "Neither Agree nor Disagree",
            "Agree", "Strongly Agree",
        ],
        "canonical": True,
    },
    {
        "name": "agreement-5pt-undecided",
        "labels": [
            "Strongly Disagree", "Disagree", "Undecided",
            "Agree", "Strongly Agree",
        ],
        "canonical": False,
    },
    {
        "name": "agreement-6pt-degree",
        "labels": [
            "Disagree Strongly", "Disagree Moderately", "Disagree Slightly",
            "Agree Slightly", "Agree Moderately", "Agree Strongly",
        ],
        "canonical": False,
    },
    {
        "name": "agreement-6pt-completeness",
        "labels": [
            "Completely Disagree", "Mostly Disagree", "Slightly Disagree",
            "Slightly Agree", "Mostly Agree", "Completely Agree",
        ],
        "canonical": False,
    },
    {
        "name": "agreement-6pt-strength",
        "labels": [
            "Disagree Strongly", "Disagree", "Slightly Disagree",
            "Slightly Agree", "Agree", "Agree Strongly",
        ],
        "canonical": False,
    },
    {
        "name": "agreement-6pt-very-strongly",
        "labels": [
            "Disagree Very Strongly", "Disagree Strongly", "Disagree",
            "Agree", "Agree Strongly", "Agree Very Strongly",
        ],
        "canonical": False,
    },
    {
        "name": "agreement-7pt",
        "labels": [
            "Strongly Disagree", "Disagree", "Somewhat Disagree",
            "Neither Agree nor Disagree",
            "Somewhat Agree", "Agree", "Strongly Agree",
        ],
        "canonical": True,
    },

    # ── Satisfaction ─────────────────────────────────────────────────────
    {
        "name": "satisfaction-5pt",
        "labels": [
            "Very Dissatisfied", "Dissatisfied", "Neutral",
            "Satisfied", "Very Satisfied",
        ],
        "canonical": True,
    },
    {
        "name": "satisfaction-5pt-neither",
        "labels": [
            "Very Dissatisfied", "Dissatisfied",
            "Neither Satisfied nor Dissatisfied",
            "Satisfied", "Very Satisfied",
        ],
        "canonical": False,
    },
    {
        "name": "satisfaction-5pt-degree",
        "labels": [
            "Not at All Satisfied", "Slightly Satisfied",
            "Moderately Satisfied", "Very Satisfied",
            "Extremely Satisfied",
        ],
        "canonical": False,
    },
    {
        "name": "satisfaction-7pt",
        "labels": [
            "Completely Dissatisfied", "Mostly Dissatisfied",
            "Somewhat Dissatisfied",
            "Neither Satisfied nor Dissatisfied",
            "Somewhat Satisfied", "Mostly Satisfied",
            "Completely Satisfied",
        ],
        "canonical": True,
    },
    {
        "name": "satisfaction-7pt-moderately",
        "labels": [
            "Very Dissatisfied", "Moderately Dissatisfied",
            "Slightly Dissatisfied", "Neutral",
            "Slightly Satisfied", "Moderately Satisfied",
            "Very Satisfied",
        ],
        "canonical": False,
    },

    # ── Quality ──────────────────────────────────────────────────────────
    {
        "name": "quality-3pt",
        "labels": ["Poor", "Fair", "Good"],
        "canonical": False,
    },
    {
        "name": "quality-4pt",
        "labels": ["Very Poor", "Poor", "Good", "Very Good"],
        "canonical": False,
    },
    {
        "name": "quality-4pt-acceptable",
        "labels": ["Very Poor", "Poor", "Acceptable", "Very Good"],
        "canonical": False,
    },
    {
        "name": "quality-5pt",
        "labels": ["Poor", "Fair", "Good", "Very Good", "Excellent"],
        "canonical": True,
    },
    {
        "name": "quality-5pt-acceptable",
        "labels": [
            "Very Poor", "Poor", "Acceptable", "Good", "Very Good",
        ],
        "canonical": False,
    },
    {
        "name": "quality-5pt-average",
        "labels": [
            "Very Poor", "Below Average", "Average",
            "Above Average", "Excellent",
        ],
        "canonical": False,
    },
    {
        "name": "quality-5pt-very",
        "labels": ["Very Poor", "Poor", "Fair", "Good", "Very Good"],
        "canonical": False,
    },
    {
        "name": "quality-7pt",
        "labels": [
            "Very Poor", "Poor", "Fair", "Good",
            "Very Good", "Excellent", "Exceptional",
        ],
        "canonical": False,
    },

    # ── Frequency ────────────────────────────────────────────────────────
    {
        "name": "frequency-4pt",
        "labels": ["Never", "Rarely", "Sometimes", "Often"],
        "canonical": True,
    },
    {
        "name": "frequency-4pt-seldom",
        "labels": ["Never", "Seldom", "Some of the Time", "Most of the Time"],
        "canonical": False,
    },
    {
        "name": "frequency-5pt",
        "labels": ["Never", "Rarely", "Sometimes", "Often", "Always"],
        "canonical": True,
    },
    {
        "name": "frequency-5pt-seldom",
        "labels": [
            "Never", "Seldom", "About Half the Time",
            "Usually", "Always",
        ],
        "canonical": False,
    },
    {
        "name": "frequency-5pt-very-often",
        "labels": ["Never", "Rarely", "Sometimes", "Very Often", "Always"],
        "canonical": False,
    },
    {
        "name": "frequency-5pt-almost",
        "labels": [
            "Never", "Almost Never", "Occasionally",
            "Almost Every Time", "Every Time",
        ],
        "canonical": False,
    },
    {
        "name": "frequency-5pt-great-deal",
        "labels": [
            "Never", "Rarely", "Occasionally",
            "A Moderate Amount", "A Great Deal",
        ],
        "canonical": False,
    },
    {
        "name": "frequency-6pt-very",
        "labels": [
            "Never", "Very Rarely", "Rarely",
            "Occasionally", "Frequently", "Very Frequently",
        ],
        "canonical": False,
    },

    # ── Likelihood ───────────────────────────────────────────────────────
    {
        "name": "likelihood-3pt",
        "labels": ["Not Likely", "Somewhat Likely", "Very Likely"],
        "canonical": False,
    },
    {
        "name": "likelihood-4pt",
        "labels": [
            "Definitely Won't", "Probably Won't",
            "Probably Will", "Definitely Will",
        ],
        "canonical": False,
    },
    {
        "name": "likelihood-5pt",
        "labels": [
            "Extremely Unlikely", "Unlikely", "Neutral",
            "Likely", "Extremely Likely",
        ],
        "canonical": True,
    },
    {
        "name": "likelihood-6pt",
        "labels": [
            "Definitely Not", "Probably Not", "Possibly",
            "Probably", "Very Probably", "Definitely",
        ],
        "canonical": False,
    },

    # ── Importance ───────────────────────────────────────────────────────
    {
        "name": "importance-3pt",
        "labels": ["Not Important", "Moderately Important", "Very Important"],
        "canonical": False,
    },
    {
        "name": "importance-5pt",
        "labels": [
            "Not Important", "Slightly Important",
            "Moderately Important", "Very Important",
            "Extremely Important",
        ],
        "canonical": True,
    },
    {
        "name": "importance-5pt-not-at-all",
        "labels": [
            "Not at All Important", "Slightly Important",
            "Moderately Important", "Very Important",
            "Extremely Important",
        ],
        "canonical": False,
    },
    {
        "name": "importance-5pt-fairly",
        "labels": [
            "Not Important", "Slightly Important",
            "Fairly Important", "Important", "Very Important",
        ],
        "canonical": False,
    },
    {
        "name": "importance-5pt-essential",
        "labels": [
            "Not at All Important", "Of Little Importance",
            "Of Average Importance", "Very Important",
            "Absolutely Essential",
        ],
        "canonical": False,
    },
    {
        "name": "importance-7pt",
        "labels": [
            "Not at All Important", "Low Importance",
            "Slightly Important", "Neutral",
            "Moderately Important", "Very Important",
            "Extremely Important",
        ],
        "canonical": True,
    },

    # ── Priority ─────────────────────────────────────────────────────────
    {
        "name": "priority-5pt",
        "labels": [
            "Not a Priority", "Low Priority", "Medium Priority",
            "High Priority", "Essential",
        ],
        "canonical": True,
    },
    {
        "name": "priority-7pt",
        "labels": [
            "Not a Priority", "Low Priority", "Somewhat Priority",
            "Neutral", "Moderate Priority", "High Priority",
            "Essential Priority",
        ],
        "canonical": False,
    },

    # ── Effectiveness ────────────────────────────────────────────────────
    {
        "name": "effectiveness-5pt",
        "labels": [
            "Not Effective", "Slightly Effective",
            "Moderately Effective", "Very Effective",
            "Extremely Effective",
        ],
        "canonical": True,
    },

    # ── Familiarity ──────────────────────────────────────────────────────
    {
        "name": "familiarity-5pt",
        "labels": [
            "Not at All Familiar", "Slightly Familiar",
            "Somewhat Familiar", "Moderately Familiar",
            "Extremely Familiar",
        ],
        "canonical": True,
    },

    # ── Awareness ────────────────────────────────────────────────────────
    {
        "name": "awareness-5pt",
        "labels": [
            "Not at All Aware", "Slightly Aware",
            "Somewhat Aware", "Moderately Aware",
            "Extremely Aware",
        ],
        "canonical": True,
    },

    # ── Concern ──────────────────────────────────────────────────────────
    {
        "name": "concern-5pt",
        "labels": [
            "Not at All Concerned", "Slightly Concerned",
            "Somewhat Concerned", "Moderately Concerned",
            "Extremely Concerned",
        ],
        "canonical": True,
    },

    # ── Influence ────────────────────────────────────────────────────────
    {
        "name": "influence-5pt",
        "labels": [
            "Not at All Influential", "Slightly Influential",
            "Somewhat Influential", "Very Influential",
            "Extremely Influential",
        ],
        "canonical": True,
    },

    # ── Difficulty ───────────────────────────────────────────────────────
    {
        "name": "difficulty-5pt",
        "labels": [
            "Very Difficult", "Difficult", "Neutral",
            "Easy", "Very Easy",
        ],
        "canonical": True,
    },

    # ── Acceptability ────────────────────────────────────────────────────
    {
        "name": "acceptability-7pt",
        "labels": [
            "Totally Unacceptable", "Unacceptable",
            "Slightly Unacceptable", "Neutral",
            "Slightly Acceptable", "Acceptable",
            "Perfectly Acceptable",
        ],
        "canonical": True,
    },

    # ── Appropriateness ──────────────────────────────────────────────────
    {
        "name": "appropriateness-7pt",
        "labels": [
            "Absolutely Inappropriate", "Inappropriate",
            "Slightly Inappropriate", "Neutral",
            "Slightly Appropriate", "Appropriate",
            "Absolutely Appropriate",
        ],
        "canonical": True,
    },

    # ── Comparison ───────────────────────────────────────────────────────
    {
        "name": "comparison-5pt",
        "labels": [
            "Much Worse", "Somewhat Worse", "About the Same",
            "Somewhat Better", "Much Better",
        ],
        "canonical": True,
    },
    {
        "name": "comparison-5pt-higher-lower",
        "labels": [
            "Much Lower", "Lower", "About the Same",
            "Higher", "Much Higher",
        ],
        "canonical": False,
    },
    {
        "name": "comparison-5pt-change",
        "labels": [
            "Much Worse", "Somewhat Worse", "Stayed the Same",
            "Somewhat Better", "Much Better",
        ],
        "canonical": False,
    },

    # ── Expectations ─────────────────────────────────────────────────────
    {
        "name": "expectations-7pt",
        "labels": [
            "Far Below", "Moderately Below", "Slightly Below",
            "Met Expectations",
            "Slightly Above", "Moderately Above", "Far Above",
        ],
        "canonical": True,
    },

    # ── Support / Opposition ─────────────────────────────────────────────
    {
        "name": "support-5pt",
        "labels": [
            "Strongly Oppose", "Somewhat Oppose", "Neutral",
            "Somewhat Favor", "Strongly Favor",
        ],
        "canonical": True,
    },

    # ── Desirability ─────────────────────────────────────────────────────
    {
        "name": "desirability-5pt",
        "labels": [
            "Very Undesirable", "Undesirable", "Neutral",
            "Desirable", "Very Desirable",
        ],
        "canonical": True,
    },

    # ── Reflect Me ───────────────────────────────────────────────────────
    {
        "name": "reflect-me-7pt",
        "labels": [
            "Very Untrue of Me", "Untrue of Me",
            "Somewhat Untrue of Me", "Neutral",
            "Somewhat True of Me", "True of Me",
            "Very True of Me",
        ],
        "canonical": True,
    },

    # ── Beliefs ──────────────────────────────────────────────────────────
    {
        "name": "beliefs-7pt",
        "labels": [
            "Very Untrue of What I Believe",
            "Untrue of What I Believe",
            "Somewhat Untrue of What I Believe",
            "Neutral",
            "Somewhat True of What I Believe",
            "True of What I Believe",
            "Very True of What I Believe",
        ],
        "canonical": False,
    },

    # ── Knowledge of Action ──────────────────────────────────────────────
    {
        "name": "knowledge-of-action-7pt",
        "labels": [
            "Never True", "Rarely True",
            "Sometimes but Infrequently True", "Neutral",
            "Sometimes True", "Usually True", "Always True",
        ],
        "canonical": False,
    },

    # ── Truth ────────────────────────────────────────────────────────────
    {
        "name": "truth-7pt",
        "labels": [
            "Almost Never True", "Rarely True", "Usually Not True",
            "Occasionally True", "Often True", "Usually True",
            "Almost Always True",
        ],
        "canonical": False,
    },

    # ── Level / Degree (generic unipolar) ────────────────────────────────
    {
        "name": "level-3pt",
        "labels": ["Low", "Medium", "High"],
        "canonical": True,
    },
    {
        "name": "level-4pt-value",
        "labels": ["None", "Low", "Moderate", "High"],
        "canonical": False,
    },
    {
        "name": "level-5pt",
        "labels": [
            "Very Low", "Below Average", "Average",
            "Above Average", "Very High",
        ],
        "canonical": False,
    },
    {
        "name": "degree-3pt",
        "labels": ["Not at All", "Moderately", "Extremely"],
        "canonical": False,
    },
    {
        "name": "degree-5pt",
        "labels": [
            "Not at All", "Slightly", "Moderately",
            "Very", "Extremely",
        ],
        "canonical": False,
    },
    {
        "name": "extent-4pt",
        "labels": [
            "Not at All", "Very Little", "Somewhat",
            "To a Great Extent",
        ],
        "canonical": False,
    },

    # ── Problem Severity ─────────────────────────────────────────────────
    {
        "name": "problem-4pt",
        "labels": [
            "Not at All a Problem", "Minor Problem",
            "Moderate Problem", "Serious Problem",
        ],
        "canonical": True,
    },

    # ── Barriers ─────────────────────────────────────────────────────────
    {
        "name": "barriers-4pt",
        "labels": [
            "Not a Barrier", "Somewhat of a Barrier",
            "Moderate Barrier", "Extreme Barrier",
        ],
        "canonical": True,
    },

    # ── Responsibility ───────────────────────────────────────────────────
    {
        "name": "responsibility-4pt",
        "labels": [
            "Not at All Responsible", "Somewhat Responsible",
            "Mostly Responsible", "Completely Responsible",
        ],
        "canonical": True,
    },

    # ── Probability ──────────────────────────────────────────────────────
    {
        "name": "probability-5pt",
        "labels": [
            "Not Probable", "Somewhat Improbable", "Neutral",
            "Somewhat Probable", "Very Probable",
        ],
        "canonical": True,
    },

    # ── Consideration ────────────────────────────────────────────────────
    {
        "name": "consideration-3pt",
        "labels": [
            "Would Not Consider", "Might or Might Not Consider",
            "Definitely Consider",
        ],
        "canonical": True,
    },

    # ── Balance / Amount ─────────────────────────────────────────────────
    {
        "name": "balance-3pt",
        "labels": ["Too Little", "About Right", "Too Much"],
        "canonical": True,
    },
    {
        "name": "strictness-3pt",
        "labels": ["Too Lenient", "About Right", "Too Strict"],
        "canonical": False,
    },
    {
        "name": "harshness-3pt",
        "labels": ["Too Lenient", "About Right", "Too Harsh"],
        "canonical": False,
    },
    {
        "name": "weight-3pt",
        "labels": ["Too Light", "About Right", "Too Heavy"],
        "canonical": False,
    },
]


# ═══════════════════════════════════════════════════════════════════════════════
# Internal Helpers
# ═══════════════════════════════════════════════════════════════════════════════


def _strip_bom(text: str) -> str:
    """Remove UTF-8 BOM if present."""
    return text.lstrip("\ufeff")


# -- N/A detection ------------------------------------------------------------
# #592 slab 1: _NA_PREFIXES/_is_na MOVED to services/missing_values.py — the
# declared-missing predicate module, where they are the DEFAULT rule set for
# columns with no declaration. Re-exported here under the old names (imported
# at the top of this file) so the many existing importers (grouping,
# code_analysis, computed_columns, data_quality, export_r, …) are unchanged;
# slab 2 migrates call sites to the column-aware predicate.


# -- LimeSurvey header parsing ------------------------------------------------

_LS_QUESTION_RE = re.compile(r"^([A-Z]\d{2}[A-Z]\d{2})\.\s*(.+)$")
_CODE_DOT_TEXT_RE = re.compile(r"^(\S+)\.\s+(.+)$")


def parse_header(header: str) -> dict:
    """
    Parse a LimeSurvey-style header into structured parts.

    Returns dict with column_code, group_code, column_text, raw_code.
    """
    header = header.strip()
    m = _LS_QUESTION_RE.match(header)
    if m:
        code = m.group(1)
        group = code.split("Q")[0] if "Q" in code else None
        return {
            "column_code": code,
            "group_code": group,
            "column_text": m.group(2).strip(),
            "raw_code": code,
        }
    m = _CODE_DOT_TEXT_RE.match(header)
    if m:
        return {
            "column_code": None,
            "group_code": None,
            "column_text": m.group(2).strip(),
            "raw_code": m.group(1),
        }
    return {
        "column_code": None,
        "group_code": None,
        "column_text": header,
        "raw_code": None,
    }


# -- Name-like heuristic -------------------------------------------------------

_GENERIC_CODE_RE = re.compile(r'^[A-Z]*\d+$')
_LIMESURVEY_CODE_RE = re.compile(r'^[A-Z]\d{2}[A-Z]\d{2}$')


def _is_name_like(code: str | None) -> bool:
    """Check if a parsed raw_code looks like a meaningful column name (not a generic code)."""
    if not code:
        return False
    if len(code) <= 3:
        return False
    if _GENERIC_CODE_RE.match(code):
        return False
    if _LIMESURVEY_CODE_RE.match(code):
        return False
    if not any(c.isalpha() for c in code):
        return False
    return True


# -- Skip-column detection ----------------------------------------------------

_SKIP_CODES = {
    "id", "submitdate", "lastpage", "startlanguage", "seed",
    "startdate", "datestamp", "ipaddr", "referurl", "token", "optout",
}

_SKIP_HEADERS = {
    "response id", "respondent", "respondent id", "last page",
    "start language", "ip address", "referring url",
}

_SKIP_SUBSTRINGS = [
    "date submitted", "date started", "date last action",
]


def _is_skip_column(header: str, raw_code: str | None) -> bool:
    """Check if a column header looks like survey platform metadata."""
    lower = header.strip().lower()
    if lower in _SKIP_CODES or lower in _SKIP_HEADERS:
        return True
    if raw_code and raw_code.strip().lower() in _SKIP_CODES:
        return True
    return any(sub in lower for sub in _SKIP_SUBSTRINGS)


# -- Identifier detection (#414) -----------------------------------------------
#
# Participant/row identity codes (P001, R-17, respondent names). Header-hint-
# gated (scoping DEC-9): value shape alone must never trigger — a near-unique
# numeric measure is not an ID. Runs BEFORE the skip check in
# `_detect_column_type` because the skip lists swallow id-family headers
# ("id", "respondent id"), silently discarding the identity column.
#
# Two keyword tiers:
#   strong — the header names a PERSON concept; trusted even for 1..N values
#   weak   — bare id-words; demoted back to skip when the values are just a
#            sequential row counter (LimeSurvey's `id` column)
# "response" is a negative signal: a "Response ID" is a platform response key,
# not a person (bare camelCase `ResponseId` never matches — no word boundary).

_IDENTIFIER_STRONG_RE = re.compile(
    r"\b(?:participant|respondent|subject|pid)\b", re.IGNORECASE,
)
_IDENTIFIER_WEAK_RE = re.compile(
    r"\b(?:id|ids|uid|identifier)\b", re.IGNORECASE,
)
_IDENTIFIER_NEGATIVE_RE = re.compile(r"\bresponse\b", re.IGNORECASE)

IDENTIFIER_MIN_UNIQUENESS_RATIO = 0.95  # identity values are (near-)unique per row
IDENTIFIER_MAX_AVG_LEN = 40             # codes are short; prose runs long
IDENTIFIER_MAX_AVG_TOKENS = 4           # "Maria Lopez" yes, a sentence no
IDENTIFIER_MIN_SUBSTANTIVE = 3          # too few rows to judge uniqueness


def _normalize_header_words(text: str | None) -> str:
    """Lower + collapse ``_``/``-``/``.`` to spaces so ``\\b`` can fire —
    Python regex treats ``_`` as a word character, so ``\\bid\\b`` never
    matches inside ``participant_id`` (the `_header_signals_percentage`
    lesson)."""
    if not text:
        return ""
    return re.sub(r"[_\-\.]+", " ", text).lower()


def _is_sequential_counter(substantive_set: set[str]) -> bool:
    """True when the values are a dense integer sequence starting at 0/1 —
    a platform row counter, not an identity referenced by other sources."""
    try:
        ints = {int(v) for v in substantive_set}
    except ValueError:
        return False
    return min(ints) in (0, 1) and (max(ints) - min(ints) + 1) == len(ints)


def _is_identifier_column(
    header: str,
    raw_code: str | None,
    substantive_set: set[str],
    substantive_list: list[str],
) -> bool:
    """#414 / DEC-9: header-hint-gated participant-identifier detection."""
    words = _normalize_header_words(header)
    code_words = _normalize_header_words(raw_code)
    if _IDENTIFIER_NEGATIVE_RE.search(words) or _IDENTIFIER_NEGATIVE_RE.search(code_words):
        return False
    strong = bool(
        _IDENTIFIER_STRONG_RE.search(words) or _IDENTIFIER_STRONG_RE.search(code_words)
    )
    weak = bool(
        _IDENTIFIER_WEAK_RE.search(words) or _IDENTIFIER_WEAK_RE.search(code_words)
    )
    if not (strong or weak):
        return False
    n = len(substantive_list)
    if n < IDENTIFIER_MIN_SUBSTANTIVE:
        return False
    unique_count = len(substantive_set)
    if (unique_count / n) < IDENTIFIER_MIN_UNIQUENESS_RATIO:
        return False
    avg_len = sum(len(v) for v in substantive_set) / unique_count
    if avg_len > IDENTIFIER_MAX_AVG_LEN:
        return False
    avg_tokens = sum(len(v.split()) for v in substantive_set) / unique_count
    if avg_tokens > IDENTIFIER_MAX_AVG_TOKENS:
        return False
    # A bare id-word over a dense 1..N counter is platform metadata — keep skip.
    if not strong and _is_sequential_counter(substantive_set):
        return False
    return True


# -- Demographic detection -----------------------------------------------------

_DEMOGRAPHIC_KEYWORDS = {
    "gender", "race", "age", "ethnicity", "role", "sex", "income", "education",
}

_DEMOGRAPHIC_RE = re.compile(
    r"\b(?:" + "|".join(_DEMOGRAPHIC_KEYWORDS) + r")\b", re.IGNORECASE,
)


# -- Percentage header detection -----------------------------------------------
#
# #358: replace the greedy "all integer + 0<=min<=max<=100 + max>=10" rule
# (which captured Tenure, Years_Experience, integer Test_Score as percentage)
# with a stricter "header signal required" rule. Falls back to numeric when
# no `%` glyph and no keyword — researchers can still manually override via
# the dataset import preview's type dropdown.
#
# Keyword list covers common research column-naming vocab. Word boundaries
# match the existing `_DEMOGRAPHIC_RE` precedent so e.g. "rate" doesn't
# match inside "narrate" but does match inside "completion_rate".
_PERCENTAGE_KEYWORDS = {
    "pct", "percent", "percentage", "rate", "share",
    "proficiency", "coverage", "uptake", "participation",
    "compliance", "completion",
}

_PERCENTAGE_KEYWORD_RE = re.compile(
    r"\b(?:" + "|".join(_PERCENTAGE_KEYWORDS) + r")\b", re.IGNORECASE,
)


def _header_signals_percentage(header: str | None) -> bool:
    """Match `_PERCENTAGE_KEYWORD_RE` against a normalized header.

    The naive `\\bpct\\b` against raw `Pct_FRL` doesn't match because
    Python regex `\\b` treats `_` as a word character — there's no
    word-to-non-word transition after `pct`. Real-world percentage
    column names almost always use `_` / `-` separators
    (`Pct_FRL`, `response_rate`, `coverage-2024`), so normalize them
    to spaces first. Letter-to-letter sequences like `narrate` stay
    glued (and correctly do NOT match `rate`).
    """
    if not header:
        return False
    normalized = re.sub(r"[_\-\.]+", " ", header)
    return bool(_PERCENTAGE_KEYWORD_RE.search(normalized))


def _is_demographic(text: str) -> bool:
    """Match short text containing a demographic keyword at a word boundary."""
    if len(text) > 40:
        return False
    return bool(_DEMOGRAPHIC_RE.search(text))


_SUBTYPE_KEYWORDS = {
    "role": {"role", "position", "title", "department"},
    "race": {"race", "ethnicity"},
    "gender": {"gender", "sex"},
    "age": {"age"},
}


def _detect_demographic_subtype(header_text: str) -> str | None:
    """Detect the demographic subtype from the column header text."""
    lower = header_text.lower()
    for subtype, keywords in _SUBTYPE_KEYWORDS.items():
        for kw in keywords:
            if re.search(r'\b' + kw + r'\b', lower):
                return subtype
    return None


# -- Boolean detection ---------------------------------------------------------

_BOOLEAN_PAIRS = [
    {"yes", "no"}, {"true", "false"}, {"1", "0"}, {"y", "n"}, {"t", "f"},
]


def _is_boolean(values: set[str]) -> bool:
    if not values or len(values) > 2:
        return False
    lower = {v.lower() for v in values}
    return any(lower.issubset(pair) for pair in _BOOLEAN_PAIRS)


# -- Numeric helpers -----------------------------------------------------------

_CURRENCY_RE = re.compile(r"[\$\u20ac\u00a3\u00a5]")  # $ € £ ¥
_PERCENT_SUFFIX_RE = re.compile(r"\d\s*%$")


def _strip_numeric(value: str) -> float | None:
    """Strip formatting characters ($, EUR, GBP, %, commas) and parse as float."""
    s = value.strip()
    if not s:
        return None
    cleaned = re.sub(r"[\$\u20ac\u00a3\u00a5,%]", "", s).strip()
    try:
        n = float(cleaned)
        return n if math.isfinite(n) else None
    except (ValueError, OverflowError):
        return None


def _analyze_numeric(values: list[str], header: str | None = None) -> dict | None:
    """
    Analyze values for numeric patterns.

    Returns dict with column_type (ColumnType), numeric_format, numeric_min,
    numeric_max -- or None if values are not all numeric.

    The ``header`` parameter (#358) gates percentage classification: a column
    is only classified as PERCENTAGE when (a) at least one value carries a
    `%` glyph, or (b) the column header matches `_PERCENTAGE_KEYWORD_RE`
    (pct/percent/rate/share/proficiency/coverage/uptake/participation/
    compliance/completion). All other integer columns in [0,100] — including
    years-of-tenure, age ranges, count-of-events, integer test scores —
    fall back to NUMERIC. Researchers can manually override via the
    dataset import preview's type dropdown.
    """
    if not values:
        return None

    nums = []
    has_currency = False
    has_percent = False
    all_integer = True

    for v in values:
        s = v.strip()
        if _CURRENCY_RE.search(s):
            has_currency = True
        if _PERCENT_SUFFIX_RE.search(s):
            has_percent = True
        n = _strip_numeric(s)
        if n is None:
            return None
        nums.append(n)
        if not n.is_integer():
            all_integer = False

    min_val = min(nums)
    max_val = max(nums)

    # Header keyword check (#358). Defensive against None / empty header so
    # direct unit-test callers without a header still get integer/decimal
    # classification correctly.
    header_signals_percentage = _header_signals_percentage(header)

    # Determine format
    if has_currency:
        fmt = "currency"
    elif has_percent:
        fmt = "percentage"
    elif header_signals_percentage:
        fmt = "percentage"
    elif all_integer:
        fmt = "integer"
    else:
        fmt = "decimal"

    qtype = ColumnType.PERCENTAGE if fmt == "percentage" else ColumnType.NUMERIC

    return {
        "column_type": qtype,
        "numeric_format": fmt,
        "numeric_min": min_val,
        "numeric_max": max_val,
    }


# -- Scale matching ------------------------------------------------------------


# A column matches a known scale even when a few of its distinct values aren't in
# the scale, as long as the matched values clearly dominate (#364). This guards
# against BOTH failure modes: (a) a single misspelled Likert label ("Srongly
# Disagree") dropping a clean ordinal column to nominal and forcing the researcher
# to re-type every affected column at import, and (b) a genuinely nominal column
# coincidentally overlapping a scale on one or two labels being mis-typed ordinal.
_SCALE_MAX_UNMATCHED = 2


def _scale_match_within_tolerance(
    matched: set[str], unmatched: set[str],
) -> bool:
    """Whether a column's value set matches a scale despite a few stray values.

    Requires at least one matched label, no more than `_SCALE_MAX_UNMATCHED`
    distinct unmatched values, and matched values to outnumber unmatched at
    least 2:1. With zero unmatched (the old strict-subset case) this is always
    True, so previously-matching columns keep matching.
    """
    if not matched:
        return False
    if len(unmatched) > _SCALE_MAX_UNMATCHED:
        return False
    if len(matched) < 2 * len(unmatched):
        return False
    return True


def _match_scale(values: set[str]) -> tuple[str, list[str]] | None:
    """
    Find the best matching known scale for a set of values.

    Matching rules (in priority order):
      1. Tolerant match: most data values must appear in the scale, allowing a
         small number of stray values (typos) — see `_scale_match_within_tolerance`
      2. Minimum 2 unique substantive values required
      3. Minimum 50% coverage: matched data values must cover >= 50% of scale labels
      4. Tightest fit: fewest labels wins
      5. Best coverage: highest percentage of labels present in data wins
      6. Canonical preference: canonical=True wins over canonical=False
      7. Alphabetical tiebreaker on name

    Returns (scale_name, ordered_labels) or None.
    """
    if not values or len(values) < 2:
        return None
    lower_vals = {v.lower() for v in values}
    matches: list[tuple[dict, float]] = []
    for scale in KNOWN_SCALES:
        lower_labels = {label.lower() for label in scale["labels"]}
        matched = lower_vals & lower_labels
        unmatched = lower_vals - lower_labels
        if not _scale_match_within_tolerance(matched, unmatched):
            continue
        # Coverage is the fraction of the SCALE's labels present in the matched
        # (in-scale) data — stray values don't count toward or against it.
        coverage = len(matched) / len(scale["labels"])
        if coverage >= 0.5:
            matches.append((scale, coverage))
    if not matches:
        return None
    matches.sort(key=lambda x: (
        len(x[0]["labels"]),       # tightest fit (fewest labels)
        -x[1],                     # best coverage (highest %)
        not x[0]["canonical"],     # canonical preference (True first)
        x[0]["name"],              # alphabetical tiebreaker
    ))
    best = matches[0][0]
    return (best["name"], best["labels"])


# -- Numeric value computation for answers -------------------------------------


def _coerce_scale_codes(scale_values: list[float]) -> list[int | float]:
    """Store an integral scale code as an int, so both import paths agree (#28).

    The CSV path derives codes from `range(1, n+1)` and stores `[1, 2, 3]`; the
    .sav path receives them as JSON floats and would store `[1.0, 2.0, 3.0]` for
    the same logical scale. `routers/export_r.py` emits `scale_values` verbatim as
    R factor levels, so the divergence would surface in exported scripts.
    """
    return [int(v) if float(v).is_integer() else float(v) for v in scale_values]


def _compute_value_numeric(
    raw_value: str,
    question_type: str,
    scale_labels: list[str] | None,
    scale_values: list[float] | None = None,
    missing_rules: list | None = None,
) -> float | None:
    """Compute the numeric encoding for a cell value.

    ``scale_values`` (#28) supplies the codes an ordinal scale's labels actually
    carry, parallel to ``scale_labels``. SPSS files know them (a scale may be
    0-based, or skip codes); CSV imports do not and pass None, which keeps the
    historical positional 1..N encoding byte-for-byte. A length mismatch falls
    back to positional rather than silently mis-encoding.

    ``missing_rules`` (#592) is the COLUMN's parsed missing declaration —
    None = undeclared, the recognized-N/A defaults (behavior unchanged for
    every caller that doesn't pass it). A declared column's rules REPLACE the
    defaults, so a declared "99" encodes NULL and a declared-[] column's
    "N/A" encodes as data.
    """
    if is_missing(raw_value, missing_rules):
        return None

    if question_type == ColumnType.ORDINAL.value:
        if scale_labels:
            if scale_values and len(scale_values) == len(scale_labels):
                codes = [float(v) for v in scale_values]
            else:
                codes = [float(i + 1) for i in range(len(scale_labels))]
            label_map = {l.lower(): codes[i] for i, l in enumerate(scale_labels)}
            return label_map.get(raw_value.strip().lower())
        # #580: an ordinal column with NO scale labels (a bare-numeric Likert item
        # the user overrode to ordinal — inference only ever suggests ORDINAL when
        # a known TEXT scale matches) used to return None here, so value_numeric was
        # NULL in every cell and the column silently vanished from every numeric
        # analysis — violating the VALUE_NUMERIC_TYPES/SCALE_SCORE_ELIGIBLE_TYPES
        # contract that ordinal's value_numeric is reliably populated. A bare number
        # IS its own code, so fall back to the numeric parse (identical to how a
        # NUMERIC column encodes the same cell). A non-numeric cell in such a column
        # still yields None, exactly as an out-of-scale label would.
        return _strip_numeric(raw_value)

    if question_type in (ColumnType.NUMERIC.value, ColumnType.PERCENTAGE.value):
        return _strip_numeric(raw_value)

    if question_type == ColumnType.BINARY.value:
        lower = raw_value.strip().lower()
        if lower in ("yes", "true", "1", "y", "t"):
            return 1.0
        if lower in ("no", "false", "0", "n", "f"):
            return 0.0
        return None

    return None


# -- Column type detection -----------------------------------------------------

# #380: high-cardinality categorical detection. A non-numeric column with >10
# distinct values used to fall straight through to open_text, which excluded it
# from analysis (frequency/group-by/cross-tab) and blocked recodes — wrong for
# demographic categoricals like industry sector (18 NAICS labels), geography, or
# detailed ethnicity. We now classify such a column as NOMINAL when it looks like
# a set of repeated short labels rather than free prose. The three signals:
#   - bounded cardinality (a 200-category "variable" is not analytically useful)
#   - low uniqueness ratio (free text is near-unique; labels repeat)
#   - short average label length (labels are short; prose runs long)
# uniqueness ratio is the primary discriminator; avg length is the backstop.
# Tuned against the scenario-4 Family Leave Survey (Industry_Sector: 18 unique,
# ratio 0.045, avg len 16) and a genuine-comment control that must stay open_text.
# #575: a numbers-only column with more distinct values than this is treated as a
# continuous measure, not a labellable scale — the wizard won't seed a code editor.
VALUE_LABEL_SEED_MAX_CODES = 30

NOMINAL_MAX_CARDINALITY = 100        # ceiling — beyond this, default to open_text
NOMINAL_MAX_UNIQUENESS_RATIO = 0.5   # unique/n must be below this (labels repeat)
NOMINAL_MAX_AVG_LABEL_LEN = 30       # avg label length (chars) — prose runs longer


def _looks_like_nominal_labels(substantive_set: set[str], substantive_list: list[str]) -> bool:
    """#380: heuristic for a high-cardinality categorical (repeated short labels)
    vs genuine free text. Caller has already ruled out numeric and <=10-unique."""
    n = len(substantive_list)
    unique_count = len(substantive_set)
    if n == 0 or unique_count == 0:
        return False
    if unique_count > NOMINAL_MAX_CARDINALITY:
        return False
    if (unique_count / n) >= NOMINAL_MAX_UNIQUENESS_RATIO:
        return False
    avg_label_len = sum(len(v) for v in substantive_set) / unique_count
    return avg_label_len <= NOMINAL_MAX_AVG_LABEL_LEN


def _detect_column_type(
    header: str,
    parsed: dict,
    substantive_set: set[str],
    substantive_list: list[str],
    col_idx: int,
) -> dict:
    """
    Auto-detect the suggested type for a CSV column.

    Returns a dict with suggested_type, scale info, and numeric metadata.
    """
    result: dict = {
        "suggested_type": ColumnType.OPEN_TEXT.value,
        "suggested_scale_name": None,
        "suggested_scale_labels": None,
        # #28: only a format that KNOWS its scale codes fills this (SPSS .sav, via
        # sav_import.apply_sav_metadata). Inference over CSV text never can, so
        # None here means "positional 1..N" downstream.
        "suggested_scale_values": None,
        "suggested_scale_unmatched": None,
        "suggested_demographic_subtype": None,
        "numeric_format": None,
        "numeric_min": None,
        "numeric_max": None,
    }

    # 0. Identifier (#414) — MUST run before skip: the skip lists swallow
    # id-family headers ("id", "respondent id"), discarding the identity column.
    if _is_identifier_column(header, parsed["raw_code"], substantive_set, substantive_list):
        result["suggested_type"] = ColumnType.IDENTIFIER.value
        return result

    # 1. Skip (platform metadata)
    if _is_skip_column(header, parsed["raw_code"]):
        result["suggested_type"] = ColumnType.SKIP.value
        return result

    # 2. Demographic (short headers only — check parsed question text, not raw header)
    if _is_demographic(parsed["column_text"]):
        result["suggested_type"] = ColumnType.DEMOGRAPHIC.value
        result["suggested_demographic_subtype"] = _detect_demographic_subtype(parsed["column_text"])
        return result

    if not substantive_set:
        return result  # defaults to open_text

    # 3. Binary
    if _is_boolean(substantive_set):
        result["suggested_type"] = ColumnType.BINARY.value
        return result

    # 4. Small cardinality (<=10 unique): scale first, then numeric, then nominal
    if len(substantive_set) <= 10:
        match = _match_scale(substantive_set)
        if match:
            result["suggested_type"] = ColumnType.ORDINAL.value
            result["suggested_scale_name"] = match[0]
            result["suggested_scale_labels"] = match[1]
            # Surface any values not in the matched scale (#364). These import
            # with value_numeric=None (blank) — the researcher should review them
            # as likely typos. Preserve original casing + first-seen order.
            label_lower = {l.lower() for l in match[1]}
            unmatched = [v for v in substantive_list if v.lower() not in label_lower]
            seen: set[str] = set()
            unmatched_unique = [
                v for v in unmatched if not (v.lower() in seen or seen.add(v.lower()))
            ]
            result["suggested_scale_unmatched"] = unmatched_unique or None
            return result

        # #358: pass header so the percentage keyword check can fire
        numeric = _analyze_numeric(list(substantive_set), header=header)
        if numeric:
            result["suggested_type"] = numeric["column_type"].value
            result["numeric_format"] = numeric["numeric_format"]
            result["numeric_min"] = numeric["numeric_min"]
            result["numeric_max"] = numeric["numeric_max"]
            return result

        result["suggested_type"] = ColumnType.NOMINAL.value
        return result

    # 5. High cardinality (>10 unique)
    numeric = _analyze_numeric(list(substantive_set), header=header)  # #358
    if numeric:
        result["suggested_type"] = numeric["column_type"].value
        result["numeric_format"] = numeric["numeric_format"]
        result["numeric_min"] = numeric["numeric_min"]
        result["numeric_max"] = numeric["numeric_max"]
        return result

    # 5b. High-cardinality categorical (#380): repeated short labels, not prose
    if _looks_like_nominal_labels(substantive_set, substantive_list):
        result["suggested_type"] = ColumnType.NOMINAL.value
        return result

    # 6. Open text
    result["suggested_type"] = ColumnType.OPEN_TEXT.value
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# Excel (.xlsx) adapter (#523)
# ═══════════════════════════════════════════════════════════════════════════════
#
# .xlsx support is a format ADAPTER: the workbook is converted to CSV text at the
# router boundary and everything downstream (type inference, N/A handling, import)
# runs the existing CSV pipeline unchanged. Keep it that way — new formats should
# adapt into CSV text here, never fork the inference/import code paths.

# Structural caps: a .xlsx is a ZIP, so a small upload can inflate enormously.
# These bound the parse work independently of the 50 MB upload cap.
# ── The real size gate: CELLS (#799/#803) ────────────────────────────────────
# The byte and dimension caps beside this one are cheap PRE-FILTERS; neither
# bounds what an import actually costs.
#
#   * BYTES vary ~4x by format — a compressed .xlsx expands into roughly four
#     times its size in CSV — so the same 50 MB budget buys wildly different
#     work depending on which file the researcher happens to have.
#   * DIMENSIONS MULTIPLY. 100,000 rows and 500 columns are each defensible on
#     their own and authorise **50,000,000 cells** together — 16x the file that
#     already exceeded every memory target in this codebase.
#
# What costs time and memory is CELLS, and it is linear in them: MEASURED at
# 23.4 / 23.6 / 24.0 s per million cells across 410K / 1.03M / 2.05M-cell
# imports of the same real file.
#
# 4,000,000 is set ABOVE the largest real dataset this has been driven against
# (GSS: 75,699 x 41 = 3,103,659) on purpose. Sizing the cap to what fits the
# <256 MB resident target would put it near 2M cells and REFUSE an ordinary
# research dataset, which is the tool declining real work.
#
# ⚠️ **The memory budget is two numbers, not one, and this is the deliberate
# part.** `<256 MB` is a RESIDENT target — a steady-state property of a server
# answering requests, and the paginated grid honours it (96 MB per page, down
# from 5,877 MB). An import is a one-off TRANSIENT: measured at 297 MB (CSV) and
# 346 MB (.xlsx) for 3.1M cells, so ~450 MB at this cap. That allowance is
# chosen and stated here rather than discovered later.
MAX_DATASET_CELLS = 4_000_000


class DatasetTooLargeError(ValueError):
    """Refused: over `MAX_DATASET_CELLS` (#803).

    ⚠️ A DISTINCT type, not a bare `ValueError`, because the preview endpoint
    catches `(ValueError, csv.Error, TypeError)` and replaces it with "Unable to
    parse CSV file. Check the file format and try again." — a diagnosis it has
    not established, about a file that parses perfectly well. That is the #797
    defect exactly, and a shared exception type is how it would have recurred.
    The router catches this one FIRST and shows its message verbatim.
    """


def cell_cap_exceeded_message(n_cols: int) -> str:
    """The refusal for a STREAMING path, which bails before it has counted.

    ⚠️ Deliberately does NOT quote a row total. The caller stops the moment the
    cap is crossed, so it does not know how many rows the file has — and a
    message naming the count at the point of the bail would state a number that
    is simply wrong, which is the #797 lesson (report what you know, never a
    plausible-looking guess).
    """
    return (
        f"This dataset is over the {MAX_DATASET_CELLS:,}-value limit at "
        f"{n_cols:,} columns. Importing fewer columns — the wizard can skip any "
        "you don't need — or splitting the file by rows will bring it under."
    )


def cell_count_error(n_rows: int, n_cols: int) -> str | None:
    """The refusal message for an over-cap dataset, or None if it fits.

    ONE function so CSV, .xlsx and .sav refuse at the same size for the same
    reason — the three formats had three different limits expressed in three
    different units, and none of them was cells.
    """
    cells = n_rows * n_cols
    if cells <= MAX_DATASET_CELLS:
        return None
    return (
        f"This dataset is {n_rows:,} rows x {n_cols:,} columns = {cells:,} values, "
        f"over the {MAX_DATASET_CELLS:,} limit. Importing fewer columns — the "
        "wizard can skip any you don't need — or splitting the file by rows will "
        "bring it under."
    )


MAX_XLSX_ROWS = 100_000
MAX_XLSX_COLS = 500

XLSX_MAGIC = b"PK\x03\x04"  # xlsx files are ZIP containers


class XlsxImportError(ValueError):
    """User-facing .xlsx parse/validation failure (surfaced as HTTP 400)."""


def is_xlsx_upload(filename: str | None, content: bytes) -> bool:
    """True when the upload should take the .xlsx adapter path.

    Requires BOTH the extension and the ZIP magic — a mis-renamed CSV falls
    through to the text path (where it may still parse), and a renamed
    non-zip binary fails fast instead of confusing openpyxl.
    """
    return bool(filename) and filename.lower().endswith(".xlsx") and content[:4] == XLSX_MAGIC


def _xlsx_cell_to_str(value) -> str:
    """Stringify a cell the way Excel's own save-as-CSV would (CSV parity).

    Order matters: bool is a subclass of int, so it must be checked first.
    """
    import datetime as _dt

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        # openpyxl yields 3.0 for a typed 3 — trim so value_text matches the CSV twin.
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return str(value)
    if isinstance(value, _dt.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (_dt.date, _dt.time)):
        return value.isoformat()
    return str(value)


def xlsx_to_csv_text(content: bytes, sheet_name: str | None = None) -> tuple[str, list[str]]:
    """Convert one worksheet of a .xlsx upload into CSV text.

    Returns (csv_text, sheet_names). ``sheet_name`` None selects the first sheet.
    Formula cells carry the file's cached computed value (``data_only=True``); a
    workbook saved without computed caches yields blanks for them.

    Raises XlsxImportError for anything the user should fix (bad zip, unknown
    sheet, empty sheet, over-cap dimensions).
    """
    import io as _io
    import zipfile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError, OSError) as e:
        raise XlsxImportError(f"Unable to read the Excel file: {e}") from e

    try:
        sheet_names = list(wb.sheetnames)
        if not sheet_names:
            raise XlsxImportError("The Excel workbook contains no worksheets.")
        target = sheet_name or sheet_names[0]
        if target not in sheet_names:
            raise XlsxImportError(f'Worksheet "{target}" was not found in the workbook.')
        ws = wb[target]

        # #803: refuse on the sheet's DECLARED dimensions, before any cell is
        # read — an over-cap workbook must not cost the memory it is being
        # refused for. openpyxl's max_row/max_column can OVERCOUNT (formatting
        # residue, trimmed later), so this only ever refuses what is genuinely
        # over; the authoritative check runs on the trimmed dimensions below.
        declared = cell_count_error(ws.max_row or 0, ws.max_column or 0)
        if declared:
            raise XlsxImportError(declared)

        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_XLSX_ROWS:
                raise XlsxImportError(
                    f"The worksheet has more than {MAX_XLSX_ROWS:,} rows. "
                    "Split the data into smaller files and import them separately."
                )
            if len(row) > MAX_XLSX_COLS:
                raise XlsxImportError(
                    f"The worksheet has more than {MAX_XLSX_COLS} columns."
                )
            rows.append([_xlsx_cell_to_str(v) for v in row])
    finally:
        wb.close()

    # Excel sheets often report phantom trailing rows/columns (formatting residue).
    # Trim fully-empty trailing rows, then size every row to the header's width.
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    if not rows:
        raise XlsxImportError(f'Worksheet "{target}" has no data.')

    header = rows[0]
    while header and header[-1] == "":
        header.pop()
    if not header:
        raise XlsxImportError(f'Worksheet "{target}" has no header row.')
    width = len(header)

    # #803: the authoritative check, on the TRIMMED dimensions. The pre-read
    # check above uses openpyxl's declared size, which can overcount.
    trimmed = cell_count_error(len(rows) - 1, width)
    if trimmed:
        raise XlsxImportError(trimmed)

    out = _io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(header)
    for row in rows[1:]:
        sized = row[:width] + [""] * (width - len(row[:width]))
        writer.writerow(sized)

    return out.getvalue(), sheet_names


def preview_dataset_csv(
    file_contents: str,
    missing_rules_by_column: dict[str, list] | None = None,
) -> dict:
    """
    Parse a survey CSV and return per-column analysis with auto-detected types.

    Args:
        file_contents: The CSV file as a decoded string (BOM handled internally).
        missing_rules_by_column: #592 slab 5 — declared missing rules keyed by
            HEADER name, known before import only for formats that carry their
            own declaration (``.sav``). Columns absent from the map fall back to
            the recognized-N/A defaults, which is every CSV/XLSX column: at
            preview time no DatasetColumn exists, so there is nothing to declare
            on yet. This is a PRE-pass, not an overlay: type detection,
            ``na_count`` and ``numeric_min``/``max`` all consume the substantive
            set, so a post-hoc fix cannot reach them. Without it, preserving
            .sav's user-missing codes (#596) makes a "Refused" cell read as real
            text and flips `suggested_type` ordinal→nominal — and for a
            non-ordinal column nothing downstream corrects it, so the flip
            persists into the imported column.

    Returns:
        Dict with ``total_rows`` and ``columns`` list.  Each column entry
        contains: column_name, column_index, sample_values, unique_count,
        empty_count, empty_percent, na_count, all_numeric, avg_text_length,
        suggested_type, suggested_scale_name, suggested_scale_labels,
        suggested_column_code, suggested_group_code, suggested_column_text,
        numeric_format, numeric_min, numeric_max.
    """
    text = _strip_bom(file_contents)
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []

    # Collect all values per column
    col_all_values: dict[str, list[str]] = {h: [] for h in headers}
    total_rows = 0

    # #803: a plain .csv declares no dimensions, so the cap can only be applied
    # while reading. Bail the MOMENT it is crossed rather than after the count —
    # accumulating an over-cap file into `col_all_values` would spend exactly the
    # memory the cap exists to refuse.
    n_cols = len(headers)
    max_rows_for_cap = MAX_DATASET_CELLS // n_cols if n_cols else None

    for row in reader:
        total_rows += 1
        if max_rows_for_cap is not None and total_rows > max_rows_for_cap:
            raise DatasetTooLargeError(cell_cap_exceeded_message(n_cols))
        for h in headers:
            col_all_values[h].append(row.get(h, "").strip())

    columns = []
    for col_idx, header in enumerate(headers):
        all_vals = col_all_values[header]
        non_empty = [v for v in all_vals if v]
        empty_count = len(all_vals) - len(non_empty)

        # Unique values preserving first-seen order
        unique_ordered = list(dict.fromkeys(non_empty))

        # Substantive = non-empty AND non-missing (drives type detection,
        # na_count, and the numeric min/max below).
        # #592 slab 5: column-aware when the FORMAT carried a declaration
        # (.sav's user-missing), else the recognized-N/A defaults — which is
        # every text-format column, since no DatasetColumn exists to declare on
        # until import. This is the one remaining bare-`_is_na` site and it is
        # allowlisted in the fail-closed scan for exactly that reason.
        preview_rules = (missing_rules_by_column or {}).get(header)
        substantive_list = [v for v in non_empty if not is_missing(v, preview_rules)]
        substantive_set = set(substantive_list)
        na_count = len(non_empty) - len(substantive_list)

        # Stats
        sample_values = unique_ordered[:5]
        unique_count = len(set(non_empty))
        empty_percent = (
            round(empty_count / total_rows * 100, PREVIEW_STATS_PRECISION) if total_rows else 0.0
        )
        all_numeric = bool(substantive_set) and all(
            _strip_numeric(v) is not None for v in substantive_set
        )
        avg_text_length = (
            round(sum(len(v) for v in non_empty) / len(non_empty), PREVIEW_STATS_PRECISION)
            if non_empty
            else 0.0
        )

        # #575: the complete sorted distinct code set for a likely scale (all
        # numeric + bounded cardinality), so the wizard's value-labels editor can
        # seed every code, not just the 5 sample_values. Skip continuous measures.
        distinct_numeric_values = None
        if all_numeric and unique_count <= VALUE_LABEL_SEED_MAX_CODES:
            parsed_codes = {_strip_numeric(v) for v in substantive_set}
            distinct_numeric_values = sorted(c for c in parsed_codes if c is not None)

        # Parse header
        parsed = parse_header(header)

        # Detect type
        detection = _detect_column_type(
            header, parsed, substantive_set, substantive_list, col_idx,
        )

        columns.append({
            "column_name": header,
            "column_index": col_idx,
            "sample_values": sample_values,
            "unique_count": unique_count,
            "empty_count": empty_count,
            "empty_percent": empty_percent,
            "na_count": na_count,
            "all_numeric": all_numeric,
            "avg_text_length": avg_text_length,
            "suggested_type": detection["suggested_type"],
            "suggested_scale_name": detection["suggested_scale_name"],
            "suggested_scale_labels": detection["suggested_scale_labels"],
            "suggested_scale_values": detection["suggested_scale_values"],
            "suggested_scale_unmatched": detection["suggested_scale_unmatched"],
            "distinct_numeric_values": distinct_numeric_values,
            "suggested_column_code": parsed["column_code"],
            "suggested_group_code": parsed["group_code"],
            "suggested_column_text": parsed["column_text"],
            "suggested_column_name": parsed["raw_code"] if _is_name_like(parsed.get("raw_code")) else None,
            "suggested_demographic_subtype": detection.get("suggested_demographic_subtype"),
            "numeric_format": detection["numeric_format"],
            "numeric_min": detection["numeric_min"],
            "numeric_max": detection["numeric_max"],
        })

    return {"total_rows": total_rows, "columns": columns}


def _scan_source_rows(text: str, column_configs: list[dict]) -> tuple[int, dict, dict]:
    """ONE streaming pass over the CSV, for everything the import needs to know
    about the data BEFORE it writes anything (#799).

    Returns ``(row_count, distinct_numeric, na_values)``.

    ⚠️ **This replaces `data_rows = list(reader)`, and the reason is memory:**
    MEASURED on a real GSS extract (75,699 x 41), that list materialised
    3,103,700 Python `str` objects and took peak RSS from 223 MB to **511 MB** —
    twice the <256 MB backend target, for a file well inside the 50 MB upload
    cap.

    ⚠️ **It is also FASTER, which the naive fix would not have been.** The two
    scans it replaces lived INSIDE the per-column loop, so the row list was
    walked once per qualifying column — 4 numeric columns meant 4 walks. Simply
    swapping the list for a fresh `csv.reader` each time would have re-parsed a
    43 MB string once per column. Accumulating every column's answer in a single
    pass costs one parse total.

    Both predicates come from the CONFIG, not from the database, so this can run
    before any column exists:

    * numeric/percentage columns need their DISTINCT values — `_analyze_numeric`
      takes `list(set(...))`, so a set is what it actually wanted;
    * ordinal columns with scale labels need the set of cells their effective
      missing rule recognises, to seed the auto recode's exclude channel.
    """
    want_numeric: dict[int, list] = {}
    want_na: dict[int, list] = {}
    for cfg in column_configs:
        if cfg.get("skip"):
            continue
        idx = cfg["column_index"]
        qtype = cfg.get("column_type", "")
        rules = cfg.get("missing_values")
        if qtype in (ColumnType.NUMERIC.value, ColumnType.PERCENTAGE.value):
            want_numeric[idx] = rules
        if (
            qtype == ColumnType.ORDINAL.value
            and cfg.get("scale_labels")
            and not cfg.get("cells_are_codes")
        ):
            want_na[idx] = rules

    distinct_numeric: dict[int, set] = {i: set() for i in want_numeric}
    na_values: dict[int, set] = {i: set() for i in want_na}

    reader = csv.reader(io.StringIO(text))
    next(reader, None)  # header
    row_count = 0
    for row in reader:
        row_count += 1
        n = len(row)
        for idx, rules in want_numeric.items():
            if idx < n:
                cell = row[idx].strip()
                if cell and not is_missing(cell, rules):
                    distinct_numeric[idx].add(cell)
        for idx, rules in want_na.items():
            if idx < n:
                cell = row[idx].strip()
                if cell and is_missing(cell, rules):
                    na_values[idx].add(cell)
    return row_count, distinct_numeric, na_values


def import_dataset_csv(
    db: Session,
    project_id: int,
    name: str,
    column_configs: list[dict],
    file_contents: str,
    description: str | None = None,
    source: str | None = None,
    participant_link_column_index: int | None = None,
) -> dict:
    """
    Import a dataset CSV into the database.

    All writes happen in a single transaction — nothing is committed until
    every object has been created successfully.

    Each row gets a system-generated record identifier (R0001, R0002,
    etc.) based on CSV row order.  Participant linking (#414) runs in the
    same transaction when `participant_link_column_index` names an
    identifier column; otherwise it remains a post-import operation via
    the row-link endpoints / retro link-by-column.

    Args:
        db: SQLAlchemy session.
        project_id: The project to import into.
        name: Display name for the Dataset.
        column_configs: Per-column configuration.  Each dict may contain:
            column_index (int), skip (bool), column_type (str),
            column_text (str), column_code (str|None),
            group_code (str|None), group_label (str|None),
            scale_labels (list[str]|None).
        file_contents: The CSV file as a decoded string.
        description: Optional description.
        source: Optional source platform name (e.g. "LimeSurvey").

    Returns:
        Summary dict: dataset_id, columns_created, rows_created,
        values_created, recognized_missing_*, participant_link_report
        (None unless linking ran).
    """
    text = _strip_bom(file_contents)
    headers = next(csv.reader(io.StringIO(text)))
    # #799: ONE streaming pass instead of a retained row list — see
    # `_scan_source_rows`. The list cost 288 MB on a real import and was walked
    # once per qualifying column.
    row_count, distinct_numeric_by_idx, na_values_by_idx = _scan_source_rows(
        text, column_configs,
    )
    # #803: the cap is enforced on the OPERATION, not only at the wizard. The
    # preview endpoint refuses first and more cheaply, but scripts and direct API
    # callers never pass it — the #589 lesson, restated for size.
    _over = cell_count_error(row_count, len(headers))
    if _over:
        raise DatasetTooLargeError(_over)

    # Build config lookup by column index
    cfg_by_idx: dict[int, dict] = {cfg["column_index"]: cfg for cfg in column_configs}

    # Auto-ID padding: len(str(row_count)) + 2 extra zeros
    pad_width = len(str(row_count)) + 2

    # -- 1. Create dataset -----------------------------------------------------
    dataset = Dataset(
        project_id=project_id,
        name=name,
        description=description,
        source=source,
        import_config=json.dumps(column_configs),
    )
    db.add(dataset)
    db.flush()

    # -- 2. Create columns (non-skipped) ----------------------------------------
    columns: dict[int, DatasetColumn] = {}  # col_idx -> DatasetColumn
    seq = 0

    for cfg in sorted(column_configs, key=lambda c: c["column_index"]):
        col_idx = cfg["column_index"]
        if cfg.get("skip") or cfg.get("column_type") == ColumnType.SKIP.value:
            continue

        qtype = ColumnType(cfg["column_type"])
        scale_labels = cfg.get("scale_labels")
        # #28: an SPSS import supplies the scale's real codes (possibly 0-based or
        # gapped). Anything else omits them and keeps the positional 1..N encoding.
        scale_values = cfg.get("scale_values")
        if scale_values and len(scale_values) != len(scale_labels or []):
            logger.warning(
                "scale_values/scale_labels length mismatch on column %s (%s vs %s) — "
                "falling back to positional encoding",
                col_idx, len(scale_values), len(scale_labels or []),
            )
            scale_values = None

        # Scale metadata
        scale_labels_json = None
        scale_values_json = None
        scale_pts = None
        # #575: a cells-are-codes column defers ALL scale handling (metadata +
        # recode + substitution) to the apply_value_labels post-pass, so it's
        # created bare here and the cell loop stores the raw numeric code.
        cells_are_codes = bool(cfg.get("cells_are_codes"))
        if qtype == ColumnType.ORDINAL and scale_labels and not cells_are_codes:
            scale_labels_json = json.dumps(scale_labels)
            scale_values_json = json.dumps(
                _coerce_scale_codes(scale_values)
                if scale_values
                else list(range(1, len(scale_labels) + 1))
            )
            scale_pts = len(scale_labels)

        # #592: the column's declared missing rules (config-borne — the wizard
        # in slab 4, .sav in slab 5). None = the recognized-N/A defaults.
        col_missing_rules = cfg.get("missing_values")

        # Numeric metadata (computed from data)
        n_fmt: str | None = None
        n_min: float | None = None
        n_max: float | None = None
        if qtype in (ColumnType.NUMERIC, ColumnType.PERCENTAGE):
            # #799: precomputed in the single scan pass — already DISTINCT,
            # which is what `_analyze_numeric` reduced it to anyway.
            col_vals = distinct_numeric_by_idx.get(col_idx, set())
            # #358: pass the CSV header (not column_text override) so the
            # percentage keyword check uses the original column name.
            col_header = headers[col_idx] if col_idx < len(headers) else None
            info = _analyze_numeric(list(col_vals), header=col_header)
            if info:
                n_fmt = info["numeric_format"]
                n_min = info["numeric_min"]
                n_max = info["numeric_max"]

        column = DatasetColumn(
            dataset_id=dataset.id,
            column_code=cfg.get("column_code") or f"C{seq + 1:03d}",
            column_name=cfg.get("column_name"),
            group_code=cfg.get("group_code"),
            group_label=cfg.get("group_label"),
            column_text=cfg.get(
                "column_text",
                headers[col_idx] if col_idx < len(headers) else "",
            ),
            column_type=qtype,
            sequence_order=seq,
            scale_labels=scale_labels_json,
            scale_values=scale_values_json,
            scale_points=scale_pts,
            numeric_min=n_min,
            numeric_max=n_max,
            numeric_format=n_fmt,
            demographic_subtype=cfg.get("demographic_subtype"),
            # `is not None` — an explicit [] declaration ("nothing is missing")
            # must persist, never fold into the NULL default (the falsy-zero rule).
            missing_values=(
                json.dumps(col_missing_rules)
                if col_missing_rules is not None else None
            ),
        )
        db.add(column)
        columns[col_idx] = column
        seq += 1

    db.flush()  # get column IDs

    # -- 2b. Create RecodeDefinitions for ordinal columns ----------------------
    for col_idx, column in columns.items():
        cfg = cfg_by_idx.get(col_idx, {})
        qtype_str = cfg.get("column_type", "")
        scale_labels = cfg.get("scale_labels")

        # #575: cells-are-codes columns get their primary scale_map from the
        # apply_value_labels post-pass, not here.
        if qtype_str != ColumnType.ORDINAL.value or not scale_labels or cfg.get("cells_are_codes"):
            continue

        # Build mapping: label -> code. The primary scale_map recode is a SECOND
        # owner of value_numeric — `append_import` re-applies it to new rows, and
        # the recode workbench re-applies it on demand. It must agree with
        # `_compute_value_numeric`, or an SPSS 0-based scale imports as 0..3 and
        # then silently rewrites to 1..4 on the first append (#28).
        scale_values = cfg.get("scale_values")
        if scale_values and len(scale_values) == len(scale_labels):
            codes = _coerce_scale_codes(scale_values)
        else:
            codes = list(range(1, len(scale_labels) + 1))
        mapping = {label: codes[i] for i, label in enumerate(scale_labels)}

        # Pre-scan data rows for missing values (#592: column-aware — the
        # exclude channel seeds FROM the effective rule, §J.2)
        col_missing_rules = cfg.get("missing_values")
        # #799: precomputed in the single scan pass.
        na_values = na_values_by_idx.get(col_idx, set())

        exclude_values_json = json.dumps(sorted(na_values)) if na_values else None

        # Name: use scale point count
        recode_name = f"{len(scale_labels)}-point scale"

        recode_def = RecodeDefinition(
            column_id=column.id,
            name=recode_name,
            recode_type=RecodeType.SCALE_MAP,
            output_type=OutputType.NUMERIC,
            mapping=json.dumps(mapping),
            exclude_values=exclude_values_json,
            is_primary=True,
            is_auto_detected=True,
            sequence_order=0,
        )
        db.add(recode_def)

    db.flush()  # get recode definition IDs

    # -- 3. Process data rows -> rows + values ----------------------------------
    values_created = 0
    # #415: track values recognized as missing (N/A / refusal labels) so the
    # import results screen can disclose the silent missing-handling (#381/#384).
    recognized_missing_count = 0
    recognized_missing_labels: set[str] = set()

    # #796b: BATCHED. This loop used to `db.flush()` once per row and `db.add()`
    # an ORM instance per cell. MEASURED on a real GSS extract (75,699 x 41 =
    # 3,103,659 values): **374.8s**, six minutes for one file and past any
    # timeout a client can reasonably offer. Batching the row flush (75,699
    # round trips -> 38) and inserting values via a Core executemany took it to
    # **76.4s — 4.9x** — with identical row/value counts, record identifiers and
    # uuids.
    #
    # Rows still become ORM objects: there are only tens of thousands, and
    # `DatasetRow` carries a Python-side `uuid` default (the Track J identity
    # spine) that a Core insert would not apply. VALUES go through Core:
    # `DatasetValue` has no defaults and no post-insert consumer in this
    # function, so nothing needs the instances.
    #
    # ⚠️ **This is a SPEED fix and NOT a memory fix — do not read it as one.**
    # Peak RSS was 521 MB before and 533 MB after, and the staged measurement
    # says why: baseline 59 MB -> **223 MB** after openpyxl's workbook read ->
    # **511 MB** after `data_rows = list(reader)` materialises 3.1M Python str
    # objects. The per-cell ORM instances were never the driver. Both real
    # drivers predate this loop and neither is addressed here (see #799); the
    # <256 MB backend target is still exceeded on a file this size.
    #
    # ⚠️ The batch sizes bound what THIS loop adds on top, not the total.
    #
    # ⚠️ **The ORM row insert is deliberately NOT converted to Core, and this is
    # measured rather than assumed.** SQLAlchemy emits one INSERT per row for
    # this mapper even under `add_all` (RETURNING is available and the page size
    # is 1000, so the reason is the mapper, not the dialect). A Core insert with
    # RETURNING is **3.0x** faster on 75,699 rows — but that is 5.7s -> 1.9s
    # against a 76.4s import, **5% of the total**, and it would require spelling
    # `DatasetRow`'s Python-side `uuid` and `created_at` defaults here, where
    # they would silently diverge the day the model changes. Not worth it. The
    # values were the win; the rows are not.
    ROW_BATCH = 2_000
    VALUE_BATCH = 10_000
    pending_values: list[dict] = []

    def _drain_values() -> None:
        if pending_values:
            db.execute(sa_insert(DatasetValue), pending_values)
            pending_values.clear()

    # #799: stream the rows a SECOND time rather than holding them. Two parses
    # of the CSV text total (this and `_scan_source_rows`) replace one parse plus
    # a retained 3.1M-object list — measured at ~0.8s per parse against a ~76s
    # import, i.e. ~1% of the time for 288 MB of memory.
    source_rows = csv.reader(io.StringIO(text))
    next(source_rows, None)  # header
    batch_start = 0
    while True:
        batch = list(islice(source_rows, ROW_BATCH))
        if not batch:
            break
        ds_rows = [
            DatasetRow(
                dataset_id=dataset.id,
                participant_id=None,
                # System-generated record identifier — numbering is unchanged
                # from the per-row loop this replaces.
                row_identifier=f"R{str(batch_start + i + 1).zfill(pad_width)}",
                submitted_at=None,
            )
            for i in range(len(batch))
        ]
        db.add_all(ds_rows)
        db.flush()  # ONE flush per batch, not per row — populates ds_row.id

        for ds_row, data_row in zip(ds_rows, batch):
            for col_idx, column in columns.items():
                if col_idx >= len(data_row):
                    continue
                cell = data_row[col_idx].strip()
                if not cell:
                    continue

                cfg = cfg_by_idx.get(col_idx, {})
                col_missing_rules = cfg.get("missing_values")

                # #415: recognized-missing accounting. Mirrors the per-column
                # na_count in preview_dataset_csv and the value-keyed compute rule
                # (missing everywhere; #592: column-aware when the config declares).
                # value_text still stores the raw label; value_numeric lands None.
                if is_missing(cell, col_missing_rules):
                    recognized_missing_count += 1
                    if len(recognized_missing_labels) < 25:
                        recognized_missing_labels.add(cell)

                if cfg.get("cells_are_codes"):
                    # #575: the cell IS the numeric code; keep it (value_text stays the
                    # raw code). apply_value_labels substitutes the label + owns the
                    # scale metadata/recode in the post-pass below. Passing scale_labels
                    # to _compute here would route to label→code and NULL a bare code.
                    # #592: a declared-missing code stores NULL, never its number.
                    value_numeric = (
                        None if is_missing(cell, col_missing_rules)
                        else _strip_numeric(cell)
                    )
                else:
                    value_numeric = _compute_value_numeric(
                        cell, cfg.get("column_type", ""), cfg.get("scale_labels"),
                        cfg.get("scale_values"),
                        missing_rules=col_missing_rules,
                    )

                col_type = cfg.get("column_type", "")
                wc = len(cell.split()) if col_type == "open_text" and cell.strip() else None

                # #607: a labelled missing rule substitutes its label into the cell,
                # exactly as the declare endpoint, the append channel, and the .sav
                # adapter do — otherwise the same code renders two ways in one
                # column ("99" here, "Refused" everywhere else) and the append
                # dedup fingerprint misses precisely the rows it exists to match.
                # `recognized_missing_labels` above records the RAW cell (the
                # disclosure lists what the file carried).
                # A plain dict, never an ORM instance: 3.1M `DatasetValue` objects
                # in one identity map is what cost 521 MB (#796b).
                pending_values.append({
                    "row_id": ds_row.id,
                    "column_id": column.id,
                    "value_text": matched_missing_label(cell, col_missing_rules) or cell,
                    "value_numeric": value_numeric,
                    "word_count": wc,
                })
                values_created += 1

            if len(pending_values) >= VALUE_BATCH:
                _drain_values()

        # ⚠️ The record-identifier counter. `range(0, n, ROW_BATCH)` used to
        # advance this; the streaming loop must do it by hand, and a fixture
        # with only ONE batch cannot tell the difference — every batch would
        # restart at R0000001.
        batch_start += len(batch)

    _drain_values()
    db.flush()

    # -- 3b. Declared value labels (#575) --------------------------------------
    # For each cells-are-codes column, apply the authored code→label dictionary
    # the SAME way the retro path and .sav import do — substitute the label into
    # value_text, keep the code in value_numeric, set scale metadata + a primary
    # scale_map. Reusing apply_value_labels (vs re-implementing inline) keeps
    # undeclared codes numeric and handles nominal, which _compute_value_numeric
    # would silently NULL. Lazy import: value_labels imports from this module.
    value_label_unlabeled: dict[int, list[float]] = {}
    codes_columns = [
        (idx, col) for idx, col in columns.items()
        if cfg_by_idx.get(idx, {}).get("cells_are_codes")
    ]
    if codes_columns:
        from .value_labels import apply_value_labels

        for col_idx, column in codes_columns:
            cfg = cfg_by_idx.get(col_idx, {})
            labels = cfg.get("scale_labels")
            values = cfg.get("scale_values")
            if not labels or not values or len(labels) != len(values):
                logger.warning(
                    "cells_are_codes column %s missing/mismatched labels/values — "
                    "skipping value-label substitution", col_idx,
                )
                continue
            pairs = [(float(code), label) for code, label in zip(values, labels)]
            result = apply_value_labels(db, column, pairs, target_type=column.column_type)
            if result["unlabeled_codes"]:
                value_label_unlabeled[col_idx] = result["unlabeled_codes"]

    # -- 4. Participant linking (#414, DEC-6) ------------------------------------
    # `is not None` is load-bearing: column index 0 is a valid link column.
    participant_link_report = None
    if participant_link_column_index is not None:
        link_col = columns.get(participant_link_column_index)
        if link_col is None or link_col.column_type != ColumnType.IDENTIFIER:
            raise ValueError(
                "participant_link_column_index must reference a non-skipped identifier column"
            )
        # Function-level import: participant_linking imports _is_na from this
        # module at top level, so the reverse edge must stay lazy.
        from .participant_linking import link_rows_by_identifier_column

        participant_link_report = link_rows_by_identifier_column(
            db,
            project_id=project_id,
            dataset_id=dataset.id,
            column_id=link_col.id,
        )

    return {
        "dataset_id": dataset.id,
        "columns_created": len(columns),
        "rows_created": row_count,
        "values_created": values_created,
        "recognized_missing_count": recognized_missing_count,
        "recognized_missing_labels": sorted(recognized_missing_labels),
        "participant_link_report": participant_link_report,
        "value_label_unlabeled": value_label_unlabeled,
    }
