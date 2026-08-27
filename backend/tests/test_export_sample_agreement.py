"""The R and Excel exports of one project agree about what the SAMPLE is (#822).

**What they disagreed about, measured on a real 75,699-record survey.** The R
export blanks a declared-missing cell (`grep -c Inapplicable gss_data.csv` → 0;
R read `<NA> 32276`). The Excel export wrote the sentinel TEXT — `.i:
Inapplicable`, `.d: Do not Know/Cannot Choose`, `.n: No answer` as ordinary cell
values, **1,099,939 of them across the workbook** — and its Data Dictionary had
no column that could correct the impression. So the two files answered "how many
people responded?" differently, and a recipient averaging a column in Excel
silently included "No answer" rows.

**Why a cross-export test and not two per-export tests.** Each export was
internally consistent and separately defensible; the defect existed only in the
relationship between them, and nothing in the suite had ever compared them. That
is the same two-halves-of-one-fact shape as #732/#742/#746 — each file reads
correctly on its own while the pair drifts.

⚠️ **The blanking REVERSES a documented decision** (#592 §I.10 / #611e, recorded
in `export_excel.py` as deliberate): the data sheet used to be a raw-data escape
hatch where a declared "99" stayed visible. The developer's call on 2026-08-25
was that two exports of one project must not disagree about the sample, and that
the Data Dictionary — not the cell — is where the distinction between "Do not
know", "No answer" and "Inapplicable" belongs. Both halves ship together for
that reason.
"""

import asyncio
import csv
import io
import json
import zipfile

import pytest
from openpyxl import load_workbook

from app.models.dataset import ColumnType, Dataset, DatasetColumn, DatasetRow, DatasetValue
from app.models.project import Project
from app.models.recode import OutputType, RecodeDefinition, RecodeType
from app.models.user import User
from app.routers.export_excel import export_datasets_excel
from app.routers.export_r import export_r_data

PID = 822

# The declaration under test: two labelled sentinels on a nominal column.
REGION_RULES = [
    {"value": "98", "label": "Refused"},
    {"value": "99", "label": "No answer"},
]

# row_identifier, pid, region, age_text, age_num, site
ROWS = [
    ("R1", "P01", "North", "34", 34.0, "Clinic A"),
    ("R2", "P02", "South", "41", 41.0, "Clinic B"),
    # Declared missing: the cell TEXT is the rule's label (#607), and the
    # declaration nulls value_numeric — the state a real declared column is in.
    ("R3", "P03", "Refused", "-99", None, "Clinic A"),
    ("R4", "P04", "No answer", "-99", None, "N/A"),
    # An identifier that LOOKS missing. R never blanks one (#533: an ID is a
    # join key, not an analysis value) and Excel must not either.
    ("R5", "N/A", "North", "29", 29.0, "Clinic B"),
]


def _run(coro):
    return asyncio.run(coro)


@pytest.fixture
def sample_project(db_session):
    db = db_session
    db.add(Project(id=PID, name="Sample agreement", user_id=1))
    db.flush()
    db.add(Dataset(id=PID, project_id=PID, name="survey"))
    db.flush()

    # Column codes are already valid R identifiers, so the Excel header
    # (`column_code`) and the R header (`_make_r_identifier(column_code)`) are
    # the same string and the two files can be matched column by column.
    cols = [
        DatasetColumn(id=8221, dataset_id=PID, column_code="pid", column_name="pid",
                      column_text="Participant", column_type=ColumnType.IDENTIFIER,
                      sequence_order=0, display_order=0),
        # NOMINAL, declared. R emits its TEXT, so this is the column where the
        # two exports could differ on the cell itself.
        DatasetColumn(id=8222, dataset_id=PID, column_code="region", column_name="region",
                      column_text="Region", column_type=ColumnType.NOMINAL,
                      missing_values=json.dumps(REGION_RULES),
                      sequence_order=1, display_order=1),
        # NUMERIC with a RANGE rule — the shape a per-row checkbox cannot
        # express (#609), and the one whose sentinel is a plausible number.
        DatasetColumn(id=8223, dataset_id=PID, column_code="age", column_name="age",
                      column_text="Age", column_type=ColumnType.NUMERIC,
                      missing_values=json.dumps([{"lo": -99, "hi": -1}]),
                      sequence_order=2, display_order=2),
        # DEMOGRAPHIC and UNDECLARED: the `_is_na` DEFAULTS arm. A test that
        # only covered declared rules could not tell "declaration-aware" from
        # "defaults re-inlined" (the #592 two-sided rule).
        DatasetColumn(id=8224, dataset_id=PID, column_code="site", column_name="site",
                      column_text="Site", column_type=ColumnType.DEMOGRAPHIC,
                      sequence_order=3, display_order=3),
    ]
    for c in cols:
        db.add(c)
    db.flush()

    for i, (rid, pid_v, region, age_t, age_n, site) in enumerate(ROWS):
        row = DatasetRow(id=8230 + i, dataset_id=PID, row_identifier=rid)
        db.add(row)
        db.flush()
        db.add(DatasetValue(row_id=row.id, column_id=8221, value_text=pid_v))
        db.add(DatasetValue(row_id=row.id, column_id=8222, value_text=region))
        db.add(DatasetValue(row_id=row.id, column_id=8223, value_text=age_t,
                            value_numeric=age_n))
        db.add(DatasetValue(row_id=row.id, column_id=8224, value_text=site))

    # A non-primary recode definition, so the dictionary's Recode columns have
    # something real to describe (the #822 rider).
    db.add(RecodeDefinition(
        id=8240, column_id=8222, name="Region 2-way",
        recode_type=RecodeType.CATEGORY_GROUP, output_type=OutputType.CATEGORICAL,
        mapping=json.dumps({"North": "N", "South": "S"}), is_primary=False,
    ))
    db.flush()
    return db.query(User).filter(User.id == 1).one()


async def _collect(resp) -> bytes:
    """Drain a StreamingResponse; both endpoints return one."""
    out = []
    async for chunk in resp.body_iterator:
        out.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return b"".join(out)


def _excel_sheets(user, db):
    async def go():
        return await _collect(export_datasets_excel(project_id=PID, user=user, db=db))
    return load_workbook(io.BytesIO(_run(go())))


def _r_data_rows(user, db):
    raw = _run(_export_r_zip(user, db))
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        name = next(n for n in zf.namelist() if n.endswith("_data.csv"))
        text = zf.read(name).decode("utf-8")
    return list(csv.DictReader(io.StringIO(text)))


async def _export_r_zip(user, db) -> bytes:
    return await _collect(export_r_data(project_id=PID, user=user, db=db))


def _excel_data_grid(wb) -> tuple[list[str], list[list]]:
    ws = wb["survey"]
    rows = list(ws.iter_rows(values_only=True))
    return [h for h in rows[0]], [list(r) for r in rows[1:]]


class TestTheTwoExportsAgree:
    def test_the_same_cells_are_empty_in_both_files(self, db_session, sample_project):
        user = sample_project
        headers, grid = _excel_data_grid(_excel_sheets(user, db_session))
        r_rows = _r_data_rows(user, db_session)

        # Self-check FIRST: a comparison over an empty column set, or over a
        # fixture with nothing missing in it, passes by finding nothing.
        shared = [h for h in headers if h in (r_rows[0].keys() if r_rows else {})]
        assert set(shared) >= {"pid", "region", "age", "site"}, \
            f"columns did not line up between the two exports: {headers} vs {list(r_rows[0])}"
        assert len(r_rows) == len(ROWS) == len(grid)

        disagreements = []
        for i, r_row in enumerate(r_rows):
            excel_row = dict(zip(headers, grid[i]))
            for col in shared:
                r_empty = (r_row[col] or "").strip() == ""
                x_val = excel_row.get(col)
                x_empty = x_val is None or str(x_val).strip() == ""
                if r_empty != x_empty:
                    disagreements.append(
                        f"row {i + 1} {col}: R={'empty' if r_empty else r_row[col]!r} "
                        f"Excel={'empty' if x_empty else x_val!r}"
                    )
        assert disagreements == [], (
            "The R and Excel exports must not disagree about which cells hold a "
            "response — a recipient comparing them would get a different N.\n"
            + "\n".join(disagreements)
        )

        # ...and the fixture really exercised it: at least one cell IS blank.
        blanked = sum(
            1 for i in range(len(grid))
            for col in ("region", "age", "site")
            if str(dict(zip(headers, grid[i])).get(col) or "").strip() == ""
        )
        assert blanked >= 4, f"fixture produced only {blanked} blanked cells"

    def test_no_sentinel_text_survives_into_either_file(self, db_session, sample_project):
        user = sample_project
        wb = _excel_sheets(user, db_session)
        headers, grid = _excel_data_grid(wb)
        flat = " | ".join(str(v) for row in grid for v in row if v is not None)
        for sentinel in ("Refused", "No answer", "-99"):
            assert sentinel not in flat, \
                f"{sentinel!r} is a declared non-answer and must not read as data"
        r_rows = _r_data_rows(user, db_session)
        r_flat = " | ".join(v or "" for row in r_rows for v in row.values())
        for sentinel in ("Refused", "No answer", "-99"):
            assert sentinel not in r_flat

    def test_an_identifier_that_looks_missing_survives_in_both(self, db_session, sample_project):
        # #533's carve-out, carried across. Blanking it here would make the
        # Excel export unjoinable exactly where the R export stays joinable —
        # the same divergence this fix removes, pointing the other way.
        user = sample_project
        headers, grid = _excel_data_grid(_excel_sheets(user, db_session))
        pid_col = headers.index("pid")
        assert "N/A" in [str(r[pid_col]) for r in grid]
        r_rows = _r_data_rows(user, db_session)
        assert "N/A" in [r["pid"] for r in r_rows]


class TestTheDictionaryExplainsWhatTheCellsNoLongerSay:
    def test_missing_values_column_names_every_rule(self, db_session, sample_project):
        wb = _excel_sheets(sample_project, db_session)
        ws = wb["Data Dictionary"]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0])
        assert "Missing Values" in headers, \
            "blanking the cells without this column loses the distinction between " \
            "'Do not know', 'No answer' and 'Inapplicable' entirely"
        mi, ci = headers.index("Missing Values"), headers.index("Code")
        by_code = {r[ci]: r[mi] for r in rows[1:]}
        assert by_code["region"] == "98 = Refused; 99 = No answer"
        assert by_code["age"] == "-99 to -1"
        # The undeclared column says which state it is in, in the same words the
        # authoring tri-state uses (#609) — not a blank cell, which would be
        # indistinguishable from "nothing is missing".
        assert by_code["site"] == "Automatic"

    def test_the_recode_columns_describe_what_the_data_sheet_did(self, db_session, sample_project):
        # ⚠️ The filed rider said these were "blank on every row". Checked here
        # rather than assumed: `region` has a non-primary definition, the data
        # sheet emits a `region [Region 2-way]` column for it, and the
        # dictionary must describe the same rule.
        wb = _excel_sheets(sample_project, db_session)
        headers, _ = _excel_data_grid(wb)
        assert "region [Region 2-way]" in headers

        rows = list(wb["Data Dictionary"].iter_rows(values_only=True))
        dh = list(rows[0])
        ci, ni, ti, mi = (dh.index("Code"), dh.index("Recode Name"),
                          dh.index("Recode Type"), dh.index("Mapping"))
        region_rows = [r for r in rows[1:] if r[ci] == "region"]
        assert len(region_rows) == 1
        assert region_rows[0][ni] == "Region 2-way"
        assert region_rows[0][ti] == "category_group"
        assert "North -> N" in str(region_rows[0][mi])
