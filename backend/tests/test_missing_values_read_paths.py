"""#592 slab 2 — the read-time surfaces are COLUMN-AWARE.

Every surface that used to hardcode the recognized-N/A defaults now judges a
cell by ITS column's declared ``missing_values`` rules, with REPLACE
semantics. The load-bearing assertion on each surface is therefore TWO-SIDED:
on a column declaring only ``99``,

- ``"99"`` IS missing (never a group / category / factor level / operand),
  even when ``value_numeric`` is still populated (the pre-re-apply / #594
  writer shape — identification keys on value_text, §I.4);
- ``"Prefer not to say"`` is NOT missing (the defaults were REPLACED, §I.7).

A one-sided test cannot tell "declaration-aware" from "defaults hardcoded" —
the defaults also drop "Prefer not to say" and keep "99".

Surfaces pinned here: the grouping loader (per-column judging), the metrics
resolvers + computers (frequency / proportion / mean incl. the pooled domain
path, C2), dataset cross-tab, Data Quality classification, the R export's
factor levels, the recode workbench's value-frequency ``is_na`` flag, and
computed-column expressions.
"""
import json

import pytest

from app.models.project import Project
from app.models.dataset import Dataset, DatasetColumn, DatasetRow, DatasetValue
from app.models.analysis_domain import AnalysisDomain, AnalysisDomainMember
from app.models.metric import MetricDefinition

DECLARE_99 = json.dumps([{"value": "99", "label": "Refused"}])


def _seed_column(db, *, dataset_id=1, col_id=10, project_id=1,
                 column_type="ordinal", missing_values=DECLARE_99,
                 cells=(("3", 3.0), ("99", 99.0), ("Prefer not to say", None)),
                 row_start=100, make_project=True):
    """A column with a declaration and the canonical two-sided cell set.

    "99" keeps its value_numeric on purpose — the pre-re-apply shape — so
    every pin also proves identification keys on value_text (§I.4)."""
    if make_project:
        db.add(Project(id=project_id, name="P", user_id=1))
        db.flush()
        db.add(Dataset(id=dataset_id, project_id=project_id, name=f"S{dataset_id}"))
        db.flush()
    col = DatasetColumn(
        id=col_id, dataset_id=dataset_id, column_code=f"Q{col_id}",
        column_text=f"Q{col_id}", column_type=column_type,
        sequence_order=col_id, missing_values=missing_values,
    )
    db.add(col)
    db.flush()
    for i, (vt, vn) in enumerate(cells):
        row = DatasetRow(id=row_start + i, dataset_id=dataset_id)
        db.add(row)
        db.flush()
        db.add(DatasetValue(row_id=row.id, column_id=col_id,
                            value_text=vt, value_numeric=vn))
    db.flush()
    return col


class TestGroupingLoader:
    def test_declared_column_replaces_the_defaults(self, db_session):
        from app.services.grouping import load_grouping_values
        col = _seed_column(
            db_session,
            cells=(("Female", None), ("Male", None),
                   ("99", 99.0), ("Prefer not to say", None)),
        )
        groups = set(load_grouping_values(db_session, col.id, None).values())
        assert "99" not in groups, "declared-missing 99 formed a group"
        assert "Prefer not to say" in groups, (
            "REPLACE semantics: the declaration must supersede the defaults"
        )
        assert {"Female", "Male"} <= groups

    def test_each_column_judged_by_its_own_rules(self, db_session):
        """One loader call over two columns: the SAME text "99" is missing on
        the declared column and a real group on the undeclared one."""
        from app.services.grouping import load_grouping_values_for_columns
        declared = _seed_column(db_session, cells=(("99", 99.0),))
        db_session.add(Dataset(id=2, project_id=1, name="S2"))
        db_session.flush()
        undeclared = _seed_column(
            db_session, dataset_id=2, col_id=20, missing_values=None,
            cells=(("99", 99.0),), row_start=200, make_project=False,
        )
        result = load_grouping_values_for_columns(
            db_session, [declared.id, undeclared.id], None,
        )
        assert 100 not in result, "declared column's 99 must be missing"
        assert result.get(200) == "99", "undeclared column's 99 is substantive"


class TestMetricsResolvers:
    def _metric(self, db, col_id, metric_type, config="{}"):
        m = MetricDefinition(
            project_id=1, name=f"m-{metric_type}", metric_type=metric_type,
            input_source_type="dataset_column", input_source_id=col_id,
            config=config,
        )
        db.add(m)
        db.flush()
        return m

    def test_frequency_two_sided(self, db_session):
        from app.services.metrics import compute_metric
        col = _seed_column(db_session)
        results = compute_metric(
            db_session, self._metric(db_session, col.id, "frequency_distribution"))
        data = json.loads(results[0].result_data)
        assert "99" not in data["counts"], "declared-missing 99 is a category"
        assert data["counts"].get("Prefer not to say") == 1, (
            "REPLACE: 'Prefer not to say' is substantive on this column"
        )
        assert results[0].valid_n == 2  # "3" + "Prefer not to say"
        assert results[0].total_n == 3

    def test_mean_excludes_declared_missing_despite_numeric(self, db_session):
        """§I.4 / the #594 read-side: value_numeric is still 99.0 on the
        missing cell — the resolver's text-keyed mark keeps it out anyway."""
        from app.services.metrics import compute_metric
        col = _seed_column(db_session)
        results = compute_metric(db_session, self._metric(db_session, col.id, "mean"))
        data = json.loads(results[0].result_data)
        assert data["mean"] == pytest.approx(3.0), (
            f"mean absorbed the declared-missing 99: {data['mean']}"
        )
        assert results[0].valid_n == 1

    def test_undeclared_column_keeps_the_defaults_end_to_end(self, db_session):
        """The defaults half, end-to-end: with NO declaration the resolver must
        still mark recognized-N/A text (a mutant that only honors declared
        rules and skips the defaults would pass every declared-column test)."""
        from app.services.metrics import compute_metric
        col = _seed_column(
            db_session, missing_values=None,
            cells=(("Yes", None), ("N/A", None), ("Prefer not to say", None)),
        )
        results = compute_metric(
            db_session, self._metric(db_session, col.id, "frequency_distribution"))
        data = json.loads(results[0].result_data)
        assert "N/A" not in data["counts"]
        assert "Prefer not to say" not in data["counts"]
        assert results[0].valid_n == 1

    def test_proportion_denominator_two_sided(self, db_session):
        from app.services.metrics import compute_metric
        col = _seed_column(db_session)
        results = compute_metric(db_session, self._metric(
            db_session, col.id, "proportion",
            config=json.dumps({"mode": "values", "threshold_values": ["3"]}),
        ))
        data = json.loads(results[0].result_data)
        # countable = {"3", "Prefer not to say"} — 99 out (declared), PNTS in
        assert results[0].valid_n == 2
        assert data["proportion"] == pytest.approx(0.5)

    def test_domain_pooled_path_marks_per_column(self, db_session):
        """C2: the pooled domain path discards col_id — the mark must be set
        in the resolver. Two member columns, one declares 99, one doesn't:
        the SAME code is missing in one and data in the other."""
        from app.services.metrics import compute_metric
        col_a = _seed_column(db_session, cells=(("3", 3.0), ("99", 99.0)))
        col_b = _seed_column(
            db_session, col_id=11, missing_values=None,
            cells=(("5", 5.0), ("99", 99.0)), row_start=150, make_project=False,
        )
        domain = AnalysisDomain(id=1, project_id=1, name="D")
        db_session.add(domain)
        db_session.flush()
        for seq, cid in enumerate([col_a.id, col_b.id]):
            db_session.add(AnalysisDomainMember(
                domain_id=1, member_type="column", member_id=cid,
                sequence_order=seq,
            ))
        db_session.flush()
        m = MetricDefinition(
            project_id=1, name="pooled mean", metric_type="mean",
            input_source_type="dataset_domain", input_source_id=1, config="{}",
        )
        db_session.add(m)
        db_session.flush()
        results = compute_metric(db_session, m)
        data = json.loads(results[0].result_data)
        # col_a's 99 excluded (declared); col_b's 99 included (undeclared)
        assert data["mean"] == pytest.approx((3 + 5 + 99) / 3)


def test_cross_tabulation_two_sided(db_session):
    from app.services.cross_tabulation import compute_cross_tabulation
    row_col = _seed_column(
        db_session, column_type="nominal",
        cells=(("Female", None), ("99", 99.0), ("Prefer not to say", None)),
    )
    # Undeclared second axis on the same rows
    col_col = DatasetColumn(
        id=11, dataset_id=1, column_code="Q11", column_text="Q11",
        column_type="nominal", sequence_order=11,
    )
    db_session.add(col_col)
    db_session.flush()
    for row_id in (100, 101, 102):
        db_session.add(DatasetValue(row_id=row_id, column_id=11, value_text="Yes"))
    db_session.flush()

    out = compute_cross_tabulation(db_session, 1, row_col.id, col_col.id)
    assert "99" not in out["row_values"], "declared-missing 99 is a cross-tab row"
    assert "Prefer not to say" in out["row_values"], "REPLACE semantics"
    assert "Female" in out["row_values"]


def test_data_quality_classification_two_sided(db_session):
    """REPLACE semantics at DQ, isolated from the #595 convergence.

    Every cell carries a value_numeric on purpose. Since slab 5, DQ's "valid"
    means "analysis will use this cell", so an ordinal cell with NO code reads
    missing too (every mean drops it). The shared fixture leaves "Prefer not to
    say" code-less, which would make this pin green for the wrong reason — it
    would report missing whether or not the declaration was honored. Giving it a
    code puts the decision back where this test is looking: on the declaration.
    """
    from app.services.data_quality import compute_missing_summary
    col = _seed_column(db_session, cells=(("3", 3.0), ("99", 99.0),
                                          ("Prefer not to say", 4.0)))
    out = compute_missing_summary(db_session, 1, [col.id])
    var = out["variables"][0]
    assert var["n_na"] == 1, "the declaration decides — only 99"
    assert var["n_valid"] == 2, (
        "REPLACE: the declaration ousts the defaults, so 'Prefer not to say' is "
        "a substantive answer on THIS column"
    )


def test_export_r_observed_values_two_sided(db_session):
    from app.routers.export_r import _get_observed_values
    col = _seed_column(
        db_session, column_type="nominal",
        cells=(("Female", None), ("99", 99.0), ("Prefer not to say", None)),
    )
    observed = _get_observed_values(db_session, col)
    assert "99" not in observed, "declared-missing 99 became an R factor level"
    assert "Prefer not to say" in observed, "REPLACE semantics"


def test_recode_value_frequency_flag_two_sided(db_session):
    from app.services.recode import get_value_frequencies
    col = _seed_column(db_session)
    flags = {r["value_text"]: r["is_na"] for r in
             get_value_frequencies(db_session, col.id)}
    assert flags["99"] is True, "the workbench flag must honor the declaration"
    assert flags["Prefer not to say"] is False, "REPLACE semantics"
    assert flags["3"] is False


class TestComputedColumns:
    def test_declared_missing_is_null_even_with_numeric(self, db_session):
        """The ColumnRef seam checks missing BEFORE the numeric branch, keyed
        on value_text (§I.4): a declared-missing "99" cell that still carries
        value_numeric=99.0 must evaluate NULL, not 198."""
        from app.services.computed_columns import evaluate_computed_column
        col = _seed_column(db_session, cells=(("3", 3.0), ("99", 99.0)))
        comp = DatasetColumn(
            id=90, dataset_id=1, column_code="X1", column_text="Doubled",
            column_type="numeric", sequence_order=90, source="computed",
            expression=f"[Q{col.id}] * 2",
            depends_on_column_ids=json.dumps([col.id]),
        )
        db_session.add(comp)
        db_session.flush()
        evaluate_computed_column(db_session, comp)
        out = {v.row_id: v.value_numeric for v in
               db_session.query(DatasetValue).filter(DatasetValue.column_id == 90)}
        assert out[100] == 6.0
        assert out[101] is None, "declared-missing 99 evaluated as a number"

    def test_replace_semantics_keep_declared_substantive_text(self, db_session):
        """On a declared column, "Prefer not to say" is genuine TEXT — an ==
        comparison must still reach it (pre-#592 the defaults nulled it)."""
        from app.services.computed_columns import evaluate_computed_column
        col = _seed_column(db_session)
        comp = DatasetColumn(
            id=91, dataset_id=1, column_code="X2", column_text="Declined?",
            column_type="numeric", sequence_order=91, source="computed",
            expression=f'IF([Q{col.id}] == "Prefer not to say", 1, 0)',
            depends_on_column_ids=json.dumps([col.id]),
        )
        db_session.add(comp)
        db_session.flush()
        evaluate_computed_column(db_session, comp)
        out = {v.row_id: v.value_numeric for v in
               db_session.query(DatasetValue).filter(DatasetValue.column_id == 91)}
        assert out[102] == 1.0, "REPLACE: declared column's PNTS is matchable text"
        assert out[101] is None, "declared-missing 99 propagates NULL"
        assert out[100] == 0.0


# ═══════════════════════════════════════════════════════════════════════════
# #611 — the close-out review's uncovered declared-rules surfaces.
# Same two-sided discipline: declared "99" excluded AND a defaults-shaped text
# kept, per surface, so "declaration-aware" is distinguishable from "defaults
# re-inlined". (The participant_linking pin the filing asked for is REFUTED —
# identifier columns cannot legally carry a declaration; the endpoint 400s
# IDENTIFIER/OPEN_TEXT — so that surface's defaults test + the 400 pin close it.)
# ═══════════════════════════════════════════════════════════════════════════
import asyncio

from app.models.participant import Participant
from app.models.statistical_test import StatisticalTest
from app.models.user import User


def _run(coro):
    return asyncio.run(coro)


def test_participant_group_map_two_sided(db_session):
    """#611a — `_build_participant_group_map` under a DECLARATION (its existing
    tests cover only the defaults + the role exemption): a declared "99"
    respondent joins no group, while "Prefer not to say" — substantive under
    REPLACE — forms a real group."""
    from app.services.code_analysis import _build_participant_group_map
    db = db_session
    db.add(Project(id=1, name="P", user_id=1))
    db.flush()
    db.add(Dataset(id=1, project_id=1, name="S"))
    db.flush()
    col = DatasetColumn(
        id=10, dataset_id=1, column_code="Q10", column_text="Grade",
        column_type="demographic", demographic_subtype="Grade",
        sequence_order=0, missing_values=DECLARE_99,
    )
    db.add(col)
    db.flush()
    for i, val in enumerate(("8", "99", "Prefer not to say"), start=1):
        db.add(Participant(id=i, project_id=1, identifier=f"P{i}"))
        db.flush()
        row = DatasetRow(id=100 + i, dataset_id=1, participant_id=i)
        db.add(row)
        db.flush()
        db.add(DatasetValue(row_id=row.id, column_id=10, value_text=val))
    db.flush()

    pmap = _build_participant_group_map(db, 1, "Grade")
    assert pmap.get(1) == "8"
    assert 2 not in pmap, "declared-missing 99 joined a demographic group"
    assert pmap.get(3) == "Prefer not to say", "REPLACE semantics"


def test_linkable_rows_labels_two_sided(db_session):
    """#611d — `get_linkable_rows` skips missing values when building row
    labels: a declared "99" identifies nothing, "Prefer not to say" (REPLACE)
    is a legitimate label."""
    from app.routers.dataset import get_linkable_rows
    col = _seed_column(
        db_session, column_type="nominal",
        cells=(("Maple School", None), ("99", 99.0), ("Prefer not to say", None)),
    )
    user = db_session.query(User).filter(User.id == 1).one()
    out = _run(get_linkable_rows(project_id=1, dataset_id=1, user=user, db=db_session))
    by_row = {r["row_id"]: r["display_values"] for r in out["rows"]}
    assert by_row[100] == ["Maple School"]
    assert by_row[101] == [], "declared-missing 99 was spent on a row label"
    assert by_row[102] == ["Prefer not to say"], "REPLACE semantics"
    assert col.id  # (silence unused warning)


class TestReliabilityFoldDeclared:
    """#611c — `build_row_item_matrix`'s `excluded or row.missing` fold
    (services/statistical_tests.py — the Cronbach/split-half gather) had no
    missing coverage at all; the resolver mark is pinned for compute_metric
    but this is a distinct consumer."""

    def _domain_with_two_items(self, db, col_a_missing, col_a_cells):
        db.add(Project(id=1, name="P", user_id=1))
        db.flush()
        db.add(Dataset(id=1, project_id=1, name="S"))
        db.flush()
        col_a = DatasetColumn(
            id=10, dataset_id=1, column_code="A", column_text="A",
            column_type="ordinal", sequence_order=0,
            missing_values=col_a_missing,
        )
        col_b = DatasetColumn(
            id=11, dataset_id=1, column_code="B", column_text="B",
            column_type="ordinal", sequence_order=1,
        )
        db.add_all([col_a, col_b])
        db.flush()
        b_cells = [("4", 4.0), ("2", 2.0), ("3", 3.0), ("5", 5.0)]
        for i, (a_cell, b_cell) in enumerate(zip(col_a_cells, b_cells)):
            row = DatasetRow(id=100 + i, dataset_id=1)
            db.add(row)
            db.flush()
            db.add(DatasetValue(row_id=row.id, column_id=10,
                                value_text=a_cell[0], value_numeric=a_cell[1]))
            db.add(DatasetValue(row_id=row.id, column_id=11,
                                value_text=b_cell[0], value_numeric=b_cell[1]))
        domain = AnalysisDomain(id=1, project_id=1, name="D")
        db.add(domain)
        db.flush()
        for seq, cid in enumerate([10, 11]):
            db.add(AnalysisDomainMember(
                domain_id=1, member_type="column", member_id=cid,
                sequence_order=seq,
            ))
        test = StatisticalTest(
            project_id=1, test_type="cronbachs_alpha",
            target_type="analysis_domain", target_id=1, config="{}",
        )
        db.add(test)
        db.flush()
        return test

    def test_declared_sentinel_is_listwise_excluded(self, db_session):
        """The declared "99" row drops (even though its value_numeric is still
        99.0 — identification keys on value_text, §I.4)."""
        from app.services.statistical_tests import compute_statistical_test
        test = self._domain_with_two_items(
            db_session, DECLARE_99,
            [("3", 3.0), ("99", 99.0), ("2", 2.0), ("4", 4.0)],
        )
        result = compute_statistical_test(db_session, test)
        assert result["n"] == 3, "declared-missing 99 row fed the item matrix"

    def test_declared_nothing_keeps_default_missing_text(self, db_session):
        """REPLACE, the [] arm: an "N/A" cell WITH a code on a []-declared
        column is a real item response — the defaults must not re-drop it."""
        from app.services.statistical_tests import compute_statistical_test
        test = self._domain_with_two_items(
            db_session, "[]",
            [("N/A", 4.0), ("3", 3.0), ("2", 2.0), ("4", 4.0)],
        )
        result = compute_statistical_test(db_session, test)
        assert result["n"] == 4, "[]-declared column re-applied the defaults"


class TestDataQualityPatternsAndMcar:
    """#611f (added by the scope) — `compute_missing_patterns` and
    `compute_littles_mcar` each thread `rules_by_col` independently of the
    pinned `compute_missing_summary`; per-surface threading is the #598-class
    risk."""

    def test_patterns_two_sided(self, db_session):
        """TWO declared rows vs ONE defaults-shaped row — deliberately
        asymmetric counts: with one of each, a defaults-hardcoded mutant
        flags the OTHER row and the missing COUNT comes out identical
        (patterns carry masks + counts, not row ids)."""
        from app.services.data_quality import compute_missing_patterns
        col = _seed_column(db_session, cells=(("3", 3.0), ("99", 99.0),
                                              ("99", 99.0),
                                              ("Prefer not to say", 4.0)))
        out = compute_missing_patterns(db_session, 1, [col.id])
        missing_rows = sum(p["count"] for p in out["patterns"] if p["pattern"][0])
        assert missing_rows == 2, (
            "exactly the two declared '99' rows are missing patterns — "
            "'Prefer not to say' is substantive under REPLACE"
        )
        assert out["total_rows"] == 4

    def test_mcar_two_sided(self, db_session):
        """Two numeric-eligible columns; the ONLY missing cell is the declared
        "99" (still carrying value_numeric — the no-resurrection rider). A
        defaults-hardcoded mutant also drops col B's []-declared "N/A"(4.0),
        creating a third pattern and shifting df."""
        from app.services.data_quality import compute_littles_mcar
        # Two "99" rows: MCAR skips patterns with fewer than 2 cases, so a
        # single declared-missing row would leave df non-positive.
        col_a = _seed_column(
            db_session,
            cells=(("3", 3.0), ("99", 99.0), ("5", 5.0),
                   ("4", 4.0), ("99", 99.0), ("2", 2.0)),
        )
        # Second column on the SAME rows (patterns/MCAR are row-wise).
        col_b = DatasetColumn(
            id=11, dataset_id=1, column_code="Q11", column_text="Q11",
            column_type="ordinal", sequence_order=11, missing_values="[]",
        )
        db_session.add(col_b)
        db_session.flush()
        for row_id, (vt, vn) in zip(
            (100, 101, 102, 103, 104, 105),
            (("2", 2.0), ("4", 4.0), ("N/A", 4.0),
             ("3", 3.0), ("5", 5.0), ("4", 4.0)),
        ):
            db_session.add(DatasetValue(row_id=row_id, column_id=11,
                                        value_text=vt, value_numeric=vn))
        db_session.flush()
        out = compute_littles_mcar(db_session, 1, [col_a.id, col_b.id])
        assert out["eligibility"]["eligible"], out["eligibility"]["reason"]
        assert out["result"]["n_patterns"] == 2 and out["result"]["df"] == 1, (
            "patterns must be exactly {complete, A-missing}: the declared 99 "
            "is the only missing cell; B's []-kept 'N/A' stays complete"
        )


def test_excel_export_asymmetry_is_deliberate(db_session):
    """#611e — the Excel datasets export is the researcher's RAW-data escape
    hatch: the raw value column keeps a declared-missing "99" VISIBLE (the
    deliberate asymmetry with the R export's `_text_cell`, which blanks it),
    while the recode column beside it honors the declaration (blank)."""
    import io as _io
    from openpyxl import load_workbook
    from app.models.recode import RecodeDefinition, RecodeType, OutputType
    from app.routers.export_excel import export_datasets_excel

    col = _seed_column(db_session, cells=(("3", 3.0), ("99", None)))
    db_session.add(RecodeDefinition(
        column_id=col.id, name="tens", recode_type=RecodeType.SCALE_MAP,
        output_type=OutputType.NUMERIC,
        mapping=json.dumps({"3": 30, "99": 990}),
        is_primary=False, sequence_order=0,
    ))
    db_session.flush()

    # The export reads column_type.value; reload so the seeded string coerces
    # to the enum (values_callable) the way real rows arrive.
    db_session.expire_all()
    user = db_session.query(User).filter(User.id == 1).one()
    resp = _run(export_datasets_excel(project_id=1, user=user, db=db_session))
    body = b"".join(_run(_drain(resp)))
    wb = load_workbook(_io.BytesIO(body))
    ws = wb["S1"] if "S1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Find the raw + recode columns for Q10 in the header row, then the "99" row.
    headers = [c.value for c in ws[1]]
    raw_idx = next(i for i, h in enumerate(headers, 1) if h and "Q10" in str(h)
                   and "tens" not in str(h))
    rec_idx = next(i for i, h in enumerate(headers, 1) if h and "tens" in str(h))
    raw_vals = {ws.cell(row=r, column=raw_idx).value: r
                for r in range(2, ws.max_row + 1)}
    assert "99" in raw_vals, (
        "the raw column must keep the declared '99' VISIBLE — this sheet is "
        "the raw-data escape hatch, the deliberate asymmetry with R's blanking"
    )
    r99, r3 = raw_vals["99"], raw_vals["3"]
    assert ws.cell(row=r99, column=rec_idx).value in (None, ""), (
        "the recode column must honor the declaration (blank for 99)"
    )
    assert ws.cell(row=r3, column=rec_idx).value == 30


async def _drain(resp):
    return [chunk async for chunk in resp.body_iterator]
