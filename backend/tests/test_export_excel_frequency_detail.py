"""The Excel export's Frequency Distributions section (#766).

`export_datasets_excel` read `result_data["distribution"]` — a key **no producer
has ever written**. `compute_frequency_distribution` returns `counts` /
`percentages` / `scale_order`, so the loop ran zero times and the section was
emitted as bare column headers with no data beneath them. `frequency_distribution`
is the DEFAULT metric type (`schemas/metric.py`), which makes it the most common
metric in the app — absent from every Excel export, silently, with no test over
this sheet at all.

⚠️ **These tests compute the metric for real rather than hand-writing
`result_data`.** That is the whole point: a fixture that hand-wrote a
`distribution` key would have passed against the broken code and pinned the
defect in place. The only thing that can catch a producer/consumer key mismatch
is running the actual producer. Same lesson as the R oracles — an assertion
anchored to our own assumption tests the assumption, not the code.
"""

import asyncio
import io
import json

import pytest
from openpyxl import load_workbook

from app.models.dataset import ColumnType, Dataset, DatasetColumn, DatasetRow, DatasetValue
from app.models.metric import MetricDefinition
from app.models.project import Project
from app.models.recode import RecodeDefinition
from app.models.user import User
from app.services.metrics import compute_metric

PID = 766

# Column indices into a row of the Metrics Detail sheet, named rather than
# inlined: queue #42 inserted the two CI columns BEFORE Valid N, and a bare
# `r[4]` silently started reading the lower bound as a sample size.
METRIC, OPTION, COUNT, PERCENTAGE, CI_LOWER, CI_UPPER, VALID_N = range(7)


def _run(coro):
    return asyncio.run(coro)


@pytest.fixture
def freq_project(db_session):
    """One ordinal column with a declared scale, one level nobody chose.

    The unchosen level is deliberate: the frequency computer zero-fills declared
    levels (#577), so the export must place it in DECLARED order rather than
    wherever a dict happened to put it.
    """
    db_session.add(Project(id=PID, name="P766", user_id=1))
    db_session.flush()
    db_session.add(Dataset(id=PID, project_id=PID, name="Survey"))
    db_session.flush()

    col = DatasetColumn(
        id=PID, dataset_id=PID, column_code="Sat", column_name="Sat",
        # The enum instance, not the bare string: `export_datasets_excel` reads
        # `column_type.value` for the Data Dictionary sheet, and a fixture that
        # assigns a str never round-trips through the DB to become one.
        column_text="Satisfaction", column_type=ColumnType.ORDINAL,
        sequence_order=0, display_order=0,
        scale_labels=json.dumps(["Low", "Neutral", "High"]),
        scale_values=json.dumps([1, 2, 3]),
        scale_points=3,
    )
    db_session.add(col)
    db_session.flush()
    db_session.add(RecodeDefinition(
        id=PID, column_id=col.id, name="scale", recode_type="scale_map",
        output_type="numeric", is_primary=True, sequence_order=0,
        mapping=json.dumps({"Low": 1, "Neutral": 2, "High": 3}),
    ))
    db_session.flush()

    # 3 Low, 0 Neutral, 1 High — 4 valid responses.
    for value in ("Low", "Low", "Low", "High"):
        row = DatasetRow(dataset_id=PID)
        db_session.add(row)
        db_session.flush()
        db_session.add(DatasetValue(row_id=row.id, column_id=col.id, value_text=value))
    db_session.flush()

    metric = MetricDefinition(
        project_id=PID, name="Satisfaction Freq",
        metric_type="frequency_distribution",
        input_source_type="dataset_column",
        input_source_id=col.id,
        config="{}",
    )
    db_session.add(metric)
    db_session.flush()
    compute_metric(db_session, metric)
    db_session.flush()
    return col, metric


def _detail_rows(db):
    from app.routers.export_excel import export_datasets_excel
    from tests.test_export_formula_injection import _stream_to_bytes

    user = db.get(User, 1)
    response = export_datasets_excel(project_id=PID, user=user, db=db)
    wb = load_workbook(io.BytesIO(_stream_to_bytes(response)))
    assert "Metrics Detail" in wb.sheetnames, "precondition: the sheet is emitted"
    ws = wb["Metrics Detail"]
    return [[c.value for c in row] for row in ws.iter_rows()]


class TestFrequencyDistributionsReachTheSheet:
    def test_the_section_has_data_rows_and_not_only_headers(self, db_session, freq_project):
        rows = _detail_rows(db_session)
        headers = [r for r in rows if r[:2] == ["Metric", "Response Option"]]
        assert headers, "precondition: the section header row is present"

        options = [r for r in rows if r[1] in ("Low", "Neutral", "High")]
        assert options, (
            "the Frequency Distributions section emitted headers and no data — "
            "the consumer is reading a result_data key the producer does not write"
        )
        # ⚠️ Existence of the rows is NOT enough, and the first draft of this
        # test stopped there. The option labels come from `scale_order` while the
        # numbers come from `counts`, so a lookup pointed at a key nobody writes
        # still yields correctly-labelled rows — full of zeros. Assert that at
        # least one real observation survived the trip.
        assert any((r[2] or 0) > 0 for r in options), (
            "every count exported as zero: the labels resolved but the counts did not"
        )

    def test_counts_and_percentages_match_what_the_app_computed(self, db_session, freq_project):
        _, metric = freq_project
        rd = json.loads(metric.results[0].result_data)
        rows = {r[1]: (r[2], r[3]) for r in _detail_rows(db_session) if r[1] in rd["counts"]}

        for label, count in rd["counts"].items():
            assert rows[label][0] == count
            # Read from the payload, never recomputed here — the export used to
            # derive the percentage from count/valid_n, a second copy of one fact
            # that would drift the moment the denominator rule changes (as #592
            # already changed it once for declared-missing values).
            assert rows[label][1] == pytest.approx(rd["percentages"][label], abs=0.01)

    def test_a_declared_level_nobody_chose_appears_in_declared_order(
        self, db_session, freq_project,
    ):
        """`scale_order` is the ordering seam (#406). Iterating the counts dict
        instead would put the zero-filled level wherever insertion left it."""
        rows = _detail_rows(db_session)
        seen = [r[1] for r in rows if r[1] in ("Low", "Neutral", "High")]
        assert seen == ["Low", "Neutral", "High"]
        neutral = next(r for r in rows if r[1] == "Neutral")
        assert neutral[2] == 0, "a structural zero is a row, with a zero in it"

    def test_valid_n_is_stated_once_per_metric(self, db_session, freq_project):
        rows = _detail_rows(db_session)
        ns = [r[VALID_N] for r in rows if r[1] in ("Low", "Neutral", "High")]
        assert ns[0] == 4, "4 valid responses"
        assert all(n is None for n in ns[1:]), (
            "repeated on every row it would read as a per-option denominator"
        )


class TestPerCategoryIntervalsReachTheSheet:
    """queue #42 — export parity for the margin of error.

    The screen and the exported workbook must not disagree about what a number
    is (the #732 class: the CSV reported eta-squared while the screen reported
    omega-squared). A spreadsheet has no tooltip, so the method disclosure gets
    a row of its own.
    """

    def test_each_option_carries_its_interval(self, db_session, freq_project):
        _, metric = freq_project
        rd = json.loads(metric.results[0].result_data)
        rows = {
            r[1]: (r[CI_LOWER], r[CI_UPPER])
            for r in _detail_rows(db_session) if r[1] in rd["counts"]
        }
        for label in rd["scale_order"]:
            lo, hi = rows[label]
            assert lo == pytest.approx(rd["ci_lower_by_label"][label], abs=0.01)
            assert hi == pytest.approx(rd["ci_upper_by_label"][label], abs=0.01)

    def test_the_zero_count_level_gets_an_honest_upper_bound(self, db_session, freq_project):
        """Wilson, not Wald — which is the whole reason the estimator matters
        here. At p = 0 Wald collapses to [0, 0], asserting certainty from four
        observations; Wilson states how high the true rate could still be."""
        rows = {r[1]: (r[CI_LOWER], r[CI_UPPER]) for r in _detail_rows(db_session)}
        lo, hi = rows["Neutral"]
        assert lo == 0
        assert hi > 0, "a category nobody chose is not thereby impossible"

    def test_the_sheet_states_that_the_intervals_are_not_simultaneous(
        self, db_session, freq_project,
    ):
        """The governing rule of B9: an interval must say what kind it is. Seven
        per-category binomial intervals do NOT jointly cover at 95%, and a
        spreadsheet gives the reader nowhere else to learn that."""
        text = " ".join(
            str(c) for row in _detail_rows(db_session) for c in row if c is not None
        )
        assert "Wilson" in text
        assert "not simultaneous" in text
