"""Tests for ROADMAP 12d (i) — CSV / Excel formula-injection defang.

Two helpers in `app/routers/export_helpers.py`:
- `csv_safe(value)` prepends an apostrophe to strings starting with =, @,
  tab, or CR. Numbers / None / safe strings pass through unchanged.
- `excel_set_safe(cell, value)` sets the cell value, then forces
  data_type='s' for strings starting with '=' (defangs openpyxl's
  auto-formula tagging at cell.py:198-199).

Scope is the high-impact subset of OWASP's CSV-injection prefix list. We
intentionally exclude '+' and '-' to avoid false-positives on legitimate
negative numbers in respondent demographic free-text. See helper docstring
for rationale.

This file holds three layers:
1. Helper unit tests — direct calls, no fixtures.
2. Polluted-fixture integration sweep — instantiate a project where every
   user-typed field starts with '=cmd|...' or '@SUM(...)' and verify the
   export bytes don't surface a raw formula prefix at any field's start.
3. Excel cell-type tests — verify excel_set_safe forces data_type='s'.
"""
import asyncio
import csv as csv_module
import io
import json
import zipfile
from datetime import datetime, timezone

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.user import User
from app.models.project import Project
from app.models.dataset import Dataset, DatasetColumn, DatasetRow, DatasetValue
from app.models.conversation import Conversation
from app.models.segment import Segment
from app.models.speaker import Speaker
from app.models.code import Code
from app.models.code_application import CodeApplication
from app.models.code_category import CodeCategory
from app.models.memo import Memo
from app.models.note import Note
from app.models.analysis_domain import AnalysisDomain, AnalysisDomainMember
from app.models.metric import MetricDefinition, ComputedResult
from app.routers.export_helpers import csv_safe, excel_set_safe


def _run(coro):
    return asyncio.run(coro)


# Common attack payloads. The first character is what matters — Excel /
# Sheets / LibreOffice key off the leading byte to decide formula vs text.
ATTACK_EQUALS = "=cmd|'/c calc'!A1"
ATTACK_AT = "@SUM(1+1)"
ATTACK_TAB = "\t=evil"
ATTACK_CR = "\r=evil"
LEGITIMATE_NEGATIVE = "-1"  # Common "decline to answer" sentinel; not defanged.


# ── 1. Helper unit tests ─────────────────────────────────────────────────────


class TestCsvSafe:
    def test_equals_prefix_gets_apostrophe(self):
        assert csv_safe(ATTACK_EQUALS) == "'" + ATTACK_EQUALS

    def test_at_prefix_gets_apostrophe(self):
        assert csv_safe(ATTACK_AT) == "'" + ATTACK_AT

    def test_tab_prefix_gets_apostrophe(self):
        assert csv_safe(ATTACK_TAB) == "'" + ATTACK_TAB

    def test_cr_prefix_gets_apostrophe(self):
        assert csv_safe(ATTACK_CR) == "'" + ATTACK_CR

    def test_negative_number_is_NOT_defanged(self):
        # Documented tradeoff: '-1' as 'decline to answer' is common in
        # research data; defanging it would break R numeric-column auto-typing.
        assert csv_safe(LEGITIMATE_NEGATIVE) == LEGITIMATE_NEGATIVE

    def test_plus_sign_NOT_defanged(self):
        # Same tradeoff: '+5' could be legitimate signed numeric data.
        assert csv_safe("+5") == "+5"

    def test_normal_string_unchanged(self):
        assert csv_safe("Wellness Score") == "Wellness Score"

    def test_empty_string_unchanged(self):
        assert csv_safe("") == ""

    def test_numbers_pass_through(self):
        assert csv_safe(42) == 42
        assert csv_safe(3.14) == 3.14

    def test_none_passes_through(self):
        assert csv_safe(None) is None

    def test_bool_passes_through(self):
        assert csv_safe(True) is True
        assert csv_safe(False) is False

    def test_equals_in_middle_is_safe(self):
        # Only the leading char is risky; '=' inside a string is benign.
        assert csv_safe("Score = mean") == "Score = mean"


class TestExcelSetSafe:
    def test_equals_string_forces_string_data_type(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        excel_set_safe(cell, ATTACK_EQUALS)
        assert cell.value == ATTACK_EQUALS
        # Without the helper, openpyxl's _bind_value tags this as 'f'.
        assert cell.data_type == "s"

    def test_normal_string_keeps_default_string_type(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        excel_set_safe(cell, "Wellness Score")
        assert cell.value == "Wellness Score"
        assert cell.data_type == "s"

    def test_at_prefix_xlsx_does_NOT_get_apostrophe(self):
        # In xlsx the cell type is authoritative; '@' strings are still
        # type-'s' by default, so no apostrophe is needed (unlike csv_safe).
        # We verify the value lands literal, not prefixed.
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        excel_set_safe(cell, ATTACK_AT)
        assert cell.value == ATTACK_AT  # not "'@SUM(1+1)"

    def test_numeric_value_unaffected(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        excel_set_safe(cell, 42)
        assert cell.value == 42
        assert cell.data_type == "n"

    def test_round_trip_through_load_workbook(self):
        # Persist + reload to confirm Excel won't see a formula on open.
        wb = Workbook()
        ws = wb.active
        excel_set_safe(ws.cell(row=1, column=1), ATTACK_EQUALS)
        excel_set_safe(ws.cell(row=2, column=1), "normal")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        loaded = load_workbook(buf)
        loaded_cell = loaded.active.cell(row=1, column=1)
        assert loaded_cell.value == ATTACK_EQUALS
        assert loaded_cell.data_type == "s"

    def test_short_equals_string_does_not_trigger_helper(self):
        # openpyxl only auto-formula-tags strings of length > 1, so a single
        # '=' character is already type-'s'. The helper guards on len > 1.
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        excel_set_safe(cell, "=")
        assert cell.value == "="
        # openpyxl tags single-char strings as 's' regardless.
        assert cell.data_type == "s"


# ── 2. Polluted-fixture integration sweep ────────────────────────────────────


def _make_engine():
    engine = create_engine("sqlite:///:memory:")

    @event.listens_for(engine, "connect")
    def set_sqlite_pragma(dbapi_conn, _):
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    return engine


@pytest.fixture(scope="function")
def polluted_session():
    """A project where every user-typed field starts with a formula prefix.

    Lets each endpoint test stream-export and assert no field starts with
    a raw formula prefix (which would inject when Excel/Sheets opens the CSV).
    """
    engine = _make_engine()
    Session = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    s = Session()

    user = User(id=1, username="testuser", password_hash="x", is_admin=True)
    s.add(user)
    project = Project(id=1, name=ATTACK_EQUALS + " Project", user_id=1)
    s.add(project)

    # Conversation + speaker + segment with poisoned text
    conv = Conversation(id=1, project_id=1, name=ATTACK_EQUALS + " Convo")
    s.add(conv)
    speaker = Speaker(id=1, project_id=1, name=ATTACK_AT + " Speaker", is_facilitator=0)
    s.add(speaker)
    seg = Segment(
        id=1, conversation_id=1, speaker_id=1, sequence_order=0,
        text=ATTACK_EQUALS + " segment text",
    )
    s.add(seg)

    # Code + category
    cat = CodeCategory(id=1, project_id=1, name=ATTACK_AT + " Cat", display_order=0)
    s.add(cat)
    code = Code(
        id=1, project_id=1, numeric_id=1, name=ATTACK_EQUALS + " Code",
        description=ATTACK_AT + " desc", is_active=True, category_id=1,
    )
    s.add(code)

    # Apply the code to the segment
    s.flush()
    ca = CodeApplication(segment_id=1, code_id=1, user_id=1)
    s.add(ca)

    # Memo + note
    memo = Memo(
        project_id=1, numeric_id=1,
        content=ATTACK_EQUALS + " memo body",
        entity_type="project", entity_id=1,
    )
    s.add(memo)
    note = Note(conversation_id=1, segment_id=1, sequence_number=1,
                content=ATTACK_AT + " note body")
    s.add(note)

    # Dataset with poisoned column code + text
    ds = Dataset(id=1, project_id=1, name=ATTACK_EQUALS + " Dataset")
    s.add(ds)
    col_q1 = DatasetColumn(
        id=1, dataset_id=1, column_code=ATTACK_EQUALS + "Q1",
        column_text=ATTACK_AT + " question?", column_type="ordinal",
        sequence_order=0, display_order=0,
    )
    col_demo = DatasetColumn(
        id=2, dataset_id=1, column_code="GENDER",
        column_text="Gender", column_type="demographic",
        sequence_order=1, display_order=1,
    )
    s.add_all([col_q1, col_demo])

    row = DatasetRow(id=1, dataset_id=1, row_identifier=ATTACK_EQUALS + "R1")
    s.add(row)
    s.flush()
    s.add_all([
        DatasetValue(id=1, row_id=1, column_id=1, value_text="3", value_numeric=3.0),
        # Demographic free-text starting with a payload — the realistic abuse vector
        DatasetValue(id=2, row_id=1, column_id=2, value_text=ATTACK_EQUALS + " gendervalue"),
    ])

    # Analysis domain + auto-created scale-score metric (Tier-3 path)
    domain = AnalysisDomain(id=1, project_id=1, name=ATTACK_EQUALS + " Domain")
    s.add(domain)
    s.flush()
    s.add(AnalysisDomainMember(
        domain_id=1, member_type="column", member_id=1, sequence_order=0,
    ))
    metric = MetricDefinition(
        id=1, project_id=1,
        name=f"{domain.name} Score",  # The Tier-3 widened risk surface
        metric_type="domain_aggregate",
        config=json.dumps({"child_metric_type": "mean", "child_config": {}, "aggregation": "mean"}),
        input_source_type="dataset_domain",
        input_source_id=1,
        sequence_order=1,
        origin="human",
        origin_context="crosswalk_auto",
        stale=False,
    )
    s.add(metric)
    s.flush()
    s.add(ComputedResult(
        metric_definition_id=1,
        result_data=json.dumps({
            "aggregate_value": 3.0,
            "column_means": {ATTACK_EQUALS + "Q1": 3.0},
        }),
        valid_n=1, total_n=1,
        computed_at=datetime.now(timezone.utc).replace(tzinfo=None),
    ))
    s.commit()
    yield s
    s.close()
    engine.dispose()


# Helper: parse CSV bytes and assert no field starts with a formula prefix.
_BAD_PREFIXES = ("=", "@", "\t", "\r")


def _assert_no_raw_formula_prefix(csv_text: str, *, label: str = "csv"):
    """Parse and assert no field starts with a raw formula prefix.

    Apostrophe-prefixed cells (e.g. "'=cmd|...") are accepted — that's what
    csv_safe produces. Section headers like "Correlation Matrix (...)" and
    literal labels are also fine because they don't start with a prefix.
    """
    reader = csv_module.reader(io.StringIO(csv_text))
    for row_idx, row in enumerate(reader):
        for col_idx, field in enumerate(row):
            if field and field[0] in _BAD_PREFIXES:
                raise AssertionError(
                    f"{label} row {row_idx} col {col_idx} starts with formula "
                    f"prefix {field[0]!r}: {field!r}"
                )


def _stream_to_text(response):
    """Drain a StreamingResponse body iterator into a string."""
    chunks = []
    body = response.body_iterator
    # FastAPI StreamingResponse.body_iterator may be sync or async.
    try:
        for ch in body:
            if isinstance(ch, bytes):
                chunks.append(ch.decode("utf-8"))
            else:
                chunks.append(ch)
    except TypeError:
        # async iterator
        async def _drain():
            async for ch in body:
                chunks.append(ch.decode("utf-8") if isinstance(ch, bytes) else ch)
        asyncio.run(_drain())
    return "".join(chunks)


def _stream_to_bytes(response):
    """Drain a StreamingResponse body iterator into raw bytes (for ZIP/xlsx)."""
    chunks = []
    body = response.body_iterator
    try:
        for ch in body:
            if isinstance(ch, str):
                chunks.append(ch.encode("utf-8"))
            else:
                chunks.append(ch)
    except TypeError:
        async def _drain():
            async for ch in body:
                chunks.append(ch.encode("utf-8") if isinstance(ch, str) else ch)
        asyncio.run(_drain())
    return b"".join(chunks)


def test_export_study_csv_defangs_user_strings(polluted_session):
    from app.routers.export import export_study_csv
    user = polluted_session.get(User, 1)
    response = _run(export_study_csv(project_id=1, user=user, db=polluted_session))
    text = _stream_to_text(response)
    _assert_no_raw_formula_prefix(text, label="study_csv")


def test_export_coded_segments_csv_defangs_user_strings(polluted_session):
    from app.routers.export import export_coded_segments_csv
    user = polluted_session.get(User, 1)
    response = _run(export_coded_segments_csv(
        project_id=1, code_ids=None, exclude_facilitator=False,
        conversation_ids=None, participant_ids=None,
        user=user, db=polluted_session,
    ))
    text = _stream_to_text(response)
    _assert_no_raw_formula_prefix(text, label="coded_segments_csv")


def test_export_code_cooccurrence_csv_defangs_code_names(polluted_session):
    from app.routers.export import export_code_cooccurrence_csv
    user = polluted_session.get(User, 1)
    response = _run(export_code_cooccurrence_csv(
        project_id=1, code_ids=None, exclude_facilitator=False,
        conversation_ids=None, participant_ids=None,
        user=user, db=polluted_session,
    ))
    text = _stream_to_text(response)
    _assert_no_raw_formula_prefix(text, label="code_cooccurrence_csv")


def test_export_study_excel_metric_name_is_string_typed(polluted_session):
    """Excel: metric.name = "=cmd|... Score" must land as data_type='s' so
    Excel renders it as literal text instead of evaluating it as a formula.
    """
    from app.routers.export_excel import export_study_excel
    user = polluted_session.get(User, 1)
    response = _run(export_study_excel(
        project_id=1, user=user, db=polluted_session,
    ))
    body = _stream_to_bytes(response)
    wb = load_workbook(io.BytesIO(body))

    # Walk every cell on every sheet; any cell whose value starts with '='
    # must carry data_type='s', NOT 'f'.
    bad_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and len(cell.value) > 1:
                    if cell.data_type != "s":
                        bad_cells.append((ws.title, cell.coordinate, cell.value, cell.data_type))
    assert not bad_cells, (
        "Cells with '='-prefix value must be type-'s', got: "
        + repr(bad_cells)
    )


def test_export_datasets_excel_defangs_metric_and_column_names(polluted_session):
    """Datasets export carries the Tier-3 widened risk surface: auto metric
    name = f"{domain.name} Score" lands in Metrics Summary and Grouped sheets.
    Verify all formula-prefix string cells are type-'s'.
    """
    from app.routers.export_excel import export_datasets_excel
    user = polluted_session.get(User, 1)
    response = _run(export_datasets_excel(
        project_id=1, user=user, db=polluted_session,
    ))
    body = _stream_to_bytes(response)
    wb = load_workbook(io.BytesIO(body))

    bad_cells = []
    metric_summary_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and len(cell.value) > 1:
                    if cell.data_type != "s":
                        bad_cells.append((ws.title, cell.coordinate, cell.value, cell.data_type))
                    if "Metrics Summary" in ws.title and "Score" in cell.value:
                        metric_summary_cells.append((cell.coordinate, cell.value, cell.data_type))

    assert not bad_cells, (
        "Cells with '='-prefix value must be type-'s' in datasets export: "
        + repr(bad_cells)
    )
    # Sanity check: the Tier-3 auto-metric name DID land in the Metrics
    # Summary sheet (otherwise this test would silently pass even if the
    # surface wasn't covered).
    assert metric_summary_cells, (
        "Expected Tier-3 auto-metric name with '=' prefix in Metrics Summary "
        "but found none — fixture or wiring may have regressed."
    )


def test_export_r_data_csv_defangs_dataset_and_demographic_strings(polluted_session):
    """R-export ZIP contains a data.csv. dataset name + demographic value_text
    are user-supplied; verify they're defanged but numeric value columns
    remain numeric (no leading apostrophe on numbers)."""
    from app.routers.export_r import export_r_data
    user = polluted_session.get(User, 1)
    response = _run(export_r_data(project_id=1, user=user, db=polluted_session))
    body = _stream_to_bytes(response)

    with zipfile.ZipFile(io.BytesIO(body)) as zf:
        csv_name = next(n for n in zf.namelist() if n.endswith("_data.csv"))
        csv_bytes = zf.read(csv_name).decode("utf-8")

    # Strip BOM for parsing
    csv_bytes = csv_bytes.lstrip("﻿")
    _assert_no_raw_formula_prefix(csv_bytes, label="r_data_csv")

    # Verify numeric value column stays parseable as numeric (i.e. not
    # accidentally apostrophe-defaced).
    reader = csv_module.reader(io.StringIO(csv_bytes))
    header = next(reader)
    rows = list(reader)
    assert rows, "expected at least 1 data row"
    # Find the Q1 column (slugified). The polluted column_code starts with
    # '=' which _make_r_identifier strips entirely, so the slug falls back
    # to column_text or the literal "col" (slugified). It's enough to
    # confirm SOME numeric cell in the data row is purely digits.
    has_clean_numeric = any(
        cell.replace(".", "").replace("-", "").isdigit()
        for row in rows
        for cell in row
        if cell  # skip blanks
    )
    assert has_clean_numeric, "expected at least one clean-numeric cell in R data CSV"


def test_export_row_matrix_csv_defangs_metric_label(polluted_session):
    """row-matrix export: header `col.label` derives from metric/column names.
    Verify defang."""
    from app.routers.metrics import get_row_matrix_csv
    user = polluted_session.get(User, 1)
    response = _run(get_row_matrix_csv(
        project_id=1, metric_ids=None, user=user, db=polluted_session,
    ))
    text = _stream_to_text(response)
    _assert_no_raw_formula_prefix(text, label="row_matrix_csv")


def test_export_excerpts_csv_does_not_inject(polluted_session):
    """excerpts/all/csv: no excerpts in the polluted fixture, but the
    endpoint still emits the header row. Verify it doesn't blow up and
    no field starts with a prefix (header is all literals)."""
    from app.routers.excerpts import export_excerpts_csv
    user = polluted_session.get(User, 1)
    response = _run(export_excerpts_csv(
        project_id=1, user=user, db=polluted_session,
    ))
    text = _stream_to_text(response)
    _assert_no_raw_formula_prefix(text, label="excerpts_csv")


# ── 3. Negative-control: confirm helper actually changes behavior ────────────


def test_helper_DOES_make_a_difference(polluted_session):
    """Sanity check: writing a poisoned string to openpyxl WITHOUT the helper
    produces a formula-typed cell. This locks in the underlying behavior so
    the helper-applied tests above can't silently pass against a phantom
    risk."""
    wb = Workbook()
    ws = wb.active
    # No helper — direct assign:
    cell = ws.cell(row=1, column=1, value=ATTACK_EQUALS)
    assert cell.data_type == "f", (
        "openpyxl behavior changed — '=' strings are no longer auto-tagged "
        "as formulas. Re-evaluate whether excel_set_safe is still needed."
    )
