"""The study Excel export spans every source type (#620).

The gap these close is a fail-OPEN omission: sheets titled "Coded Data" and
"Notes" gathered through an INNER join on Conversation, so document segments
(live since documents shipped), observation clips, and every document /
observation / dataset-value note were absent from a workbook that presented
itself as the whole study. Same shape as #616, which fixed the coded-segments
CSV in slab 4c and did not touch the Excel side.

⚠️ Fixture discipline: the project below carries a conversation AND a document
AND an observation, each with its own coded segment, note and quote. A
conversation-only fixture passes under the OLD code and the new — it is exactly
what let this live so long. Every assertion here names a non-conversation row.
"""
import asyncio
import io

import pytest
from openpyxl import load_workbook

from app.models.code import Code
from app.models.code_application import CodeApplication
from app.models.conversation import Conversation
from app.models.document import Document
from app.models.excerpt import Excerpt
from app.models.memo import Memo
from app.models.note import Note
from app.models.observation import Observation
from app.models.project import Project
from app.models.segment import Segment
from app.models.user import User


def _run(coro):
    return asyncio.run(coro)


PID = 860


@pytest.fixture
def three_parent_project(db_session):
    """One project, one coded+noted+quoted unit per source type."""
    db = db_session
    db.add(Project(id=PID, name="Three Parents", user_id=1))
    db.add(Conversation(id=PID, project_id=PID, name="Interview A"))
    db.add(Document(id=PID, project_id=PID, name="Field Notes B",
                    source_filename="field_notes.docx", source_format="docx",
                    summary="A doc summary"))
    db.add(Observation(id=PID, project_id=PID, name="Playground C"))
    db.add(Code(id=PID, project_id=PID, name="Belonging", numeric_id=1,
                is_active=True, is_universal=False))
    db.flush()

    conv_seg = Segment(id=8601, conversation_id=PID, text="conversation turn text",
                       sequence_order=0)
    doc_seg = Segment(id=8602, document_id=PID, text="document paragraph text",
                      sequence_order=0)
    clip = Segment(id=8603, observation_id=PID, text="clip label",
                   start_time=10.0, end_time=20.0, sequence_order=0)
    db.add_all([conv_seg, doc_seg, clip])
    db.flush()

    for sid in (8601, 8602, 8603):
        db.add(CodeApplication(code_id=PID, user_id=1, segment_id=sid))
        db.add(Excerpt(project_id=PID, segment_id=sid))

    db.add(Note(id=8601, conversation_id=PID, segment_id=8601,
                content="a conversation note", sequence_number=1))
    db.add(Note(id=8602, document_id=PID, segment_id=8602,
                content="a document note", sequence_number=2))
    db.add(Note(id=8603, observation_id=PID, segment_id=8603,
                content="an observation note", sequence_number=3))
    db.flush()
    return db


def _sheets(db):
    """Render the workbook and read it back — the export is only correct if
    openpyxl can reopen what it wrote."""
    from app.routers.export_excel import export_study_excel
    from tests.test_export_formula_injection import _stream_to_bytes
    user = db.get(User, 1)
    response = export_study_excel(project_id=PID, user=user, db=db)
    return load_workbook(io.BytesIO(_stream_to_bytes(response)))


def _rows(ws):
    header = [c.value for c in ws[1]]
    return header, [
        {header[i]: c.value for i, c in enumerate(row) if i < len(header)}
        for row in ws.iter_rows(min_row=2)
    ]


class TestCodedDataSheetSpansAllParents:
    def test_every_source_type_has_a_row_and_says_which_it_is(self, three_parent_project):
        wb = _sheets(three_parent_project)
        header, rows = _rows(wb["Coded Data"])

        assert header[:2] == ["Source Type", "Source"]
        by_kind = {r["Source Type"]: r for r in rows}
        assert set(by_kind) == {"conversation", "document", "observation"}
        assert by_kind["document"]["Source"] == "Field Notes B"
        assert by_kind["observation"]["Source"] == "Playground C"
        assert by_kind["document"]["Text"] == "document paragraph text"

    def test_the_code_column_marks_the_non_conversation_rows(self, three_parent_project):
        """The whole point: their CODINGS were missing, not just their names."""
        wb = _sheets(three_parent_project)
        header, rows = _rows(wb["Coded Data"])
        code_col = next(h for h in header if h and h.startswith("1 - "))
        marked = {r["Source Type"] for r in rows if r[code_col] == "X"}
        assert marked == {"conversation", "document", "observation"}

    def test_speaker_columns_stay_blank_rather_than_inventing_a_value(self, three_parent_project):
        wb = _sheets(three_parent_project)
        _, rows = _rows(wb["Coded Data"])
        for r in rows:
            if r["Source Type"] in ("document", "observation"):
                assert (r["Speaker"] or "") == ""
                assert (r["Is Facilitator"] or "") == ""

    def test_the_quoted_flag_now_reaches_non_conversation_units(self, three_parent_project):
        wb = _sheets(three_parent_project)
        _, rows = _rows(wb["Coded Data"])
        assert {r["Source Type"] for r in rows if r["Quoted"] == "Yes"} == {
            "conversation", "document", "observation"
        }


class TestRatingsSheet:
    """#35 — ratings export at the APPLICATION grain, on their own sheet.

    ⚠️ The document segment is rated ZERO on purpose: the sheet's row set is
    "rated applications", and a truthiness slip would drop exactly that row
    while every other assertion here still passed.
    """

    def _rate(self, db):
        code = db.get(Code, PID)
        code.magnitude_min, code.magnitude_max, code.magnitude_step = 0.0, 10.0, 1.0
        code.magnitude_labels = '[{"value": 7.0, "label": "the main cause"}]'
        apps = {a.segment_id: a for a in db.query(CodeApplication).filter(CodeApplication.code_id == PID)}
        apps[8601].magnitude = 7.0   # conversation: rated at an anchor
        apps[8602].magnitude = 0.0   # document: rated ZERO
        # 8603 (the clip): applied, unrated — must be ABSENT.
        db.flush()

    def test_one_row_per_RATED_application_with_the_number_and_its_scale(self, three_parent_project):
        db = three_parent_project
        self._rate(db)
        wb = _sheets(db)
        assert "Ratings" in wb.sheetnames
        # Beside Coded Data, where a reader looks for coding output.
        assert wb.sheetnames.index("Ratings") == wb.sheetnames.index("Coded Data") + 1
        header, rows = _rows(wb["Ratings"])
        assert header[:10] == [
            "Source Type", "Source", "Segment ID", "Sequence", "Speaker",
            "Code", "Coder", "Rating", "Rating Scale", "Rating Anchor",
        ]
        by_kind = {r["Source Type"]: r for r in rows}
        assert set(by_kind) == {"conversation", "document"}, "the unrated clip is absent, never a 0"
        conv, doc = by_kind["conversation"], by_kind["document"]
        assert conv["Rating"] == 7 and conv["Rating Anchor"] == "the main cause"
        assert doc["Rating"] == 0 and doc["Rating"] is not None
        assert doc["Rating Anchor"] in (None, "")
        assert conv["Rating Scale"] == "0 to 10" and conv["Code"] == "Belonging"
        assert conv["Coder"] == "testuser"
        # A number a researcher can average, not the string of one.
        assert isinstance(conv["Rating"], (int, float))

    def test_no_declared_scale_means_no_sheet(self, three_parent_project):
        db = three_parent_project
        # Ratings with no instrument (hand-edited state) do not earn a sheet.
        db.query(CodeApplication).filter(CodeApplication.segment_id == 8601).one().magnitude = 3.0
        db.flush()
        assert "Ratings" not in _sheets(db).sheetnames

    def test_the_coded_data_matrix_is_untouched(self, three_parent_project):
        """The union "X" keeps its meaning; ratings never leak into it."""
        db = three_parent_project
        self._rate(db)
        header, rows = _rows(_sheets(db)["Coded Data"])
        code_col = next(h for h in header if h and h.startswith("1 - "))
        assert {r[code_col] for r in rows} == {"X"}


class TestNotesSheetSpansAllParents:
    def test_document_and_observation_notes_are_present_and_named(self, three_parent_project):
        wb = _sheets(three_parent_project)
        header, rows = _rows(wb["Notes"])

        assert "Source Type" in header and "Source" in header
        by_content = {r["Content"]: r for r in rows}
        assert set(by_content) == {
            "a conversation note", "a document note", "an observation note"
        }
        assert by_content["a document note"]["Source Type"] == "document"
        assert by_content["a document note"]["Source"] == "Field Notes B"
        assert by_content["an observation note"]["Source Type"] == "observation"
        assert by_content["an observation note"]["Source"] == "Playground C"


class TestQuotesSheet:
    def test_quotes_reach_excel_with_their_text_and_ranges(self, three_parent_project):
        """Before #620 an excerpt was a Yes/No flag; its TEXT was nowhere."""
        wb = _sheets(three_parent_project)
        assert "Quotes" in wb.sheetnames
        _, rows = _rows(wb["Quotes"])

        by_kind = {r["Source Type"]: r for r in rows}
        assert set(by_kind) == {"conversation", "document", "observation"}
        assert by_kind["document"]["Quote Text"] == "document paragraph text"
        # A whole-CLIP quote's identity is the clip's range — its label is
        # often blank, so the range is what makes the row usable.
        assert by_kind["observation"]["Start Time"] == "0:10.0"
        assert by_kind["observation"]["End Time"] == "0:20.0"
        assert by_kind["observation"]["Type"] == "whole-segment"


class TestMemoLinkNames:
    def test_a_document_memo_resolves_its_name_instead_of_a_blank(self, three_parent_project):
        """Link Type said "Document" while Link Name was empty (#620)."""
        db = three_parent_project
        db.add(Memo(project_id=PID, entity_type="document", entity_id=PID,
                    numeric_id=1, content="on the doc"))
        db.add(Memo(project_id=PID, entity_type="observation", entity_id=PID,
                    numeric_id=2, content="on the obs"))
        db.flush()

        wb = _sheets(db)
        _, rows = _rows(wb["Memos"])
        by_content = {r["Content"]: r for r in rows}
        assert by_content["on the doc"]["Link Name"] == "Field Notes B"
        assert by_content["on the obs"]["Link Name"] == "Playground C"

    def test_an_unmapped_entity_type_falls_back_rather_than_blanking(self, three_parent_project):
        db = three_parent_project
        db.add(Memo(project_id=PID, entity_type="dataset", entity_id=77,
                    numeric_id=3, content="on a dataset"))
        db.flush()

        wb = _sheets(db)
        _, rows = _rows(wb["Memos"])
        link = next(r["Link Name"] for r in rows if r["Content"] == "on a dataset")
        assert link == "Dataset 77"


class TestSummariesSheet:
    def test_a_document_summary_is_exported(self, three_parent_project):
        wb = _sheets(three_parent_project)
        header, rows = _rows(wb["Summaries"])
        assert header[0] == "Source Type"
        doc_rows = [r for r in rows if r["Source Type"] == "document"]
        assert len(doc_rows) == 1
        assert doc_rows[0]["Source"] == "Field Notes B"
        assert doc_rows[0]["Summary"] == "A doc summary"


class TestCodeSourceMatrixSheet:
    """#629 — the last sheet #620 left conversation-only.

    ⚠️ The shared `PID` is load-bearing here, not laziness: the conversation,
    the document and the observation all have id 860, which is legal because the
    three parents are independent sequences. A matrix keyed by bare source id
    collapses them into one column, and no conversation-only fixture can see it.
    """

    def test_every_source_type_gets_a_column_that_names_its_type(self, three_parent_project):
        wb = _sheets(three_parent_project)
        assert "Code-Source Matrix" in wb.sheetnames, (
            "the sheet was renamed from 'Code-Conversation Matrix' — it no "
            "longer describes only conversations"
        )
        header, rows = _rows(wb["Code-Source Matrix"])

        assert header[0] == "Code"
        assert header[-1] == "Total"
        # One column per source, each carrying its type so two same-named
        # sources of different kinds stay distinguishable.
        assert header[1:-1] == [
            "Interview A (conversation)",
            "Field Notes B (document)",
            "Playground C (observation)",
        ]

        assert len(rows) == 1
        row = rows[0]
        assert row["Code"] == "1 - Belonging"
        assert row["Interview A (conversation)"] == 1
        assert row["Field Notes B (document)"] == 1, "document segments reach the matrix"
        assert row["Playground C (observation)"] == 1, "observation clips reach the matrix"
        assert row["Total"] == 3, (
            "3 distinct coded units across 3 sources — a bare-id key would put "
            "all three in one column and the other two would read blank"
        )

    def test_the_code_column_stays_visible_when_the_axis_is_wide(self, three_parent_project):
        """Freeze panes: the axis got wider, so scrolling right must not strand
        the reader among unlabelled numbers."""
        wb = _sheets(three_parent_project)
        assert wb["Code-Source Matrix"].freeze_panes == "B2"

    def test_the_sheet_survives_a_project_with_no_conversations(self, db_session):
        """The gate was `include_matrix and codes and conversations`, so an
        observation-only project — however much coded material it held — got no
        matrix sheet at all, silently and with nothing saying why. Same shape as
        #626/#627 calling an observation-only project empty.
        """
        db = db_session
        db.add(Project(id=PID, name="Observation only", user_id=1))
        db.add(Observation(id=PID, project_id=PID, name="Playground C"))
        db.add(Code(id=PID, project_id=PID, name="Belonging", numeric_id=1,
                    is_active=True, is_universal=False))
        db.flush()
        db.add(Segment(id=8613, observation_id=PID, text="clip label",
                       start_time=1.0, end_time=2.0, sequence_order=0))
        db.flush()
        db.add(CodeApplication(code_id=PID, user_id=1, segment_id=8613))
        db.flush()

        wb = _sheets(db)
        assert "Code-Source Matrix" in wb.sheetnames, (
            "a project with zero conversations still has coded sources (#629)"
        )
        header, rows = _rows(wb["Code-Source Matrix"])
        assert header[1:-1] == ["Playground C (observation)"]
        assert rows[0]["Playground C (observation)"] == 1


class TestRatingsSheetScope:
    """#869 (h) — the sheet's row set is VISIBLE, HUMAN, rated applications, each
    clause through its chokepoint (`visible_segment_filter`, `non_consensus_filter`).

    Both exclusions were byte-identical to the hand-rolled clauses they replace,
    which is exactly why a test has to pin them: the J2-B rule exists because a
    hand-rolled `origin !=` will not STAY identical.
    """

    def test_a_consensus_row_and_a_hidden_original_never_reach_the_sheet(self, three_parent_project):
        db = three_parent_project
        code = db.get(Code, PID)
        code.magnitude_min, code.magnitude_max, code.magnitude_step = 0.0, 10.0, 1.0
        apps = {a.segment_id: a for a in db.query(CodeApplication).filter(CodeApplication.code_id == PID)}
        apps[8601].magnitude = 7.0
        apps[8602].magnitude = 0.0
        # A merged-away original: UI-unreachable, its rating must not export.
        db.add(Segment(id=8604, conversation_id=PID, text="hidden original",
                       sequence_order=1, merged_into_id=8601))
        db.flush()
        db.add(CodeApplication(code_id=PID, user_id=1, segment_id=8604, magnitude=9.0))
        # The DERIVED consensus row on the clip: its median is computed FROM the
        # coders' ratings, so exporting it beside them counts a judgement twice.
        db.add(CodeApplication(code_id=PID, user_id=None, segment_id=8603,
                               origin="consensus", magnitude=5.0))
        db.flush()

        _, rows = _rows(_sheets(db)["Ratings"])
        ratings = sorted(r["Rating"] for r in rows)
        assert ratings == [0, 7], f"a hidden or consensus rating leaked: {ratings}"
